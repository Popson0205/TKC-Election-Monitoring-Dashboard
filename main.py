import sqlite3
import shutil
import uuid
import os
import json
import logging
import io
import csv
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, Form, Request, Response, Cookie, HTTPException, Depends
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import cloudinary
import cloudinary.uploader
import hashlib, time, secrets
from collections import defaultdict


# --- WHATSAPP ALERT ---
import threading

def send_whatsapp_alert(payload: dict):
    try:
        from twilio.rest import Client
        account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
        auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
        from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")
        recipients_env = os.environ.get("WHATSAPP_RECIPIENTS", "+2349160420100,+2349039587686,+2349072707396,+2348051383900,+2348089377590")
        to_numbers = [f"whatsapp:{n.strip()}" for n in recipients_env.split(",")]
        if not account_sid or not auth_token:
            logger.warning("Twilio credentials not set — WhatsApp alert skipped.")
            return
        import json as _json
        votes = payload.get("votes", {})
        accord_votes = votes.get("ACCORD", 0)
        top_rivals = sorted(
            [(p, v) for p, v in votes.items() if p != "ACCORD" and v > 0],
            key=lambda x: -x[1]
        )[:3]
        rival_str = ", ".join([f"{p}: {v}" for p, v in top_rivals]) or "None"
        msg = (
            f"🗳 *NEW PU SUBMISSION*\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📍 *PU:* {payload.get('location', 'N/A')}\n"
            f"🏛 *Ward:* {payload.get('ward', 'N/A')} | *LGA:* {payload.get('lg', 'N/A')}\n"
            f"🔑 *PU Code:* {payload.get('pu_code', 'N/A')}\n"
            f"👤 *Officer:* {payload.get('officer_id', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"✅ *ACCORD:* {accord_votes}\n"
            f"🔴 *Rivals:* {rival_str}\n"
            f"📊 *Total Cast:* {payload.get('total_cast', 0)} | *Accredited:* {payload.get('total_accredited', 0)}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🕐 {payload.get('timestamp', '')}\n"
            f"_Powered by Popson Geospatial Services_"
        )
        client = Client(account_sid, auth_token)
        for to_number in to_numbers:
            try:
                client.messages.create(from_=f"whatsapp:{from_number}", to=to_number, body=msg)
                logger.info(f"✅ WhatsApp alert sent to {to_number} for PU: {payload.get('pu_code')}")
            except Exception as sms_err:
                logger.error(f"Failed to send to {to_number}: {sms_err}")
    except Exception as e:
        logger.error(f"WhatsApp alert failed: {type(e).__name__}: {e}", exc_info=True)


# --- INCIDENT WHATSAPP ALERT ---
def send_incident_alert(payload: dict):
    try:
        from twilio.rest import Client
        account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
        auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
        from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")
        recipients_env = os.environ.get("WHATSAPP_RECIPIENTS", "+2349160420100,+2349039587686,+2349072707396,+2348051383900,+2348089377590")
        to_numbers = [f"whatsapp:{n.strip()}" for n in recipients_env.split(",")]
        if not account_sid or not auth_token:
            logger.warning("Twilio credentials not set — incident alert skipped.")
            return

        severity = payload.get("severity", "Unknown").upper()
        severity_icon = {"CRITICAL": "🔴", "MEDIUM": "🟡", "LOW": "🟢"}.get(severity, "⚪")

        msg = (
            f"🚨 *INCIDENT REPORT ALERT*\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"{severity_icon} *Severity:* {severity}\n"
            f"⚠️ *Type:* {payload.get('incident_type', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📍 *PU:* {payload.get('location', 'N/A')}\n"
            f"🏛 *Ward:* {payload.get('ward', 'N/A')} | *LGA:* {payload.get('lg', 'N/A')}\n"
            f"🔑 *PU Code:* {payload.get('pu_code', 'N/A')}\n"
            f"👤 *Officer:* {payload.get('officer_id', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"📝 *Description:*\n{payload.get('description', 'N/A')}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"🕐 {payload.get('timestamp', '')}\n"
            f"_Powered by Popson Geospatial Services_"
        )

        client = Client(account_sid, auth_token)
        for to_number in to_numbers:
            try:
                client.messages.create(from_=f"whatsapp:{from_number}", to=to_number, body=msg)
                logger.info(f"✅ Incident alert sent to {to_number}")
            except Exception as sms_err:
                logger.error(f"Failed to send incident alert to {to_number}: {sms_err}")
    except Exception as e:
        logger.error(f"Incident alert failed: {type(e).__name__}: {e}", exc_info=True)

# --- CONFIGURATION ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

# ── CORS — restricted to same origin only ─────────────────────────────────────
from fastapi.middleware.cors import CORSMiddleware
_ALLOWED_ORIGINS = [o.strip() for o in os.environ.get("ALLOWED_ORIGINS", "").split(",") if o.strip()]
if not _ALLOWED_ORIGINS:
    _ALLOWED_ORIGINS = ["*"]   # fallback for local dev; set ALLOWED_ORIGINS in prod
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["Content-Type", "X-Requested-With"],
)
# ─────────────────────────────────────────────────────────────────────────────

# ── Security headers middleware ───────────────────────────────────────────────
from starlette.middleware.base import BaseHTTPMiddleware
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "SAMEORIGIN"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "geolocation=(self), camera=(self)"
        response.headers["Cache-Control"] = "no-store"
        return response
app.add_middleware(SecurityHeadersMiddleware)
# ─────────────────────────────────────────────────────────────────────────────

# ── Dashboard key (set DASHBOARD_KEY env var — keep it secret!) ───────────────
_DASHBOARD_KEY_HASH = hashlib.sha256(
    os.environ.get("DASHBOARD_KEY", "changeme-set-in-env").encode()
).hexdigest()
_SESSION_TOKENS: dict = {}   # token -> expiry timestamp
_SESSION_TTL = 8 * 3600      # 8 hours

def _make_session_token() -> str:
    token = secrets.token_hex(32)
    _SESSION_TOKENS[token] = time.time() + _SESSION_TTL
    return token

def _is_valid_token(token) -> bool:  # accepts str or None
    if not token:
        return False
    expiry = _SESSION_TOKENS.get(token)
    if not expiry:
        return False
    if time.time() > expiry:
        _SESSION_TOKENS.pop(token, None)
        return False
    return True

def _require_dashboard(request: Request):
    token = request.cookies.get("ds_session")
    if not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Dashboard access denied")

def _require_admin(request: Request):
    """Accepts EITHER the admin Bearer key OR a valid ds_session cookie —
    matches the pattern used by the existing /api/admin/* routes, so the
    same admin-portal login (Bearer token) works for Supabase-backed routes too."""
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    bearer_ok = auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )
    if not bearer_ok and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")

# ── OTP store ────────────────────────────────────────────────────────────────
# Structure: { officer_id: { otp, expiry, phone_hint, used, attempts, locked_until } }
_OTP_STORE: dict = {}
_OTP_TTL        = 5 * 60        # 5 minutes
_OTP_MAX_TRIES  = 3             # wrong attempts before lockout
_OTP_LOCKOUT    = 15 * 60       # 15-minute lockout after max attempts
_SUBMIT_TOKENS: dict = {}       # token -> { officer_id, expiry }
_SUBMIT_TOKEN_TTL = 30 * 60     # submission token valid 30 min

def _generate_otp() -> str:
    """Cryptographically random 6-digit OTP."""
    import random as _rnd
    return str(_rnd.SystemRandom().randint(100000, 999999))

def _mask_phone(phone: str) -> str:
    """Return +234***4567 style hint."""
    if len(phone) < 6:
        return "****"
    return phone[:4] + "***" + phone[-4:]

def _clean_phone(p: str) -> str:
    """Normalize Nigerian phone numbers to E.164 format (+234XXXXXXXXXX).
    Strips ALL non-digit chars first, then rebuilds E.164 — handles stray
    plus signs mid-string, spaces, dashes, and any other garbage.
    """
    import re as _re
    digits = _re.sub(r"[^0-9]", "", str(p).strip())
    if not digits:
        return ""
    if digits.startswith("234") and len(digits) == 13:
        return "+" + digits          # 2348012345678  -> +2348012345678
    if digits.startswith("0") and len(digits) == 11:
        return "+234" + digits[1:]   # 08012345678    -> +2348012345678
    if len(digits) == 10:
        return "+234" + digits       # 8012345678     -> +2348012345678
    if digits.startswith("234"):
        return "+" + digits          # any other 234X -> +234X
    return "+" + digits              # best-effort

def _make_submit_token(officer_id: str) -> str:
    """Short-lived signed token that authorises one submission."""
    import hmac as _hmac
    token = secrets.token_hex(24)
    expiry = time.time() + _SUBMIT_TOKEN_TTL
    # Sign: HMAC of token+officer_id with DASHBOARD_KEY_HASH as key material
    sig = _hmac.new(
        _DASHBOARD_KEY_HASH.encode(),
        (token + officer_id).encode(),
        "sha256"
    ).hexdigest()[:16]
    full = f"{token}.{sig}"
    _SUBMIT_TOKENS[full] = {"officer_id": officer_id, "expiry": expiry}
    return full

def _verify_submit_token(token: str, officer_id: str) -> bool:
    """Verify token is valid, not expired, and belongs to this officer."""
    entry = _SUBMIT_TOKENS.get(token)
    if not entry:
        return False
    if time.time() > entry["expiry"]:
        _SUBMIT_TOKENS.pop(token, None)
        return False
    if entry["officer_id"] != officer_id:
        return False
    # Single-use: remove after first successful check
    _SUBMIT_TOKENS.pop(token, None)
    return True
# ─────────────────────────────────────────────────────────────────────────────

# ── Rate limiter for validate_officer (10 attempts / 60s per IP) ──────────────
_rl_store: dict = defaultdict(list)
_RL_MAX   = 10
_RL_WINDOW = 60

def _check_rate_limit(ip: str):
    now = time.time()
    calls = [t for t in _rl_store[ip] if now - t < _RL_WINDOW]
    calls.append(now)
    _rl_store[ip] = calls
    if len(calls) > _RL_MAX:
        raise HTTPException(status_code=429, detail="Too many attempts. Try again in 60 seconds.")

# ── File upload guard ─────────────────────────────────────────────────────────
_MAX_UPLOAD_BYTES = 5 * 1024 * 1024   # 5 MB
_ALLOWED_MIME_PREFIXES = ("image/jpeg", "image/png", "image/webp", "image/gif")
_ALLOWED_MAGIC = {
    b"\xff\xd8\xff": "image/jpeg",
    b"\x89PNG":        "image/png",
    b"RIFF":           "image/webp",
    b"GIF8":           "image/gif",
}

async def _safe_read_image(upload: UploadFile, max_bytes: int = _MAX_UPLOAD_BYTES) -> bytes:
    data = await upload.read(max_bytes + 1)
    if len(data) > max_bytes:
        raise HTTPException(status_code=413, detail="Image too large (max 5MB)")
    magic = data[:4]
    ok = any(magic.startswith(sig) for sig in _ALLOWED_MAGIC)
    if not ok:
        raise HTTPException(status_code=415, detail="Unsupported file type. JPEG/PNG/WebP only.")
    return data
# ─────────────────────────────────────────────────────────────────────────────

# ── Cloudinary Configuration ──────────────────────────────────────────────────
cloudinary.config(
    cloud_name = os.environ.get("CLOUDINARY_CLOUD_NAME", ""),
    api_key    = os.environ.get("CLOUDINARY_API_KEY", ""),
    api_secret = os.environ.get("CLOUDINARY_API_SECRET", "")
)
# ─────────────────────────────────────────────────────────────────────────────

# Render-safe Pathing
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "static", "logos")
os.makedirs(LOGO_PATH, exist_ok=True)
app.mount("/logos", StaticFiles(directory=LOGO_PATH), name="logos")

STATIC_PATH = os.path.join(BASE_DIR, "static")
EC8E_PATH = os.path.join(BASE_DIR, "static", "ec8e")
os.makedirs(EC8E_PATH, exist_ok=True)
os.makedirs(STATIC_PATH, exist_ok=True)
app.mount("/static", StaticFiles(directory=STATIC_PATH), name="static")

# --- DATABASE CONNECTION ---
DB_RELEASE_URL = os.environ.get(
    "DB_RELEASE_URL",
    "https://github.com/Popson0205/Election_Dashboard03/releases/download/v1/election_v3.db"
)
DB_PATH = os.path.join(BASE_DIR, "election_v3.db")

def _ensure_db():
    """Download the SQLite DB from GitHub Releases if not already present."""
    if not os.path.exists(DB_PATH):
        import urllib.request
        logger.info(f"Downloading database from {DB_RELEASE_URL} ...")
        urllib.request.urlretrieve(DB_RELEASE_URL, DB_PATH)
        logger.info(f"Database saved to {DB_PATH}")

_ensure_db()

class _DictRow(sqlite3.Row):
    """Make sqlite3.Row behave like a dict (supports row['key'] and row.get())."""
    def get(self, key, default=None):
        try:
            return self[key]
        except (IndexError, KeyError):
            return default

class _FakeCursor:
    """Wraps sqlite3.Cursor so it supports 'with conn.cursor() as cur:' syntax."""
    def __init__(self, cur):
        self._cur = cur
    def execute(self, sql, params=()):
        self._cur.execute(sql, params)
    def fetchone(self):
        return self._cur.fetchone()
    def fetchall(self):
        return self._cur.fetchall()
    def close(self):
        self._cur.close()
    def __enter__(self):
        return self
    def __exit__(self, *args):
        self._cur.close()

class _FakeConn:
    """Wraps sqlite3 connection so it works as a context manager like psycopg2."""
    def __init__(self, conn):
        self._conn = conn
    def cursor(self):
        return _FakeCursor(self._conn.cursor())
    def commit(self):
        self._conn.commit()
    def close(self):
        self._conn.close()
    def __enter__(self):
        return self
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self._conn.commit()
        self._conn.close()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = _DictRow
    return _FakeConn(conn)

# --- SUPABASE POSTGRES CONNECTION (for field_submissions / incidents) ---
# polling_units stays on SQLite above (static reference data, safe to be
# ephemeral — it's just re-downloaded from GitHub on every boot).
# field_submissions and incidents are USER-GENERATED data that must survive
# redeploys, so they live in Supabase's actual Postgres database instead.
#
# Get this connection string from: Supabase → Settings → Database →
# Connection string → "Transaction pooler" (or "Session pooler"), then set
# it as SUPABASE_DB_URL in Render's environment variables.
import psycopg2
import psycopg2.extras

SUPABASE_DB_URL = os.environ.get("SUPABASE_DB_URL", "")
if not SUPABASE_DB_URL:
    raise RuntimeError(
        "SUPABASE_DB_URL environment variable is not set. "
        "Set it to your Supabase Postgres connection string "
        "(Supabase → Settings → Database → Connection string)."
    )
if SUPABASE_DB_URL.startswith("postgres://"):
    SUPABASE_DB_URL = SUPABASE_DB_URL.replace("postgres://", "postgresql://", 1)


class _PgDictRow(dict):
    """Dict-like row, matching the sqlite3.Row-based _DictRow interface above."""
    def get(self, key, default=None):
        return super().get(key, default)


class _PgFakeCursor:
    """Wraps a psycopg2 cursor so 'with conn.cursor() as cur:' call sites keep working."""
    def __init__(self, cur):
        self._cur = cur
    def execute(self, sql, params=()):
        self._cur.execute(sql, params)
    def fetchone(self):
        row = self._cur.fetchone()
        return _PgDictRow(row) if row is not None else None
    def fetchall(self):
        return [_PgDictRow(r) for r in self._cur.fetchall()]
    def close(self):
        self._cur.close()
    def __enter__(self):
        return self
    def __exit__(self, *args):
        self._cur.close()


class _PgFakeConn:
    """Wraps a psycopg2 connection so 'with get_pg() as conn:' call sites keep working."""
    def __init__(self, conn):
        self._conn = conn
    def cursor(self):
        return _PgFakeCursor(self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor))
    def commit(self):
        self._conn.commit()
    def close(self):
        self._conn.close()
    def __enter__(self):
        return self
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._conn.close()


def get_pg():
    conn = psycopg2.connect(SUPABASE_DB_URL)
    return _PgFakeConn(conn)

# --- DATABASE INITIALIZATION ---
def init_db():
    # ── SQLite: polling_units reference data only (officer_phone migration) ──
    try:
        conn = get_db()
        cur = conn.cursor()
        try:
            cur.execute("ALTER TABLE polling_units ADD COLUMN officer_phone TEXT")
        except Exception:
            pass  # Column already exists — fine
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ SQLite (polling_units) INIT ERROR: {e}")

    # ── Supabase Postgres: field_submissions / incidents / result_audit_log ──
    try:
        conn = get_pg()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS field_submissions (
                id SERIAL PRIMARY KEY,
                officer_id TEXT,
                state TEXT,
                lg TEXT,
                ward TEXT,
                ward_code TEXT,
                pu_code TEXT,
                location TEXT,
                reg_voters INTEGER,
                total_accredited INTEGER,
                valid_votes INTEGER,
                rejected_votes INTEGER,
                total_cast INTEGER,
                lat REAL,
                lon REAL,
                timestamp TEXT,
                votes_json TEXT,
                ec8e_image TEXT,
                reviewed INTEGER DEFAULT 0,
                reviewed_by TEXT,
                reviewed_at TEXT,
                edited_votes_json TEXT,
                edit_note TEXT,
                UNIQUE(pu_code)
            )
        """)

        # Audit log table — records every edit made in the results portal
        cur.execute("""
            CREATE TABLE IF NOT EXISTS result_audit_log (
                id          SERIAL PRIMARY KEY,
                submission_id INTEGER,
                action      TEXT,        -- 'edit' | 'approve' | 'unapprove'
                field       TEXT,        -- field that changed (or 'bulk')
                old_value   TEXT,
                new_value   TEXT,
                changed_by  TEXT,        -- 'admin' for now
                changed_at  TEXT
            )
        """)

        # ── Incidents table ────────────────────────────────────────────────────
        cur.execute("""
            CREATE TABLE IF NOT EXISTS incidents (
                id SERIAL PRIMARY KEY,
                officer_id TEXT,
                pu_code TEXT,
                ward TEXT,
                ward_code TEXT,
                lg TEXT,
                state TEXT,
                location TEXT,
                incident_type TEXT,
                severity TEXT,
                description TEXT,
                evidence_url TEXT,
                lat REAL,
                lon REAL,
                timestamp TEXT,
                status TEXT DEFAULT 'open'
            )
        """)

        conn.commit()
        conn.close()
        print("✅ Supabase Postgres tables (field_submissions/incidents) ready")
    except Exception as e:
        print(f"❌ Supabase Postgres INIT ERROR: {e}")

init_db()

# =============================================================================
# ── SUPABASE OFFICER MANAGEMENT ───────────────────────────────────────────────
# Officers uploaded via /api/upload-officers are stored permanently in Supabase.
# This survives all Render redeployments. Add SUPABASE_URL + SUPABASE_KEY env
# vars on Render, and create the officers table in your Supabase SQL editor:
#
#   CREATE TABLE officers (
#     officer_id   TEXT PRIMARY KEY,
#     lga          TEXT,
#     ward         TEXT,
#     polling_unit TEXT,
#     phone        TEXT
#   );
# =============================================================================

def get_supabase():
    """Return a Supabase client, or None if env vars are not set."""
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_KEY", "").strip()
    if not url or not key:
        logger.warning("SUPABASE_URL / SUPABASE_KEY not set — Supabase disabled.")
        return None
    try:
        from supabase import create_client
        return create_client(url, key)
    except Exception as e:
        logger.error(f"Supabase client error: {e}")
        return None

def _ensure_officers_table():
    """Check the officers table exists at startup and log a clear error if not."""
    sb = get_supabase()
    if not sb:
        return
    try:
        sb.table("officers").select("officer_id").limit(1).execute()
        logger.info("✅ Supabase 'officers' table found.")
    except Exception as e:
        logger.error(
            "❌ Supabase 'officers' table missing. Create it in Supabase SQL editor:\n"
            "  CREATE TABLE officers (\n"
            "    officer_id   TEXT PRIMARY KEY,\n"
            "    lga          TEXT,\n"
            "    ward         TEXT,\n"
            "    polling_unit TEXT,\n"
            "    phone        TEXT\n"
            "  );\n"
            f"Error: {e}"
        )

_ensure_officers_table()

def _get_supabase_officer(officer_id: str, lg: str = "") -> dict | None:
    """Look up one officer in Supabase. Returns row dict or None."""
    sb = get_supabase()
    if not sb:
        return None
    try:
        query = sb.table("officers").select("*").eq("officer_id", officer_id)
        if lg:
            query = query.ilike("lga", lg)
        res = query.limit(1).execute()
        return res.data[0] if res.data else None
    except Exception as e:
        logger.error(f"Supabase officer lookup error: {e}")
        return None


@app.post("/api/upload-officers")
async def upload_officers(request: Request, file: UploadFile = File(...)):
    """
    Admin-only. Upload a CSV of officers → saved permanently to Supabase.
    Required columns: officer_id, lga, ward, polling_unit, phone
    Existing officers are updated (upsert). Safe to re-upload an expanded list.
    """
    _require_admin(request)

    sb = get_supabase()
    if not sb:
        raise HTTPException(
            status_code=503,
            detail="Supabase not configured. Set SUPABASE_URL and SUPABASE_KEY env vars on Render."
        )

    raw = await file.read()
    try:
        text = raw.decode("utf-8-sig")   # strips Excel BOM automatically
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    required = {"officer_id", "lga", "ward", "polling_unit", "phone"}
    if not reader.fieldnames or not required.issubset({c.strip().lower() for c in reader.fieldnames}):
        raise HTTPException(
            status_code=400,
            detail=f"CSV must contain columns: {', '.join(sorted(required))}"
        )

    rows, skipped = [], []
    for i, row in enumerate(reader, start=2):
        r     = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
        oid   = r.get("officer_id", "")
        phone = _clean_phone(r.get("phone", ""))
        if not oid:
            skipped.append(f"Row {i}: missing officer_id")
            continue
        rows.append({
            "officer_id":   oid,
            "lga":          r.get("lga", ""),
            "ward":         r.get("ward", ""),
            "polling_unit": r.get("polling_unit", ""),
            "phone":        phone,
        })

    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows found in CSV.")

    total_saved = 0
    for start in range(0, len(rows), 500):
        batch = rows[start:start + 500]
        try:
            sb.table("officers").upsert(batch, on_conflict="officer_id").execute()
            total_saved += len(batch)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Database error: {e}")

    logger.info(f"✅ {total_saved} officers upserted to Supabase.")
    return {
        "status":          "success",
        "saved":           total_saved,
        "skipped":         len(skipped),
        "skipped_details": skipped[:20],
        "message":         f"{total_saved} officers saved permanently to Supabase.",
    }


@app.get("/api/officers")
async def list_officers(request: Request, lga: str = "", search: str = "", page: int = 1, filter: str = "all", page_size: int = 50):
    """Admin-only: list officers stored in Supabase, paginated and filterable."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        q = sb.table("officers").select("*").order("officer_id")
        if lga:
            q = q.ilike("lga", f"%{lga}%")
        if search:
            q = q.ilike("officer_id", f"%{search}%")
        res = q.execute()
        rows = res.data or []
        if filter == "registered":
            rows = [r for r in rows if (r.get("phone") or "").strip()]
        elif filter == "unregistered":
            rows = [r for r in rows if not (r.get("phone") or "").strip()]
        total = len(rows)
        pages = max(1, (total + page_size - 1) // page_size)
        page = max(1, min(page, pages))
        start = (page - 1) * page_size
        page_rows = rows[start:start + page_size]
        return {"status": "success", "officers": page_rows, "total": total, "page": page, "pages": pages, "count": total}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/officers/stats")
async def officer_stats(request: Request):
    """Admin-only: registration counts for the dashboard stat cards (Supabase-backed)."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        res = sb.table("officers").select("phone").execute()
        rows = res.data or []
        total = len(rows)
        with_phone = sum(1 for r in rows if (r.get("phone") or "").strip())
        return {"total": total, "with_phone": with_phone, "without_phone": total - with_phone}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/officers/update")
async def update_officer(request: Request):
    """Admin-only: update a single officer's phone number in Supabase."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        body  = await request.json()
        oid   = str(body.get("officer_id", "")).strip()
        if not oid:
            raise HTTPException(status_code=400, detail="officer_id required.")
        phone = _clean_phone(str(body.get("phone", "")))
        sb.table("officers").update({"phone": phone}).eq("officer_id", oid).execute()
        return {"status": "success", "message": f"Officer {oid} updated."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/officers/delete")
async def delete_officer(request: Request):
    """Admin-only: remove a single officer by officer_id."""
    _require_admin(request)
    sb = get_supabase()
    if not sb:
        raise HTTPException(status_code=503, detail="Supabase not configured.")
    try:
        body = await request.json()
        oid  = str(body.get("officer_id", "")).strip()
        if not oid:
            raise HTTPException(status_code=400, detail="officer_id required.")
        sb.table("officers").delete().eq("officer_id", oid).execute()
        return {"status": "success", "message": f"Officer {oid} deleted."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# END SUPABASE BLOCK
# =============================================================================

# --- API ENDPOINTS ---
@app.get("/api/validate_officer/{officer_id}")
def validate_officer(officer_id: str, request: Request, lg: str = ""):
    _check_rate_limit(request.client.host)
    officer_id = officer_id[:60]
    import re as _re
    officer_id = _re.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)
    lg = lg.strip()[:60]
    if not lg:
        return {"valid": False, "message": "Please select your LGA first."}

    # ── 1. Check Supabase officers table first ────────────────────────────────
    sb_officer = _get_supabase_officer(officer_id, lg)
    if sb_officer:
        parts = officer_id.split("-", 1)
        pu_data = {}
        if len(parts) == 2:
            ward_code, pu_code = parts[0].strip(), parts[1].strip()
            try:
                with get_db() as conn:
                    with conn.cursor() as cur:
                        cur.execute(
                            """SELECT ward, lg, location, pu_code, ward_code, state
                               FROM polling_units
                               WHERE ward_code = ? AND pu_code = ?
                               AND LOWER(state) = 'osun' AND LOWER(lg) = LOWER(?)
                               LIMIT 1""",
                            (ward_code, pu_code, lg)
                        )
                        row = cur.fetchone()
                        if row:
                            pu_data = {
                                "state":     row["state"] or "osun",
                                "ward":      row["ward"],
                                "lg":        row["lg"],
                                "location":  row["location"],
                                "pu_code":   row["pu_code"],
                                "ward_code": row["ward_code"],
                            }
            except Exception:
                pass
        if not pu_data:
            pu_data = {
                "state":     "osun",
                "ward":      sb_officer.get("ward", ""),
                "lg":        sb_officer.get("lga", lg),
                "location":  sb_officer.get("polling_unit", ""),
                "pu_code":   officer_id.split("-", 1)[1] if "-" in officer_id else officer_id,
                "ward_code": officer_id.split("-", 1)[0] if "-" in officer_id else "",
            }
        return {
            "valid":   True,
            "message": f"Access Granted: {pu_data.get('location', sb_officer.get('polling_unit', ''))}",
            **pu_data,
        }

    # ── 2. Fallback: check polling_units SQLite table ─────────────────────────
    try:
        parts = officer_id.split("-", 1)
        if len(parts) != 2:
            return {"valid": False, "message": "Invalid ID format. Expected: WARDCODE-PUCODE"}
        ward_code, pu_code = parts[0].strip(), parts[1].strip()
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT ward, lg, location, pu_code, ward_code, state
                       FROM polling_units
                       WHERE ward_code = ? AND pu_code = ?
                       AND LOWER(state) = 'osun'
                       AND LOWER(lg) = LOWER(?)
                       LIMIT 1""",
                    (ward_code, pu_code, lg)
                )
                row = cur.fetchone()
                if row:
                    return {
                        "valid":     True,
                        "message":   f"Access Granted: {row['location']}",
                        "state":     row["state"] or "osun",
                        "ward":      row["ward"],
                        "lg":        row["lg"],
                        "location":  row["location"],
                        "pu_code":   row["pu_code"],
                        "ward_code": row["ward_code"],
                    }
                else:
                    return {"valid": False, "message": "Officer ID not found in selected LGA. Check your LGA and ID."}
    except Exception as e:
        return {"valid": False, "message": f"Validation error: {str(e)}"}


@app.get("/api/states")
def get_states():
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT state FROM polling_units WHERE state = 'osun' ORDER BY state")
            rows = cur.fetchall()
            return [r["state"] for r in rows]

@app.get("/api/lgas/{state}")
def get_lgas(state: str):
    with get_db() as conn:
        with conn.cursor() as cur:
            # BUG FIX #6: Use the actual state param (lowercased) instead of hardcoded 'osun'
            cur.execute("SELECT DISTINCT lg FROM polling_units WHERE LOWER(state) = LOWER(?) ORDER BY lg", (state,))
            rows = cur.fetchall()
            return [r["lg"] for r in rows]

@app.get("/api/wards/{state}/{lg}")
def get_wards(state: str, lg: str):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT ward, ward_code FROM polling_units WHERE LOWER(state) = LOWER(?) AND lg = ? ORDER BY ward", (state, lg))
            rows = cur.fetchall()
            return [{"name": r["ward"], "code": r["ward_code"]} for r in rows]

@app.get("/api/pus/{state}/{lg}/{ward}")
def get_pus(state: str, lg: str, ward: str):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT location, pu_code FROM polling_units WHERE LOWER(state) = LOWER(?) AND lg = ? AND ward = ?", (state, lg, ward))
            rows = cur.fetchall()
            return [{"location": r["location"], "pu_code": r["pu_code"]} for r in rows]

@app.post("/submit")
async def submit(
    data: str = Form(...),
    ec8e_image: UploadFile = File(None)
):
    try:
        payload = json.loads(data)
        # Verify submission token — proves officer completed OTP auth
        submit_token = payload.get("submit_token", "")
        officer_id   = payload.get("officer_id", "")
        if not _verify_submit_token(submit_token, officer_id):
            return {"status": "error", "message": "Session expired or invalid. Please log in again."}
        votes_json = json.dumps(payload.get("votes", {}))
        ec8e_filename = None
        if ec8e_image and ec8e_image.filename:
            safe_pu = str(payload.get("pu_code", "unk")).replace("/", "_").replace(" ", "_")
            public_id = f"ec8e_forms/{safe_pu}_{uuid.uuid4().hex[:8]}"
            try:
                img_bytes = await _safe_read_image(ec8e_image)
                upload_result = cloudinary.uploader.upload(
                    img_bytes,
                    public_id=public_id,
                    resource_type="image",
                    overwrite=True
                )
                ec8e_filename = upload_result["secure_url"]
                logger.info(f"EC8E uploaded to Cloudinary: {ec8e_filename}")
            except Exception as cloud_err:
                logger.error(f"Cloudinary upload failed: {cloud_err} — saving locally")
                ext = os.path.splitext(ec8e_image.filename)[1].lower()
                local_name = f"{safe_pu}_{uuid.uuid4().hex[:8]}{ext}"
                with open(os.path.join(EC8E_PATH, local_name), "wb") as img_f:
                    shutil.copyfileobj(ec8e_image.file, img_f)
                ec8e_filename = f"/ec8e/{local_name}"
        with get_pg() as conn:
            with conn.cursor() as cur:
                cur.execute("""INSERT INTO field_submissions (
                    officer_id, state, lg, ward, ward_code, pu_code, location,
                    reg_voters, total_accredited, valid_votes, rejected_votes, total_cast,
                    lat, lon, timestamp, votes_json, ec8e_image
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
                    payload.get("officer_id"),
                    (payload.get("state") or "osun").lower(),  # always store lowercase
                    payload.get("lg"),
                    payload.get("ward"), payload.get("ward_code"), payload.get("pu_code"),
                    payload.get("location"), payload.get("reg_voters"), payload.get("total_accredited"),
                    payload.get("valid_votes"), payload.get("rejected_votes"), payload.get("total_cast"),
                    payload.get("lat"), payload.get("lon"),
                    datetime.now().isoformat(), votes_json, ec8e_filename
                ))
                conn.commit()
        alert_payload = {**payload, "timestamp": datetime.now().strftime("%d %b %Y %H:%M")}
        threading.Thread(target=send_whatsapp_alert, args=(alert_payload,)).start()
        return {"status": "success", "message": "Result Uploaded Successfully"}
    except psycopg2.IntegrityError:
        return {"status": "error", "message": "REJECTED: A submission for this Polling Unit already exists."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/ai_interpret")
async def ai_interpret(data: dict):
    STATS = {
        "lgas": 30,
        "wards": 332,
        "pus": 3763,
        "urban_hubs": ["OSOGBO", "OLORUNDA", "ILESA EAST", "IFE CENTRAL", "IWO"]
    }
    OSUN_PARTIES_AI = ["ACCORD", "AA", "AAC", "ADC", "ADP", "APGA", "APC", "APM", "APP", "BP", "NNPP", "PRP", "YPP", "ZLP"]
    party_votes = {p: data.get(p, 0) for p in OSUN_PARTIES_AI}
    acc = party_votes["ACCORD"]
    rivals = {p: v for p, v in party_votes.items() if p != "ACCORD"}

    ta = data.get('total_accredited', 0)
    rv = data.get('reg_voters', 0)
    current_lg = str(data.get('lg', "")).upper()

    total_votes = sum(party_votes.values())
    if total_votes == 0:
        return {"analysis": "SYSTEM READY: Awaiting live feed from 3,763 Polling Units across Osun State."}

    share = (acc / total_votes) * 100
    top_rival = max(rivals, key=rivals.get)
    margin = acc - rivals[top_rival]
    turnout = (ta / rv * 100) if rv > 0 else 0

    is_urban = current_lg in STATS["urban_hubs"]

    if share > 55:
        trend = "LANDSLIDE"
    elif share > 40:
        trend = "STRONG LEAD"
    else:
        trend = "BATTLEGROUND"

    location_tag = " [URBAN HUB]" if is_urban else " [RURAL SECTOR]"

    analysis = (
        f"OSUN STATISTICAL AUDIT ({current_lg}{location_tag}): "
        f"Accord is in a {trend} position with {share:.1f}% of the current tally. "
        f"Lead Margin over {top_rival}: **{margin:+,}** votes. "
    )

    if turnout > 0:
        analysis += f"Voter Productivity is at **{turnout:.1f}%**. "
        if turnout > 65:
            analysis += "⚠️ ALERT: Unusually high turnout detected; verify PU logs. "

    if is_urban and share < 45:
        analysis += "STRATEGY: Increase urban mobilization; Osogbo/Iwo volume is critical."
    elif not is_urban and share > 50:
        analysis += "STRATEGY: Rural stronghold confirmed. Protect the lead during collation."

    return {
        "analysis": analysis,
        "is_alert": turnout > 70 or margin < 100,
        "stats": {
            "turnout_gap": 100 - turnout,
            "osun_progress": f"Active in {STATS['lgas']} LGAs"
        }
    }

@app.get("/api/dashboard_filters")
def get_dash_filters(request: Request):
    _require_dashboard(request)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""SELECT DISTINCT
                               LOWER(state) as state,
                               lg,
                               ward
                           FROM polling_units
                           WHERE LOWER(state) = 'osun'
                           ORDER BY lg, ward""")
            return [dict(r) for r in cur.fetchall()]

@app.get("/export/csv")
async def export_csv(request: Request):
    _require_dashboard(request)
    try:
        import openpyxl
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=500, content={"error": "openpyxl not installed. Add openpyxl to requirements.txt and redeploy."})
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import Response as _Resp
    import io as _io
    from datetime import datetime as _dt

    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM field_submissions ORDER BY timestamp DESC")
            rows = cur.fetchall()

    PARTIES = ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Election Results"

    GREEN="00008751"; GOLD="00FFC107"; DARK="00121212"; WHITE="00FFFFFF"
    LIGHT_GREY="00F5F5F5"; ACCORD_LIGHT="00E8F5E9"
    hdr_fill=PatternFill("solid",fgColor=GREEN)
    hdr_font=Font(bold=True,color=WHITE,size=10,name="Calibri")
    party_fill=PatternFill("solid",fgColor=GOLD)
    party_font=Font(bold=True,color=DARK,size=10,name="Calibri")
    accord_fill=PatternFill("solid",fgColor=ACCORD_LIGHT)
    alt_fill=PatternFill("solid",fgColor=LIGHT_GREY)
    center=Alignment(horizontal="center",vertical="center",wrap_text=True)
    left_align=Alignment(horizontal="left",vertical="center")
    ts=Side(style="thin",color="CCCCCC")
    tb=Border(left=ts,right=ts,top=ts,bottom=ts)

    TOTAL_COLS=17+len(PARTIES)+1
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=TOTAL_COLS)
    tc=ws.cell(row=1,column=1,value="ACCORD PARTY - OSUN 2026 GOVERNORSHIP ELECTION RESULTS")
    tc.fill=PatternFill("solid",fgColor=GREEN); tc.font=Font(bold=True,color=GOLD,size=14,name="Calibri")
    tc.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=30

    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=TOTAL_COLS)
    sc=ws.cell(row=2,column=1,value=f"Exported: {_dt.now().strftime('%d %B %Y  %H:%M')}  |  Total PUs: {len(rows)}")
    sc.fill=PatternFill("solid",fgColor=GOLD); sc.font=Font(bold=True,color=DARK,size=10,name="Calibri")
    sc.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[2].height=18

    headers=["#","Officer ID","State","LGA","Ward","Ward Code","PU Code","Polling Unit",
             "Reg. Voters","Accredited","Valid Votes","Rejected","Total Cast",
             "Latitude","Longitude","Timestamp","EC8E",*PARTIES,"ACCORD TOTAL"]
    for col,h in enumerate(headers,1):
        cell=ws.cell(row=3,column=col,value=h)
        ip=h in PARTIES or h=="ACCORD TOTAL"
        cell.fill=party_fill if ip else hdr_fill
        cell.font=party_font if ip else hdr_font
        cell.alignment=center; cell.border=tb
    ws.row_dimensions[3].height=22

    for ri,r in enumerate(rows,4):
        v=json.loads(r["votes_json"]) if isinstance(r["votes_json"],str) else (r["votes_json"] or {})
        av=v.get("ACCORD",0); is_alt=(ri%2==0)
        rf=accord_fill if av>0 else (alt_fill if is_alt else None)
        row_data=[r["id"],r["officer_id"],(r["state"] or "").upper(),(r["lg"] or "").upper(),
                  (r["ward"] or "").upper(),r["ward_code"],r["pu_code"],r["location"],
                  r["reg_voters"],r["total_accredited"],r["valid_votes"],r["rejected_votes"],
                  r["total_cast"],r["lat"],r["lon"],r["timestamp"],
                  "YES" if r.get("ec8e_image") else "NO",
                  *[v.get(p,0) for p in PARTIES],av]
        for col,val in enumerate(row_data,1):
            cell=ws.cell(row=ri,column=col,value=val)
            cell.border=tb; cell.alignment=left_align if col==8 else center
            if rf: cell.fill=rf
            if col==len(headers): cell.font=Font(bold=True,color="00008751",size=10,name="Calibri")
        ws.row_dimensions[ri].height=16

    col_widths=[4,12,8,14,16,11,10,28,10,11,10,9,10,10,10,22,6]+[8]*len(PARTIES)+[12]
    for i,w in enumerate(col_widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A4"
    ws.auto_filter.ref=f"A3:{get_column_letter(TOTAL_COLS)}3"

    ws2=wb.create_sheet("Party Summary")
    ws2.merge_cells("A1:C1")
    s2t=ws2.cell(row=1,column=1,value="PARTY VOTE SUMMARY")
    s2t.fill=PatternFill("solid",fgColor=GREEN); s2t.font=Font(bold=True,color=GOLD,size=12,name="Calibri")
    s2t.alignment=Alignment(horizontal="center",vertical="center"); ws2.row_dimensions[1].height=24
    for col,h in enumerate(["Party","Total Votes","% Share"],1):
        c=ws2.cell(row=2,column=col,value=h); c.fill=hdr_fill; c.font=hdr_font; c.alignment=center; c.border=tb
    pt={p:sum((json.loads(r["votes_json"]) if isinstance(r["votes_json"],str) else (r["votes_json"] or {})).get(p,0) for r in rows) for p in PARTIES}
    gt=sum(pt.values()) or 1
    for ri2,(party,total) in enumerate(sorted(pt.items(),key=lambda x:-x[1]),3):
        pct=round((total/gt)*100,2); ia=party=="ACCORD"
        for col,val in enumerate([party,total,f"{pct}%"],1):
            c=ws2.cell(row=ri2,column=col,value=val); c.border=tb; c.alignment=center
            if ia: c.fill=PatternFill("solid",fgColor=ACCORD_LIGHT); c.font=Font(bold=True,color="00008751",size=10,name="Calibri")
    ws2.column_dimensions["A"].width=12; ws2.column_dimensions["B"].width=14; ws2.column_dimensions["C"].width=10

    buf=_io.BytesIO(); wb.save(buf); buf.seek(0)
    return _Resp(content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=Accord_Osun2026_Results.xlsx"})

# ── Dashboard authentication endpoint ────────────────────────────────────────
@app.post("/api/verify-dashboard")
async def verify_dashboard(request: Request, response: Response):
    try:
        body = await request.json()
        key = str(body.get("key", ""))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid request")
    given_hash = hashlib.sha256(key.encode()).hexdigest()
    if not secrets.compare_digest(given_hash, _DASHBOARD_KEY_HASH):
        raise HTTPException(status_code=401, detail="Invalid dashboard key")
    token = _make_session_token()
    # secure flag: True only when request came in over HTTPS
    # (Render/cloud proxies forward as HTTP internally — check X-Forwarded-Proto)
    is_https = request.headers.get("x-forwarded-proto", "http") == "https"
    response.set_cookie(
        key="ds_session",
        value=token,
        httponly=True,
        samesite="lax",
        secure=is_https,
        max_age=_SESSION_TTL,
        path="/"
    )
    return {"status": "ok"}

@app.post("/api/logout-dashboard")
async def logout_dashboard(request: Request, response: Response):
    token = request.cookies.get("ds_session")
    if token:
        _SESSION_TOKENS.pop(token, None)
    response.delete_cookie("ds_session", path="/")
    return {"status": "ok"}
# ─────────────────────────────────────────────────────────────────────────────

# ── OTP endpoints ────────────────────────────────────────────────────────────
@app.post("/api/request-otp")
async def request_otp(request: Request):
    """
    Step 1 of officer auth.
    Body: { "officer_id": "WARDCODE-PUCODE", "lg": "Osogbo" }
    Looks up phone from Supabase first, then falls back to polling_units SQLite.
    Returns: { "status": "sent", "phone_hint": "+234***4567" }
             or { "status": "no_phone", "message": "..." }
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    lg         = str(body.get("lg", "")).strip()[:60]
    import re as _re_otp
    officer_id = _re_otp.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)

    if not lg:
        raise HTTPException(status_code=400, detail="LGA is required.")

    parts = officer_id.split("-", 1)
    if len(parts) != 2:
        raise HTTPException(status_code=400, detail="Invalid officer ID format")
    ward_code, pu_code = parts[0].strip(), parts[1].strip()

    # Check lockout — key includes LGA to prevent cross-LGA lockout bleed
    otp_key      = f"{officer_id}|{lg.lower()}"
    entry        = _OTP_STORE.get(otp_key, {})
    locked_until = entry.get("locked_until", 0)
    if time.time() < locked_until:
        remaining = int(locked_until - time.time())
        raise HTTPException(
            status_code=429,
            detail=f"Too many failed attempts. Try again in {remaining // 60}m {remaining % 60}s."
        )

    # ── Resolve phone + PU data: Supabase first, SQLite fallback ─────────────
    phone   = None
    pu_data = {}

    sb_officer = _get_supabase_officer(officer_id, lg)
    if sb_officer:
        phone = _clean_phone(sb_officer.get("phone", "") or "")
        # Enrich pu_data from SQLite polling_units if available
        try:
            with get_db() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT ward, lg, location, pu_code, ward_code, state
                           FROM polling_units
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun' AND LOWER(lg) = LOWER(?)
                           LIMIT 1""",
                        (ward_code, pu_code, lg)
                    )
                    row = cur.fetchone()
                    if row:
                        pu_data = {
                            "state":     row["state"] or "osun",
                            "ward":      row["ward"],
                            "lg":        row["lg"],
                            "location":  row["location"],
                            "pu_code":   row["pu_code"],
                            "ward_code": row["ward_code"],
                        }
        except Exception:
            pass
        if not pu_data:
            pu_data = {
                "state":     "osun",
                "ward":      sb_officer.get("ward", ""),
                "lg":        sb_officer.get("lga", lg),
                "location":  sb_officer.get("polling_unit", ""),
                "pu_code":   pu_code,
                "ward_code": ward_code,
            }
    else:
        # Fallback: check polling_units SQLite table
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT ward, lg, location, pu_code, ward_code, state, officer_phone
                       FROM polling_units
                       WHERE ward_code = ? AND pu_code = ?
                       AND LOWER(state) = 'osun'
                       AND LOWER(lg) = LOWER(?)
                       LIMIT 1""",
                    (ward_code, pu_code, lg)
                )
                row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Officer ID not found. Access Denied.")

        phone = _clean_phone(row["officer_phone"] or "") if row["officer_phone"] else None
        pu_data = {
            "state":     row["state"] or "osun",
            "ward":      row["ward"],
            "lg":        row["lg"],
            "location":  row["location"],
            "pu_code":   row["pu_code"],
            "ward_code": row["ward_code"],
        }

    if not phone or len(phone) < 10:
        return {
            "status":  "no_phone",
            "message": "No phone number registered for this officer ID. Contact your supervisor."
        }

    # Generate OTP and store
    otp = _generate_otp()
    _OTP_STORE[otp_key] = {
        "otp":          otp,
        "expiry":       time.time() + _OTP_TTL,
        "phone_hint":   _mask_phone(phone),
        "phone":        phone,
        "used":         False,
        "attempts":     0,
        "locked_until": 0,
        "pu_data":      pu_data,
    }

    # Send via Twilio WhatsApp
    account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
    auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
    from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")

    if not account_sid or not auth_token:
        dev_mode = os.environ.get("OTP_DEV_MODE", "").lower() in ("1", "true", "yes")
        if dev_mode:
            logger.warning("Twilio not configured — OTP not sent (dev mode)")
            logger.info(f"[DEV OTP] {officer_id}: {otp}")
        else:
            logger.error("Twilio credentials missing — cannot send OTP")
            raise HTTPException(
                status_code=503,
                detail="OTP service is not configured. Contact your administrator."
            )
    else:
        try:
            from twilio.rest import Client as _TC
            client = _TC(account_sid, auth_token)
            msg = (
                f"🗳 *TKC FIELD COLLATION*\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"Your One-Time Password:\n\n"
                f"*{otp}*\n\n"
                f"Valid for *5 minutes*. Do NOT share this code.\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"_Powered by Popson Geospatial Services_"
            )
            client.messages.create(
                from_=f"whatsapp:{from_number}",
                to=f"whatsapp:{phone}",
                body=msg
            )
            logger.info(f"✅ OTP sent to {_mask_phone(phone)} for officer {officer_id}")
        except Exception as e:
            twilio_msg = str(e)
            logger.error(f"Twilio OTP send failed for {officer_id}: {twilio_msg}", exc_info=True)
            raise HTTPException(
                status_code=500,
                detail=f"Failed to send OTP: {twilio_msg}"
            )

    return {
        "status":     "sent",
        "phone_hint": _mask_phone(phone)
    }


@app.post("/api/verify-otp")
async def verify_otp(request: Request):
    """
    Step 2 of officer auth.
    Body: { "officer_id": "WARDCODE-PUCODE", "otp": "123456" }
    Returns: { "status": "ok", "token": "...", "pu_data": {...} }
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    lg         = str(body.get("lg", "")).strip()[:60]
    import re as _re3
    officer_id = _re3.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)
    given_otp  = str(body.get("otp", "")).strip()[:6]

    otp_key = f"{officer_id}|{lg.lower()}"
    entry = _OTP_STORE.get(otp_key)

    # No OTP requested
    if not entry:
        raise HTTPException(status_code=400, detail="No OTP requested for this officer. Start again.")

    # Lockout check
    if time.time() < entry.get("locked_until", 0):
        remaining = int(entry["locked_until"] - time.time())
        raise HTTPException(
            status_code=429,
            detail=f"Account locked. Try again in {remaining // 60}m {remaining % 60}s."
        )

    # Already used
    if entry.get("used"):
        raise HTTPException(status_code=400, detail="OTP already used. Request a new one.")

    # Expired
    if time.time() > entry["expiry"]:
        _OTP_STORE.pop(otp_key, None)
        raise HTTPException(status_code=400, detail="OTP expired. Request a new one.")

    # Wrong OTP — increment attempts
    if not secrets.compare_digest(given_otp, entry["otp"]):
        entry["attempts"] = entry.get("attempts", 0) + 1
        remaining_tries = _OTP_MAX_TRIES - entry["attempts"]
        if entry["attempts"] >= _OTP_MAX_TRIES:
            entry["locked_until"] = time.time() + _OTP_LOCKOUT
            entry["used"] = True  # invalidate
            raise HTTPException(
                status_code=429,
                detail=f"Too many wrong attempts. Account locked for {_OTP_LOCKOUT // 60} minutes."
            )
        raise HTTPException(
            status_code=401,
            detail=f"Incorrect OTP. {remaining_tries} attempt{'s' if remaining_tries != 1 else ''} remaining."
        )

    # ✅ Correct OTP — mark used, issue submission token
    entry["used"] = True
    token = _make_submit_token(officer_id)
    pu_data = entry["pu_data"]

    return {
        "status": "ok",
        "token": token,
        "officer_id": officer_id,
        "state":     pu_data["state"],
        "ward":      pu_data["ward"],
        "lg":        pu_data["lg"],
        "location":  pu_data["location"],
        "pu_code":   pu_data["pu_code"],
        "ward_code": pu_data["ward_code"],
    }
# ─────────────────────────────────────────────────────────────────────────────

# ── Admin: officer stats ─────────────────────────────────────────────────────
@app.get("/api/admin/officer-stats")
async def officer_stats(request: Request):
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing Authorization")
    given_hash = hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest()
    if not secrets.compare_digest(given_hash, _DASHBOARD_KEY_HASH):
        raise HTTPException(status_code=403, detail="Invalid key")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as total FROM polling_units WHERE LOWER(state)='osun'")
            total = cur.fetchone()["total"]
            cur.execute("SELECT COUNT(*) as c FROM polling_units WHERE LOWER(state)='osun' AND officer_phone IS NOT NULL AND officer_phone != ''")
            with_phone = cur.fetchone()["c"]
    return {"total": total, "with_phone": with_phone, "without_phone": total - with_phone}
# ─────────────────────────────────────────────────────────────────────────────

# ── Admin: set officer phone numbers ─────────────────────────────────────────
@app.post("/api/admin/set-officer-phone")
async def set_officer_phone(request: Request):
    """
    Protected by DASHBOARD_KEY (same secret as dashboard).
    Accepts JSON:
      Single:  {"officer_id": "WARDCODE-PUCODE", "phone": "+2348012345678"}
      Bulk:    {"officers": [{"officer_id": "...", "phone": "..."}, ...]}
    """
    # Verify admin key from Authorization header
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    given = auth.split(" ", 1)[1].strip()
    given_hash = hashlib.sha256(given.encode()).hexdigest()
    if not secrets.compare_digest(given_hash, _DASHBOARD_KEY_HASH):
        raise HTTPException(status_code=403, detail="Invalid admin key")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    records = body.get("officers") or ([body] if "officer_id" in body else [])
    if not records:
        raise HTTPException(status_code=400, detail="Provide officer_id+phone or officers array")

    updated, skipped, no_match = 0, 0, []
    with get_db() as conn:
        with conn.cursor() as cur:
            for rec in records:
                oid   = str(rec.get("officer_id", "")).strip()[:60]
                phone = _clean_phone(rec.get("phone", ""))
                lga   = str(rec.get("lga", "")).strip()[:80]

                # Skip rows with no phone (blank = not yet assigned, not an error)
                if not phone or len(phone) < 10:
                    skipped += 1
                    continue
                if not oid:
                    skipped += 1
                    continue
                parts = oid.split("-", 1)
                if len(parts) != 2:
                    skipped += 1
                    continue
                ward_code, pu_code = parts[0].strip(), parts[1].strip()

                if lga:
                    # Preferred: match on ward_code + pu_code + LGA (unique)
                    cur.execute(
                        """UPDATE polling_units SET officer_phone = ?
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun'
                           AND LOWER(lg) = LOWER(?)""",
                        (phone, ward_code, pu_code, lga)
                    )
                else:
                    # Fallback: no LGA provided — only safe if combo is unique
                    cur.execute(
                        """UPDATE polling_units SET officer_phone = ?
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun'""",
                        (phone, ward_code, pu_code)
                    )

                if conn._conn.total_changes > 0:
                    updated += 1
                else:
                    no_match.append(f"{oid} ({lga})")
                    skipped += 1
        conn.commit()

    result = {"status": "ok", "updated": updated, "skipped": skipped}
    if no_match:
        result["not_found"] = no_match[:20]  # return first 20 unmatched for debugging
    return result
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/submissions")
async def get_dashboard_data(request: Request):
    _require_dashboard(request)
    require_review = os.environ.get("REQUIRE_REVIEW", "0").lower() in ("1", "true", "yes")
    where = "WHERE LOWER(state) = 'osun'"
    if require_review:
        where += " AND reviewed = 1"
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT * FROM field_submissions {where} ORDER BY timestamp DESC")
            rows = cur.fetchall()
            data = []
            for r in rows:
                # Use edited votes if admin has overridden them
                votes_src = r.get('edited_votes_json') or r.get('votes_json')
                v = json.loads(votes_src) if isinstance(votes_src, str) else (votes_src or {})
                raw = r.get('ec8e_image')
                if raw:
                    ec8e_url = raw if raw.startswith('http') else f"/ec8e/{raw}"
                else:
                    ec8e_url = None
                entry = {
                    "pu_name": r['location'], "state": r['state'], "lga": r['lg'], "ward": r['ward'],
                    "latitude": r['lat'], "longitude": r['lon'],
                    "ec8e_image": ec8e_url,
                    "reg_voters": r.get('reg_voters') or 0,
                    "total_accredited": r.get('total_accredited') or 0,
                    "total_cast": r.get('total_cast') or 0,
                    "officer_id": r.get('officer_id') or '',
                    "timestamp": str(r.get('timestamp') or ''),
                    "pu_code": r.get('pu_code') or '',
                    "reviewed": bool(r.get('reviewed')),
                }
                for p in ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"]:
                    entry[f"votes_party_{p}"] = v.get(p, 0)
                data.append(entry)
            return data


@app.get("/ec8e/{filename}")
async def serve_ec8e(filename: str):
    """Legacy fallback for locally-stored EC8E images. New uploads use Cloudinary CDN."""
    import mimetypes
    filepath = os.path.join(EC8E_PATH, filename)
    if not os.path.exists(filepath):
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Image not found")
    mime = mimetypes.guess_type(filepath)[0] or "image/jpeg"
    with open(filepath, "rb") as f:
        content = f.read()
    from fastapi.responses import Response
    return Response(content=content, media_type=mime, headers={
        "Cache-Control": "public, max-age=86400",
        "Access-Control-Allow-Origin": "*"
    })



RESULTS_PORTAL_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>TKC — Election Observation Portal</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{--green:#008751;--green-light:#00b368;--gold:#ffc107;--dark:#020c06;--panel:rgba(255,255,255,0.04);--border:rgba(255,255,255,0.1)}
body{font-family:'Inter',sans-serif;background:var(--dark);color:#fff;min-height:100vh}
body::before{content:'';position:fixed;inset:0;z-index:0;background:radial-gradient(ellipse 120% 80% at 50% -10%,rgba(0,135,81,0.15) 0%,transparent 60%),linear-gradient(160deg,#020c06 0%,#041508 40%,#020c06 100%)}
.page{position:relative;z-index:1;max-width:1100px;margin:0 auto;padding:28px 20px 60px}
#loginGate{min-height:100vh;display:flex;align-items:center;justify-content:center;position:relative;z-index:1}
.login-card{background:var(--panel);border:1px solid rgba(0,135,81,0.25);border-radius:20px;padding:48px 40px;width:100%;max-width:400px;text-align:center}
.lock-icon{width:60px;height:60px;border-radius:50%;background:rgba(0,135,81,0.15);border:2px solid rgba(0,135,81,0.3);display:flex;align-items:center;justify-content:center;font-size:1.6rem;margin:0 auto 20px}
h1{font-size:1.5rem;font-weight:800;margin-bottom:6px}
.sub{color:rgba(255,255,255,0.4);font-size:0.82rem;margin-bottom:28px}
.form-control{width:100%;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:10px;color:#fff;padding:12px 16px;font-size:0.9rem;outline:none;margin-bottom:12px}
.form-control:focus{border-color:var(--green)}
.btn-primary{width:100%;background:linear-gradient(135deg,var(--green),var(--green-light));border:none;border-radius:10px;color:#fff;font-weight:700;font-size:0.9rem;padding:13px;cursor:pointer;margin-top:4px}
.btn-primary:hover{opacity:0.9}
.err-msg{color:#ff6b6b;font-size:0.78rem;margin-top:8px;display:none}
/* Portal layout */
#portal{display:none}
.portal-header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:24px;flex-wrap:wrap;gap:16px}
.badge-tag{font-size:0.62rem;font-weight:700;letter-spacing:0.15em;color:var(--gold);background:rgba(255,193,7,0.1);border:1px solid rgba(255,193,7,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:8px;text-transform:uppercase}
.portal-header h1{font-size:1.6rem;font-weight:900;margin-bottom:4px}
.portal-header p{color:rgba(255,255,255,0.45);font-size:0.8rem}
.btn-secondary{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:rgba(255,255,255,0.7);font-size:0.78rem;font-weight:600;padding:8px 16px;cursor:pointer}
.btn-secondary:hover{background:rgba(255,255,255,0.1)}
/* Filters */
.filters{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:18px;align-items:center}
.filter-select{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:#fff;padding:7px 12px;font-size:0.78rem;outline:none;cursor:pointer}
.filter-select option{background:#041508}
.search-input{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:#fff;padding:7px 14px;font-size:0.78rem;outline:none;width:220px}
.search-input:focus{border-color:var(--green)}
/* Stats pills */
.stats-row{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:22px}
.stat-pill{background:var(--panel);border:1px solid var(--border);border-radius:12px;padding:12px 18px;display:flex;flex-direction:column;align-items:center;min-width:110px}
.stat-pill .val{font-size:1.4rem;font-weight:800;line-height:1}
.stat-pill .lbl{font-size:0.65rem;color:rgba(255,255,255,0.4);margin-top:4px;text-transform:uppercase;letter-spacing:0.06em}
.val-green{color:#00cc66}.val-gold{color:#ffc107}.val-red{color:#ff6b6b}.val-blue{color:#4fc3f7}
/* Table */
.tbl-wrap{overflow-x:auto;border-radius:12px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;font-size:0.78rem}
thead th{background:rgba(255,255,255,0.04);padding:10px 10px;text-align:left;font-size:0.65rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;white-space:nowrap;border-bottom:1px solid rgba(255,255,255,0.08)}
tbody tr{border-bottom:1px solid rgba(255,255,255,0.05);transition:background 0.1s}
tbody tr:hover{background:rgba(255,255,255,0.03)}
td{padding:9px 10px;vertical-align:middle}
.badge{display:inline-block;padding:2px 8px;border-radius:20px;font-size:0.62rem;font-weight:700;letter-spacing:0.05em}
.badge-ok{background:rgba(0,204,102,0.15);color:#00cc66;border:1px solid rgba(0,204,102,0.3)}
.badge-pending{background:rgba(255,193,7,0.12);color:#ffc107;border:1px solid rgba(255,193,7,0.25)}
.badge-edited{background:rgba(79,195,247,0.12);color:#4fc3f7;border:1px solid rgba(79,195,247,0.25)}
/* Action buttons */
.btn-a{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:6px;color:#fff;cursor:pointer;font-size:0.68rem;padding:3px 8px;margin-right:3px;white-space:nowrap}
.btn-a:hover{background:rgba(255,255,255,0.12)}
.btn-approve{border-color:rgba(0,204,102,0.4);color:#00cc66}
.btn-unapprove{border-color:rgba(255,193,7,0.4);color:#ffc107}
.btn-edit-r{border-color:rgba(79,195,247,0.4);color:#4fc3f7}
/* Edit modal */
.modal-overlay{position:fixed;inset:0;background:rgba(0,0,0,0.75);z-index:1000;display:flex;align-items:center;justify-content:center;padding:20px}
.modal{background:#041508;border:1px solid rgba(0,135,81,0.3);border-radius:16px;width:100%;max-width:560px;padding:28px;max-height:90vh;overflow-y:auto}
.modal h2{font-size:1.1rem;font-weight:800;margin-bottom:4px}
.modal .sub{margin-bottom:20px}
.field-row{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}
.field-group label{font-size:0.68rem;color:rgba(255,255,255,0.4);display:block;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.06em}
.field-group input{width:100%;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.15);border-radius:8px;color:#fff;padding:8px 10px;font-size:0.82rem;outline:none}
.field-group input:focus{border-color:var(--green)}
.parties-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(100px,1fr));gap:8px;margin-bottom:16px}
.modal-actions{display:flex;gap:10px;justify-content:flex-end;margin-top:16px}
.btn-save-modal{background:linear-gradient(135deg,var(--green),var(--green-light));border:none;border-radius:8px;color:#fff;font-weight:700;font-size:0.82rem;padding:10px 22px;cursor:pointer}
.btn-cancel-modal{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.15);border-radius:8px;color:rgba(255,255,255,0.7);font-size:0.82rem;padding:10px 18px;cursor:pointer}
.note-input{width:100%;background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.15);border-radius:8px;color:#fff;padding:8px 10px;font-size:0.8rem;resize:vertical;min-height:60px;margin-bottom:12px;outline:none}
/* Pagination */
.pagination{display:flex;gap:8px;align-items:center;justify-content:flex-end;margin-top:14px;font-size:0.78rem;color:rgba(255,255,255,0.4)}
.page-btn{background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:6px;color:#fff;cursor:pointer;padding:4px 10px;font-size:0.75rem}
.page-btn:disabled{opacity:0.3;cursor:default}
/* Pending badge on dashboard link */
.pending-badge{display:inline-block;background:#ff6b6b;color:#fff;border-radius:20px;font-size:0.6rem;font-weight:800;padding:1px 6px;margin-left:6px;vertical-align:middle}
</style>
</head>
<body>

<!-- Login gate -->
<div id="loginGate">
  <div class="login-card">
    <div class="lock-icon">🔐</div>
    <h1>Results Portal</h1>
    <p class="sub">TKC Field Collation — Review & Approve</p>
    <input type="password" id="rKey" class="form-control" placeholder="Admin key" onkeydown="if(event.key==='Enter')rLogin()">
    <button class="btn-primary" id="rLoginBtn" onclick="rLogin()">Access Results Portal →</button>
    <div class="err-msg" id="rLoginErr">Incorrect key. Try again.</div>
  </div>
</div>

<!-- Portal -->
<div id="portal" class="page">
  <div class="portal-header">
    <div>
      <div class="badge-tag">Admin — Results Review Portal</div>
      <h1>📊 Submitted Results</h1>
      <p>Review, edit and approve PU submissions before they appear on the dashboard.</p>
    </div>
    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
      <a href="/admin" target="_blank" class="btn-secondary">⚙️ Officer Admin</a>
      <button class="btn-secondary" onclick="rLogout()">🔒 Lock</button>
    </div>
  </div>

  <!-- Stats -->
  <div class="stats-row">
    <div class="stat-pill"><span class="val val-gold" id="rStatTotal">—</span><span class="lbl">Total</span></div>
    <div class="stat-pill"><span class="val val-green" id="rStatApproved">—</span><span class="lbl">Approved</span></div>
    <div class="stat-pill"><span class="val val-red" id="rStatPending">—</span><span class="lbl">Pending</span></div>
    <div class="stat-pill"><span class="val val-blue" id="rStatEdited">—</span><span class="lbl">Edited</span></div>
  </div>

  <!-- Filters -->
  <div class="filters">
    <select class="filter-select" id="rFilterLga" onchange="loadResults(1)">
      <option value="">All LGAs</option>
    </select>
    <select class="filter-select" id="rFilterStatus" onchange="loadResults(1)">
      <option value="">All Status</option>
      <option value="pending">Pending Review</option>
      <option value="reviewed">Approved</option>
    </select>
    <input type="text" class="search-input" id="rSearch" placeholder="Search PU, ward, officer..." oninput="filterTable(this.value)">
    <button class="btn-secondary" onclick="approveAllVisible()" style="margin-left:auto;">✅ Approve All Visible</button>
  </div>

  <!-- Table -->
  <div class="tbl-wrap">
    <table>
      <thead>
        <tr>
          <th>#</th>
          <th>PU Code</th>
          <th>Location</th>
          <th>LGA</th>
          <th>Ward</th>
          <th>Officer</th>
          <th>ACCORD</th>
          <th>Total Cast</th>
          <th>Accredited</th>
          <th>Timestamp</th>
          <th>Status</th>
          <th>Actions</th>
        </tr>
      </thead>
      <tbody id="rTableBody">
        <tr><td colspan="12" style="text-align:center;padding:30px;color:rgba(255,255,255,0.3);">Login to view results.</td></tr>
      </tbody>
    </table>
  </div>
  <div class="pagination">
    <span id="rPageInfo"></span>
    <button class="page-btn" id="rPrevBtn" onclick="loadResults(_rPage-1)" disabled>◀ Prev</button>
    <button class="page-btn" id="rNextBtn" onclick="loadResults(_rPage+1)" disabled>Next ▶</button>
  </div>
</div>

<!-- Edit Modal -->
<div class="modal-overlay" id="editModal" style="display:none;" onclick="if(event.target===this)closeModal()">
  <div class="modal">
    <h2>✏️ Edit Submission</h2>
    <p class="sub" id="modalSub"></p>
    <input type="hidden" id="modalId">
    <div class="field-row">
      <div class="field-group"><label>Total Accredited</label><input type="number" id="mAccredited" min="0"></div>
      <div class="field-group"><label>Total Cast</label><input type="number" id="mCast" min="0"></div>
      <div class="field-group"><label>Valid Votes</label><input type="number" id="mValid" min="0"></div>
      <div class="field-group"><label>Rejected Votes</label><input type="number" id="mRejected" min="0"></div>
    </div>
    <div style="font-size:0.68rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;margin-bottom:8px;">Party Votes</div>
    <div class="parties-grid" id="modalParties"></div>
    <div style="font-size:0.68rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;margin-bottom:6px;">Edit Note</div>
    <textarea class="note-input" id="mNote" placeholder="Reason for edit (optional)..."></textarea>
    <div class="modal-actions">
      <button class="btn-cancel-modal" onclick="closeModal()">Cancel</button>
      <button class="btn-save-modal" onclick="saveEdit()">💾 Save Changes</button>
    </div>
  </div>
</div>

<script>
const PARTIES = ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"];
let _rKey = '', _rPage = 1, _rData = [], _rTotal = 0, _rPages = 1;

async function rLogin() {
    const key = document.getElementById('rKey').value.trim();
    if (!key) return;
    const btn = document.getElementById('rLoginBtn');
    const err = document.getElementById('rLoginErr');
    btn.disabled = true; btn.textContent = 'Verifying...';
    err.style.display = 'none';
    try {
        const res = await fetch('/api/admin/officer-stats', {
            headers: { 'Authorization': 'Bearer ' + key }
        });
        if (!res.ok) {
            err.style.display = 'block';
            btn.disabled = false; btn.textContent = 'Access Results Portal →';
            return;
        }
        _rKey = key;
        document.getElementById('loginGate').style.display = 'none';
        document.getElementById('portal').style.display = 'block';
        loadResults(1);
        loadLgas();
        loadSummaryStats();
    } catch(e) {
        err.textContent = 'Server error.';
        err.style.display = 'block';
        btn.disabled = false; btn.textContent = 'Access Results Portal →';
    }
}

function rLogout() {
    _rKey = '';
    document.getElementById('portal').style.display = 'none';
    document.getElementById('loginGate').style.display = 'flex';
    document.getElementById('rKey').value = '';
}

async function loadSummaryStats() {
    try {
        const res = await fetch('/api/admin/pending-review-count', {
            headers: { 'Authorization': 'Bearer ' + _rKey },
            credentials: 'include'
        });
        if (!res.ok) return;
        const d = await res.json();
        document.getElementById('rStatTotal').textContent    = d.total.toLocaleString();
        document.getElementById('rStatPending').textContent  = d.pending.toLocaleString();
        document.getElementById('rStatApproved').textContent = (d.total - d.pending).toLocaleString();
    } catch(e) {}
}

async function loadLgas() {
    try {
        // /api/lgas/osun requires no auth — returns plain string array
        const res = await fetch('/api/lgas/osun');
        if (!res.ok) return;
        const d = await res.json();
        const sel = document.getElementById('rFilterLga');
        const items = Array.isArray(d) ? d : (d.lgas || []);
        items.forEach(l => {
            const o = document.createElement('option');
            o.value = typeof l === 'string' ? l : l.lg;
            o.textContent = typeof l === 'string' ? l : l.lg;
            sel.appendChild(o);
        });
    } catch(e) {}
}

async function loadResults(page) {
    _rPage = page || 1;
    const lga    = document.getElementById('rFilterLga').value;
    const status = document.getElementById('rFilterStatus').value;
    const tbody  = document.getElementById('rTableBody');
    tbody.innerHTML = '<tr><td colspan="12" style="text-align:center;padding:24px;color:rgba(255,255,255,0.3);">Loading...</td></tr>';
    try {
        const res = await fetch(`/api/admin/results?page=${_rPage}&lga=${encodeURIComponent(lga)}&status=${status}`, {
            headers: { 'Authorization': 'Bearer ' + _rKey }
        });
        const d = await res.json();
        _rData   = d.results || [];
        _rTotal  = d.total || 0;
        _rPages  = d.pages || 1;
        renderTable(_rData);
        document.getElementById('rPageInfo').textContent = `Page ${_rPage} of ${_rPages} · ${_rTotal.toLocaleString()} results`;
        document.getElementById('rPrevBtn').disabled = _rPage <= 1;
        document.getElementById('rNextBtn').disabled = _rPage >= _rPages;
        // Count edited
        const editedCount = _rData.filter(r => r.edit_note || false).length;
        document.getElementById('rStatEdited').textContent = editedCount;
        loadSummaryStats();
    } catch(e) {
        tbody.innerHTML = '<tr><td colspan="12" style="color:#ff6b6b;text-align:center;padding:20px;">Failed to load results.</td></tr>';
    }
}

function renderTable(data) {
    const tbody = document.getElementById('rTableBody');
    if (!data.length) {
        tbody.innerHTML = '<tr><td colspan="12" style="text-align:center;padding:30px;color:rgba(255,255,255,0.3);">No results found.</td></tr>';
        return;
    }
    const offset = (_rPage - 1) * 50;
    tbody.innerHTML = data.map((r, i) => {
        const statusBadge = r.reviewed
            ? '<span class="badge badge-ok">✅ Approved</span>'
            : '<span class="badge badge-pending">⏳ Pending</span>';
        const editedBadge = r.edit_note ? ' <span class="badge badge-edited">✏️ Edited</span>' : '';
        const accord = r.votes ? (r.votes.ACCORD || 0) : 0;
        const ts = r.timestamp ? r.timestamp.substring(0,16).replace('T',' ') : '—';
        return `<tr id="rrow-${r.id}">
            <td style="color:rgba(255,255,255,0.3);">${offset + i + 1}</td>
            <td style="font-weight:700;color:#ffc107;white-space:nowrap;">${r.pu_code || '—'}</td>
            <td style="max-width:140px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${r.location||''}">${r.location || '—'}</td>
            <td style="font-size:0.72rem;color:rgba(255,255,255,0.6);">${r.lg || '—'}</td>
            <td style="font-size:0.7rem;color:rgba(255,255,255,0.5);">${r.ward || '—'}</td>
            <td style="font-size:0.72rem;">${r.officer_id || '—'}</td>
            <td style="font-weight:700;color:#00cc66;">${accord.toLocaleString()}</td>
            <td>${(r.total_cast||0).toLocaleString()}</td>
            <td>${(r.total_accredited||0).toLocaleString()}</td>
            <td style="font-size:0.68rem;color:rgba(255,255,255,0.4);white-space:nowrap;">${ts}</td>
            <td>${statusBadge}${editedBadge}</td>
            <td style="white-space:nowrap;">
                <button class="btn-a btn-edit-r" onclick="openEditModal(${r.id})">✏️ Edit</button>
                ${r.reviewed
                    ? `<button class="btn-a btn-unapprove" onclick="toggleApprove(${r.id}, false)">↩ Undo</button>`
                    : `<button class="btn-a btn-approve"   onclick="toggleApprove(${r.id}, true)">✅ Approve</button>`}
            </td>
        </tr>`;
    }).join('');
}

function filterTable(q) {
    if (!q) { renderTable(_rData); return; }
    const lq = q.toLowerCase();
    renderTable(_rData.filter(r =>
        (r.pu_code||'').toLowerCase().includes(lq) ||
        (r.location||'').toLowerCase().includes(lq) ||
        (r.ward||'').toLowerCase().includes(lq) ||
        (r.officer_id||'').toLowerCase().includes(lq)
    ));
}

async function toggleApprove(id, approve) {
    try {
        const res = await fetch(`/api/admin/results/${id}/approve`, {
            method: 'POST',
            headers: { 'Authorization': 'Bearer ' + _rKey, 'Content-Type': 'application/json' },
            body: JSON.stringify({ approve })
        });
        if (!res.ok) { alert('Failed to update approval'); return; }
        loadResults(_rPage);
    } catch(e) { alert('Server error'); }
}

async function approveAllVisible() {
    const pending = _rData.filter(r => !r.reviewed);
    if (!pending.length) { alert('No pending results on this page.'); return; }
    if (!confirm(`Approve all ${pending.length} pending results on this page?`)) return;
    for (const r of pending) {
        await fetch(`/api/admin/results/${r.id}/approve`, {
            method: 'POST',
            headers: { 'Authorization': 'Bearer ' + _rKey, 'Content-Type': 'application/json' },
            body: JSON.stringify({ approve: true })
        });
    }
    loadResults(_rPage);
}

function openEditModal(id) {
    const r = _rData.find(x => x.id === id);
    if (!r) return;
    document.getElementById('modalId').value = id;
    document.getElementById('modalSub').textContent = `${r.pu_code} · ${r.location} · ${r.lg}`;
    document.getElementById('mAccredited').value = r.total_accredited || 0;
    document.getElementById('mCast').value       = r.total_cast || 0;
    document.getElementById('mValid').value      = r.valid_votes || 0;
    document.getElementById('mRejected').value   = r.rejected_votes || 0;
    document.getElementById('mNote').value       = r.edit_note || '';
    const grid = document.getElementById('modalParties');
    grid.innerHTML = PARTIES.map(p => `
        <div class="field-group">
            <label>${p}</label>
            <input type="number" id="mp-${p}" min="0" value="${(r.votes && r.votes[p]) || 0}">
        </div>`).join('');
    document.getElementById('editModal').style.display = 'flex';
}

function closeModal() {
    document.getElementById('editModal').style.display = 'none';
}

async function saveEdit() {
    const id   = parseInt(document.getElementById('modalId').value);
    const body = {
        total_accredited: parseInt(document.getElementById('mAccredited').value) || 0,
        total_cast:       parseInt(document.getElementById('mCast').value)       || 0,
        valid_votes:      parseInt(document.getElementById('mValid').value)      || 0,
        rejected_votes:   parseInt(document.getElementById('mRejected').value)   || 0,
        edit_note:        document.getElementById('mNote').value.trim(),
        votes: {}
    };
    PARTIES.forEach(p => {
        body.votes[p] = parseInt(document.getElementById(`mp-${p}`).value) || 0;
    });
    try {
        const res = await fetch(`/api/admin/results/${id}`, {
            method: 'PUT',
            headers: { 'Authorization': 'Bearer ' + _rKey, 'Content-Type': 'application/json' },
            body: JSON.stringify(body)
        });
        if (!res.ok) { const e = await res.json(); alert(e.detail || 'Save failed'); return; }
        closeModal();
        loadResults(_rPage);
    } catch(e) { alert('Server error'); }
}
</script>
</body>
</html>
"""

ADMIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TKC — Admin Portal</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
        :root {
            --green: #008751; --green-light: #00b368; --gold: #ffc107;
            --dark: #020c06; --panel: rgba(255,255,255,0.04); --border: rgba(255,255,255,0.1);
        }
        body { font-family: 'Inter', sans-serif; background: var(--dark); color: #fff; min-height: 100vh; }
        body::before {
            content: ''; position: fixed; inset: 0; z-index: 0;
            background: radial-gradient(ellipse 120% 80% at 50% -10%, rgba(0,135,81,0.15) 0%, transparent 60%),
                        linear-gradient(160deg, #020c06 0%, #041508 40%, #020c06 100%);
        }
        .grid-bg {
            position: fixed; inset: 0; z-index: 0;
            background-image: linear-gradient(rgba(0,135,81,0.05) 1px, transparent 1px),
                              linear-gradient(90deg, rgba(0,135,81,0.05) 1px, transparent 1px);
            background-size: 60px 60px; pointer-events: none;
        }

        /* ── Layout ── */
        .page { position: relative; z-index: 1; max-width: 860px; margin: 0 auto; padding: 32px 20px 60px; }

        /* ── Login gate ── */
        #loginGate {
            min-height: 100vh; display: flex; align-items: center; justify-content: center;
        }
        .login-card {
            background: var(--panel); border: 1px solid rgba(0,135,81,0.25); border-radius: 20px;
            padding: 48px 40px; width: 100%; max-width: 400px; text-align: center;
            box-shadow: 0 0 60px rgba(0,135,81,0.1);
        }
        .lock-icon {
            width: 60px; height: 60px; border-radius: 50%;
            background: rgba(0,135,81,0.12); border: 1px solid rgba(0,135,81,0.3);
            display: flex; align-items: center; justify-content: center;
            font-size: 1.5rem; margin: 0 auto 20px;
        }
        .badge-tag {
            display: inline-block; font-size: 0.58rem; font-weight: 700;
            letter-spacing: 0.18em; text-transform: uppercase; color: var(--gold);
            background: rgba(255,193,7,0.1); border: 1px solid rgba(255,193,7,0.2);
            border-radius: 20px; padding: 3px 12px; margin-bottom: 14px;
        }
        .login-card h2 { font-size: 1.3rem; font-weight: 900; margin-bottom: 6px; }
        .login-card p { font-size: 0.78rem; color: rgba(255,255,255,0.4); margin-bottom: 28px; }

        /* ── Inputs ── */
        input[type="password"], input[type="text"], input[type="file"] {
            width: 100%; background: rgba(255,255,255,0.05);
            border: 1px solid rgba(0,135,81,0.3); border-radius: 10px;
            color: #fff; font-size: 0.9rem; padding: 12px 16px; outline: none;
            transition: border-color 0.2s; font-family: 'Inter', sans-serif;
        }
        input:focus { border-color: rgba(0,135,81,0.7); }
        input[type="file"] { cursor: pointer; padding: 10px 14px; }
        textarea {
            width: 100%; background: rgba(255,255,255,0.04);
            border: 1px solid rgba(0,135,81,0.25); border-radius: 10px;
            color: #fff; font-size: 0.82rem; padding: 12px 14px; outline: none;
            font-family: 'Inter', monospace; resize: vertical; min-height: 120px;
            transition: border-color 0.2s;
        }
        textarea:focus { border-color: rgba(0,135,81,0.6); }

        /* ── Buttons ── */
        .btn-primary {
            background: linear-gradient(135deg, #008751, #00b368); border: none;
            border-radius: 10px; color: #fff; font-size: 0.9rem; font-weight: 700;
            padding: 12px 24px; cursor: pointer; width: 100%; transition: opacity 0.2s;
        }
        .btn-primary:hover { opacity: 0.88; }
        .btn-primary:disabled { opacity: 0.45; cursor: not-allowed; }
        .btn-secondary {
            background: rgba(255,255,255,0.06); border: 1px solid var(--border);
            border-radius: 10px; color: rgba(255,255,255,0.7); font-size: 0.82rem;
            font-weight: 600; padding: 10px 20px; cursor: pointer; transition: background 0.2s;
        }
        .btn-secondary:hover { background: rgba(255,255,255,0.1); }
        .btn-gold {
            background: rgba(255,193,7,0.12); border: 1px solid rgba(255,193,7,0.3);
            border-radius: 10px; color: var(--gold); font-size: 0.82rem; font-weight: 700;
            padding: 10px 20px; cursor: pointer; text-decoration: none; display: inline-block;
            transition: background 0.2s;
        }
        .btn-gold:hover { background: rgba(255,193,7,0.22); }

        /* ── Admin portal ── */
        #adminPortal { display: none; }
        .portal-header { margin-bottom: 32px; }
        .portal-header h1 { font-size: 1.6rem; font-weight: 900; margin-bottom: 6px; }
        .portal-header p { font-size: 0.82rem; color: rgba(255,255,255,0.4); }
        .header-row { display: flex; align-items: flex-start; justify-content: space-between; gap: 16px; flex-wrap: wrap; }

        /* ── Cards ── */
        .card {
            background: var(--panel); border: 1px solid var(--border);
            border-radius: 16px; padding: 28px; margin-bottom: 20px;
        }
        .card-title {
            font-size: 0.65rem; font-weight: 700; letter-spacing: 0.15em;
            text-transform: uppercase; color: var(--gold);
            border-left: 3px solid var(--gold); padding-left: 10px;
            margin-bottom: 20px; display: block;
        }

        /* ── Upload zone ── */
        .upload-zone {
            border: 2px dashed rgba(0,135,81,0.3); border-radius: 12px;
            padding: 36px 20px; text-align: center; cursor: pointer;
            transition: all 0.2s; background: rgba(0,135,81,0.03); position: relative;
        }
        .upload-zone:hover, .upload-zone.drag-over {
            border-color: rgba(0,135,81,0.7); background: rgba(0,135,81,0.07);
        }
        .upload-zone input[type="file"] {
            position: absolute; inset: 0; opacity: 0; cursor: pointer;
            width: 100%; height: 100%; border: none; padding: 0;
        }
        .upload-icon { font-size: 2rem; margin-bottom: 10px; }
        .upload-label { font-size: 0.88rem; font-weight: 600; color: rgba(255,255,255,0.7); }
        .upload-hint { font-size: 0.72rem; color: rgba(255,255,255,0.3); margin-top: 4px; }
        .file-chosen { font-size: 0.78rem; color: var(--green-light); margin-top: 8px; font-weight: 600; }

        /* ── Preview table ── */
        .preview-wrap { overflow-x: auto; margin-top: 16px; border-radius: 8px; overflow: hidden; }
        table { width: 100%; border-collapse: collapse; font-size: 0.78rem; }
        thead tr { background: rgba(0,135,81,0.15); }
        th { padding: 8px 12px; text-align: left; color: var(--gold); font-weight: 700;
             font-size: 0.65rem; letter-spacing: 0.08em; text-transform: uppercase; }
        tbody tr { border-bottom: 1px solid rgba(255,255,255,0.05); }
        tbody tr:hover { background: rgba(255,255,255,0.03); }
        .btn-action { background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.12); border-radius: 6px; color: #fff; cursor: pointer; font-size: 0.7rem; padding: 3px 8px; margin-right: 4px; transition: background 0.15s; }
        .btn-action:hover { background: rgba(255,255,255,0.12); }
        .btn-action:disabled { opacity: 0.4; cursor: default; }
        .btn-edit { border-color: rgba(0,179,104,0.4); color: #00cc66; }
        .btn-del  { border-color: rgba(255,107,107,0.4); color: #ff6b6b; }
        .btn-save { border-color: rgba(0,179,104,0.6); color: #00cc66; background: rgba(0,135,81,0.15); }
        .btn-cancel { border-color: rgba(255,193,7,0.4); color: #ffc107; }
        .btn-filter { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.1); border-radius: 20px; color: rgba(255,255,255,0.5); cursor: pointer; font-size: 0.7rem; padding: 4px 12px; transition: all 0.15s; }
        .btn-filter:hover { background: rgba(255,255,255,0.08); color: #fff; }
        .btn-filter.active { background: rgba(0,135,81,0.2); border-color: rgba(0,135,81,0.5); color: #00cc66; font-weight: 700; }
        .phone-missing { color: rgba(255,255,255,0.25); font-style: italic; font-size: 0.72rem; }
        td { padding: 7px 12px; color: rgba(255,255,255,0.75); }
        td.valid { color: #00cc66; }
        td.invalid { color: #ff6b6b; }
        .row-num { color: rgba(255,255,255,0.25); font-size: 0.68rem; }

        /* ── Result banner ── */
        .result-banner {
            border-radius: 10px; padding: 14px 18px; margin-top: 16px;
            font-size: 0.82rem; font-weight: 600; display: none;
        }
        .result-banner.success { background: rgba(0,204,102,0.12); border: 1px solid rgba(0,204,102,0.3); color: #00cc66; }
        .result-banner.error   { background: rgba(255,107,107,0.12); border: 1px solid rgba(255,107,107,0.3); color: #ff6b6b; }
        .result-banner.info    { background: rgba(255,193,7,0.1); border: 1px solid rgba(255,193,7,0.25); color: var(--gold); }

        /* ── Stats row ── */
        .stats-row { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 20px; }
        .stat-pill {
            background: rgba(255,255,255,0.04); border: 1px solid var(--border);
            border-radius: 10px; padding: 10px 18px; text-align: center; flex: 1; min-width: 80px;
        }
        .stat-pill .val { font-size: 1.4rem; font-weight: 900; display: block; }
        .stat-pill .lbl { font-size: 0.6rem; color: rgba(255,255,255,0.35); text-transform: uppercase; letter-spacing: 0.08em; }
        .val-green { color: #00cc66; }
        .val-red   { color: #ff6b6b; }
        .val-gold  { color: var(--gold); }

        /* ── Error alert ── */
        .err-box { background: rgba(220,53,69,0.12); border: 1px solid rgba(220,53,69,0.3);
                   border-radius: 8px; color: #ff6b6b; font-size: 0.78rem; padding: 10px 14px;
                   margin-bottom: 12px; display: none; }
        .divider { border: none; border-top: 1px solid var(--border); margin: 24px 0; }
        label.field-label { display: block; font-size: 0.72rem; font-weight: 600;
                            color: rgba(255,255,255,0.5); margin-bottom: 6px; }
    </style>
</head>
<body>
<div class="grid-bg"></div>

<!-- ── Login gate ── -->
<div id="loginGate">
    <div class="login-card">
        <div class="lock-icon">🔐</div>
        <div class="badge-tag">Admin Portal</div>
        <h2>Officer Management</h2>
        <p>Enter your admin key to manage officer phone numbers.</p>
        <div class="err-box" id="loginErr"></div>
        <div style="position:relative; margin-bottom:12px;">
            <input type="password" id="adminKey" placeholder="Enter admin key"
                   autocomplete="off" onkeydown="if(event.key==='Enter')adminLogin()"
                   style="padding-right:44px;">
            <button onclick="togglePwd()" style="position:absolute;right:12px;top:50%;transform:translateY(-50%);
                background:none;border:none;color:rgba(255,255,255,0.3);cursor:pointer;font-size:1rem;">👁</button>
        </div>
        <button class="btn-primary" id="loginBtn" onclick="adminLogin()">Access Admin Portal →</button>
        <div style="margin-top:16px;"><a href="/" style="font-size:0.72rem;color:rgba(255,255,255,0.25);text-decoration:none;">← Back to home</a></div>
    </div>
</div>

<!-- ── Admin portal ── -->
<div id="adminPortal">
<div class="page">

    <div class="portal-header">
        <div class="header-row">
            <div>
                <div class="badge-tag">Admin Portal — Officer Management</div>
                <h1>📋 Officer Phone Numbers</h1>
                <p>Upload a CSV or paste officer IDs and phone numbers to register them for WhatsApp OTP authentication.</p>
            </div>
            <div style="display:flex;gap:8px;flex-wrap:wrap;align-items:center;">
                <a href="/admin/results" target="_blank" class="btn-gold" style="background:rgba(0,179,104,0.15);border-color:rgba(0,179,104,0.4);">📊 Results Portal</a>
                <a id="templateDownloadBtn" href="#" onclick="downloadTemplate(event)" class="btn-gold">⬇ Download CSV Template</a>
                <button class="btn-secondary" onclick="adminLogout()">🔒 Lock</button>
            </div>
        </div>
    </div>

    <!-- ── Stats ── -->
    <div class="stats-row">
        <div class="stat-pill"><span class="val val-gold" id="statTotal">—</span><span class="lbl">Total Officers</span></div>
        <div class="stat-pill"><span class="val val-green" id="statWithPhone">—</span><span class="lbl">With Phone</span></div>
        <div class="stat-pill"><span class="val val-red" id="statNoPhone">—</span><span class="lbl">Missing Phone</span></div>
    </div>

    <!-- ── CSV Upload ── -->
    <div class="card">
        <span class="card-title">📁 Upload CSV File</span>
        <div class="upload-zone" id="uploadZone">
            <input type="file" id="csvFile" accept=".csv,text/csv" onchange="handleFileSelect(this)">
            <div class="upload-icon">📂</div>
            <div class="upload-label">Drop your CSV here or click to browse</div>
            <div class="upload-hint">Columns: <code>officer_id</code>, <code>lga</code>, <code>ward</code>, <code>polling_unit</code>, <code>phone</code> &nbsp;·&nbsp; Use the template from this portal</div>
            <div class="file-chosen d-none" id="fileChosen"></div>
        </div>

        <div id="previewSection" style="display:none;">
            <hr class="divider">
            <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
                <span style="font-size:0.78rem;color:rgba(255,255,255,0.5);" id="previewLabel"></span>
                <button class="btn-secondary" onclick="clearFile()" style="padding:6px 14px;font-size:0.72rem;">✕ Clear</button>
            </div>
            <div class="preview-wrap"><table id="previewTable"><thead></thead><tbody></tbody></table></div>
        </div>

        <div class="result-banner" id="uploadResult"></div>
        <div style="margin-top:16px;">
            <button class="btn-primary" id="uploadBtn" onclick="submitCSV()" disabled>Upload & Register Officers</button>
        </div>
    </div>

    <!-- ── Manual paste ── -->
    <div class="card">
        <span class="card-title">✏️ Manual Entry (Paste or Type)</span>
        <p style="font-size:0.78rem;color:rgba(255,255,255,0.4);margin-bottom:14px;">
            One officer per line. Format: <code style="color:#00cc66;">officer_id,lga,phone</code><br>
            <span style="font-size:0.72rem;color:rgba(255,255,255,0.25);">LGA is required to uniquely identify the polling unit across all 30 LGAs.</span>
        </p>
        <label class="field-label">Officer entries (officer_id, lga, phone)</label>
        <textarea id="manualInput" placeholder="10-001,Osogbo,+2348012345678&#10;06-001,Ife Central,+2348023456789&#10;04-001,Atakumosa East,+2348034567890"></textarea>
        <div class="result-banner" id="manualResult"></div>
        <div style="margin-top:14px;">
            <button class="btn-primary" id="manualBtn" onclick="submitManual()">Register These Officers</button>
        </div>
    </div>

</div>

    <!-- ── All Officers Table ── -->
    <div class="card" id="officerTableCard" style="margin-top:24px;">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px;flex-wrap:wrap;gap:10px;">
            <span class="card-title" style="margin-bottom:0;">📋 Officer Management</span>
            <input type="text" id="officerSearch" placeholder="Search ID, LGA, ward, phone..." onkeyup="loadOfficerTable(1, undefined, this.value)"
                   style="background:rgba(255,255,255,0.06);border:1px solid rgba(255,255,255,0.12);border-radius:8px;color:#fff;padding:6px 12px;font-size:0.78rem;width:240px;">
        </div>
        <div style="display:flex;gap:8px;margin-bottom:12px;">
            <button class="btn-filter active" id="fAll"          onclick="setOfficerFilter('all')">All</button>
            <button class="btn-filter"        id="fRegistered"   onclick="setOfficerFilter('registered')">✅ Registered</button>
            <button class="btn-filter"        id="fUnregistered" onclick="setOfficerFilter('unregistered')">⚠️ No Phone</button>
        </div>
        <div class="preview-wrap" style="max-height:460px;overflow-y:auto;">
            <table style="width:100%;border-collapse:collapse;">
                <thead>
                    <tr style="font-size:0.68rem;color:rgba(255,255,255,0.4);text-transform:uppercase;letter-spacing:0.08em;border-bottom:1px solid rgba(255,255,255,0.08);">
                        <th style="padding:8px 6px;text-align:left;">Officer ID</th>
                        <th style="padding:8px 6px;text-align:left;">LGA</th>
                        <th style="padding:8px 6px;text-align:left;">Ward</th>
                        <th style="padding:8px 6px;text-align:left;">Polling Unit</th>
                        <th style="padding:8px 6px;text-align:left;">Phone</th>
                        <th style="padding:8px 6px;text-align:left;">Actions</th>
                    </tr>
                </thead>
                <tbody id="officerTableBody">
                    <tr><td colspan="6" style="text-align:center;padding:20px;color:rgba(255,255,255,0.3);">Login to manage officers.</td></tr>
                </tbody>
            </table>
        </div>
        <div id="officerTableInfo" style="margin-top:10px;font-size:0.72rem;color:rgba(255,255,255,0.4);display:flex;align-items:center;gap:10px;"></div>
    </div>

</div>

<script>
    let _adminKey = '';

    function togglePwd() {
        const i = document.getElementById('adminKey');
        i.type = i.type === 'password' ? 'text' : 'password';
    }

    async function adminLogin() {
        const key = document.getElementById('adminKey').value.trim();
        if (!key) return;
        const btn = document.getElementById('loginBtn');
        const err = document.getElementById('loginErr');
        btn.disabled = true; btn.textContent = 'Verifying...';
        err.style.display = 'none';
        // Test key against a protected endpoint
        try {
            const res = await fetch('/api/admin/set-officer-phone', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + key, 'Content-Type': 'application/json' },
                body: JSON.stringify({ officers: [] })
            });
            if (res.status === 401 || res.status === 403) {
                err.textContent = 'Invalid admin key.';
                err.style.display = 'block';
                btn.disabled = false; btn.textContent = 'Access Admin Portal →';
                return;
            }
            _adminKey = key;
            document.getElementById('loginGate').style.display = 'none';
            document.getElementById('adminPortal').style.display = 'block';
            loadStats();
        } catch(e) {
            err.textContent = 'Server error. Try again.';
            err.style.display = 'block';
            btn.disabled = false; btn.textContent = 'Access Admin Portal →';
        }
    }

    function adminLogout() {
        _adminKey = '';
        document.getElementById('adminPortal').style.display = 'none';
        document.getElementById('loginGate').style.display = 'flex';
        document.getElementById('adminKey').value = '';
        document.getElementById('loginBtn').disabled = false;
        document.getElementById('loginBtn').textContent = 'Access Admin Portal →';
    }

    async function _loadStatsCore() {
        try {
            const res = await fetch('/api/officers/stats', {
                headers: { 'Authorization': 'Bearer ' + _adminKey }
            });
            if (!res.ok) return;
            const d = await res.json();
            document.getElementById('statTotal').textContent     = d.total.toLocaleString();
            document.getElementById('statWithPhone').textContent = d.with_phone.toLocaleString();
            document.getElementById('statNoPhone').textContent   = d.without_phone.toLocaleString();
        } catch(e) {}
    }

    // ── CSV file handling ────────────────────────────────────────────────────
    let _parsedRows = [];

    function handleFileSelect(input) {
        const file = input.files[0];
        if (!file) return;
        document.getElementById('fileChosen').textContent = '📄 ' + file.name;
        document.getElementById('fileChosen').classList.remove('d-none');
        const reader = new FileReader();
        reader.onload = e => parseCSV(e.target.result);
        reader.readAsText(file);
    }

    // Drag-and-drop
    const zone = document.getElementById('uploadZone');
    zone.addEventListener('dragover', e => { e.preventDefault(); zone.classList.add('drag-over'); });
    zone.addEventListener('dragleave', () => zone.classList.remove('drag-over'));
    zone.addEventListener('drop', e => {
        e.preventDefault(); zone.classList.remove('drag-over');
        const file = e.dataTransfer.files[0];
        if (file) { document.getElementById('csvFile').files = e.dataTransfer.files; handleFileSelect({ files: [file] }); }
    });

    function parseCSV(text) {
        const lines = text.trim().split(/\\r?\\n/).filter(l => l.trim());
        if (!lines.length) return;
        const first = lines[0].toLowerCase();
        const hasHeader = first.includes('officer_id') || first.includes('phone') || first.includes('lga');
        const dataLines = hasHeader ? lines.slice(1) : lines;

        // Detect column positions from header
        let colOfficer = 0, colLga = 1, colWard = 2, colPu = 3, colPhone = 4;
        if (hasHeader) {
            const hCols = _splitCSVLine(lines[0].toLowerCase());
            hCols.forEach((h, i) => {
                if (h.includes('officer_id') || h === 'officer id') colOfficer = i;
                else if (h === 'lga')                               colLga     = i;
                else if (h === 'ward')                              colWard    = i;
                else if (h.includes('polling') || h.includes('unit')) colPu    = i;
                else if (h === 'phone')                             colPhone   = i;
            });
        }

        _parsedRows = [];
        dataLines.forEach((line, i) => {
            const parts      = _splitCSVLine(line);
            const officer_id = (parts[colOfficer] || '').trim();
            const lga        = (parts[colLga]     || '').trim();
            const ward       = (parts[colWard]    || '').trim();
            const pu         = (parts[colPu]      || '').trim();
            const phone      = (parts[colPhone]   || '').trim();
            const hasPhone   = phone.length >= 10;
            const valid      = officer_id.includes('-') && hasPhone;
            const noPhone    = officer_id.includes('-') && !hasPhone;
            _parsedRows.push({ officer_id, lga, ward, pu, phone, valid, noPhone, row: i + (hasHeader ? 2 : 1) });
        });
        renderPreview();
        document.getElementById('uploadBtn').disabled = _parsedRows.filter(r => r.valid).length === 0;
    }

    // CSV-aware splitter — handles quoted fields containing commas
    function _splitCSVLine(line) {
        const result = [];
        let cur = '', inQuote = false;
        for (let i = 0; i < line.length; i++) {
            const ch = line[i];
            if (ch === '"') { inQuote = !inQuote; }
            else if (ch === ',' && !inQuote) { result.push(cur); cur = ''; }
            else { cur += ch; }
        }
        result.push(cur);
        return result.map(s => s.replace(/^"|"$/g, '').trim());
    }

    function renderPreview() {
        const section  = document.getElementById('previewSection');
        const label    = document.getElementById('previewLabel');
        const valid    = _parsedRows.filter(r => r.valid).length;
        const noPhone  = _parsedRows.filter(r => r.noPhone).length;
        const invalid  = _parsedRows.filter(r => !r.valid && !r.noPhone).length;
        label.innerHTML =
            `<span style="color:#00cc66;">${valid} ready to upload</span>` +
            (noPhone  ? ` &nbsp;·&nbsp; <span style="color:#ffc107;">${noPhone} missing phone (will skip)</span>` : '') +
            (invalid  ? ` &nbsp;·&nbsp; <span style="color:#ff6b6b;">${invalid} invalid rows</span>` : '') +
            ` &nbsp;·&nbsp; ${_parsedRows.length} total`;
        section.style.display = 'block';

        const thead = document.querySelector('#previewTable thead');
        const tbody = document.querySelector('#previewTable tbody');
        thead.innerHTML = '<tr><th>#</th><th>Officer ID</th><th>LGA</th><th>Ward</th><th>Polling Unit</th><th>Phone</th><th>Status</th></tr>';
        tbody.innerHTML = _parsedRows.slice(0, 100).map(r => {
            const statusIcon = r.valid ? '✅' : r.noPhone ? '⏭ skip' : '⚠️';
            const rowClass   = r.valid ? 'valid' : r.noPhone ? '' : 'invalid';
            return `<tr>
                <td class="row-num">${r.row}</td>
                <td class="${rowClass}" style="font-weight:700;">${r.officer_id || '<em>—</em>'}</td>
                <td style="color:rgba(255,255,255,0.6);font-size:0.72rem;">${r.lga || '—'}</td>
                <td style="color:rgba(255,255,255,0.5);font-size:0.7rem;">${r.ward || '—'}</td>
                <td style="color:rgba(255,255,255,0.45);font-size:0.68rem;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${r.pu}">${r.pu || '—'}</td>
                <td class="${r.valid ? 'valid' : r.noPhone ? '' : 'invalid'}">${r.phone || '<em style="color:#555;">blank</em>'}</td>
                <td>${statusIcon}</td>
            </tr>`;
        }).join('') + (_parsedRows.length > 100 ? `<tr><td colspan="7" style="color:rgba(255,255,255,0.3);text-align:center;padding:10px;">... and ${_parsedRows.length - 100} more rows</td></tr>` : '');
    }

    function clearFile() {
        _parsedRows = [];
        document.getElementById('csvFile').value = '';
        document.getElementById('fileChosen').classList.add('d-none');
        document.getElementById('previewSection').style.display = 'none';
        document.getElementById('uploadBtn').disabled = true;
        document.getElementById('uploadResult').style.display = 'none';
    }

    async function submitCSV() {
        const valid = _parsedRows.filter(r => r.valid);
        if (!valid.length) return;
        const btn = document.getElementById('uploadBtn');
        const res_el = document.getElementById('uploadResult');
        btn.disabled = true; btn.textContent = `Uploading ${valid.length} officers...`;
        res_el.style.display = 'none';
        await _submitOfficers(valid, res_el);
        btn.disabled = false; btn.textContent = 'Upload & Register Officers';
        loadStats();
    }

    // ── Manual entry ─────────────────────────────────────────────────────────
    async function submitManual() {
        const raw = document.getElementById('manualInput').value.trim();
        if (!raw) return;
        const btn    = document.getElementById('manualBtn');
        const res_el = document.getElementById('manualResult');
        const lines  = raw.split(/\\n/).filter(l => l.trim());
        const officers = lines.map(l => {
            const p = _splitCSVLine(l);
            if (p.length >= 3) {
                return { officer_id: (p[0]||'').trim(), lga: (p[1]||'').trim(), ward: '', pu: '', phone: (p[2]||'').trim() };
            } else {
                return { officer_id: (p[0]||'').trim(), lga: '', ward: '', pu: '', phone: (p[1]||'').trim() };
            }
        }).filter(o => o.officer_id && o.phone);
        if (!officers.length) {
            showBanner(res_el, 'No valid entries. Format: officer_id,lga,phone  e.g.  10-001,Osogbo,+2348012345678', 'error');
            return;
        }
        btn.disabled = true; btn.textContent = `Registering ${officers.length} officers...`;
        res_el.style.display = 'none';
        await _submitOfficers(officers, res_el);
        btn.disabled = false; btn.textContent = 'Register These Officers';
        loadStats();
    }

    // ── Shared submit: builds a CSV in memory and posts it to the Supabase-backed
    //    /api/upload-officers endpoint, which is what actually persists officers.
    async function _submitOfficers(officers, res_el) {
        try {
            const header = 'officer_id,lga,ward,polling_unit,phone';
            const lines  = officers.map(o =>
                [o.officer_id, o.lga || '', o.ward || '', o.pu || '', o.phone]
                    .map(v => `"${String(v).replace(/"/g, '""')}"`).join(',')
            );
            const csvText = [header, ...lines].join('\\n');
            const blob = new Blob([csvText], { type: 'text/csv' });

            const formData = new FormData();
            formData.append('file', blob, 'officers_upload.csv');

            const res = await fetch('/api/upload-officers', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + _adminKey },
                body: formData
            });
            const out = await res.json();
            if (!res.ok) {
                showBanner(res_el, out.detail || 'Upload failed.', 'error');
                return;
            }
            let msg = `✅ ${out.saved} officer${out.saved !== 1 ? 's' : ''} saved permanently to Supabase`;
            if (out.skipped) {
                msg += ` · ${out.skipped} skipped`;
                if (out.skipped_details && out.skipped_details.length) {
                    msg += `<br><span style="color:#ffc107;font-size:0.78rem;">⚠️ ${out.skipped_details.slice(0,5).join('; ')}${out.skipped_details.length > 5 ? ' ...' : ''}</span>`;
                }
            }
            showBanner(res_el, msg, out.saved > 0 ? 'success' : 'info');
        } catch(e) {
            showBanner(res_el, 'Server error. Try again.', 'error');
        }
    }

    function showBanner(el, msg, type) {
        el.textContent = msg;
        el.className = 'result-banner ' + type;
        el.style.display = 'block';
    }

    async function downloadTemplate(e) {
        e.preventDefault();
        const res = await fetch('/admin/template.csv', {
            headers: { 'Authorization': 'Bearer ' + _adminKey }
        });
        if (!res.ok) { alert('Could not download template'); return; }
        const blob = await res.blob();
        const url  = URL.createObjectURL(blob);
        const a    = document.createElement('a');
        a.href = url; a.download = 'officer_phones_template.csv'; a.click();
        URL.revokeObjectURL(url);
    }

    document.addEventListener('DOMContentLoaded', () => {
        document.getElementById('adminKey').focus();
    });

    // ── Officer Management Table ─────────────────────────────────────────────
    let _officerPage = 1, _officerQ = '', _officerFilter = 'all';

    function setOfficerFilter(f) {
        _officerFilter = f;
        ['All','Registered','Unregistered'].forEach(n => {
            const el = document.getElementById('f' + n);
            if (el) el.classList.toggle('active', n.toLowerCase() === f || (f === 'all' && n === 'All'));
        });
        loadOfficerTable(1);
    }

    async function loadOfficerTable(page, filter, q) {
        _officerPage   = page  !== undefined ? page  : _officerPage;
        _officerFilter = filter !== undefined ? filter : _officerFilter;
        _officerQ      = q    !== undefined ? q    : _officerQ;
        const tbl  = document.getElementById('officerTableBody');
        const info = document.getElementById('officerTableInfo');
        if (!tbl) return;
        tbl.innerHTML = '<tr><td colspan="6" style="text-align:center;padding:20px;color:rgba(255,255,255,0.3);">Loading...</td></tr>';
        try {
            const url = `/api/officers?page=${_officerPage}&search=${encodeURIComponent(_officerQ)}&filter=${_officerFilter}`;
            const res = await fetch(url, { headers: { 'Authorization': 'Bearer ' + _adminKey } });
            const d   = await res.json();
            if (!d.officers || !d.officers.length) {
                tbl.innerHTML = '<tr><td colspan="6" style="text-align:center;padding:20px;color:rgba(255,255,255,0.3);">No officers found.</td></tr>';
                if (info) info.innerHTML = '';
                return;
            }
            tbl.innerHTML = d.officers.map(o => {
                const hasPhone = o.phone && o.phone.trim();
                const phoneDisplay = hasPhone
                    ? `<span style="color:#00cc66;font-weight:600;">${o.phone}</span>`
                    : `<span class="phone-missing">— no phone</span>`;
                const jsId = JSON.stringify(o.officer_id);
                const jsPhone = JSON.stringify(o.phone || '');
                const actions = `
                    <button class="btn-action btn-edit" onclick='editOfficerRow(${jsId},${jsPhone})'>✏️ Edit</button>
                    ${hasPhone ? `<button class="btn-action btn-del" onclick='deleteOfficer(${jsId})'>🗑 Remove</button>` : ''}`;
                return `<tr id="orow-${o.officer_id}">
                    <td style="font-weight:700;color:#ffc107;white-space:nowrap;">${o.officer_id}</td>
                    <td style="font-size:0.74rem;color:rgba(255,255,255,0.6);">${o.lga||'—'}</td>
                    <td style="font-size:0.7rem;color:rgba(255,255,255,0.5);">${o.ward||'—'}</td>
                    <td style="font-size:0.68rem;color:rgba(255,255,255,0.4);max-width:130px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;" title="${o.polling_unit||''}">${o.polling_unit||'—'}</td>
                    <td id="phone-${o.officer_id}">${phoneDisplay}</td>
                    <td style="white-space:nowrap;">${actions}</td>
                </tr>`;
            }).join('');
            if (info) info.innerHTML =
                `<span>Page ${d.page} of ${d.pages} &nbsp;·&nbsp; ${d.total.toLocaleString()} officers</span>` +
                (d.pages > 1 ? `<button class="btn-action" onclick="loadOfficerTable(${_officerPage-1})" ${_officerPage<=1?'disabled':''}>◀</button>` +
                               `<button class="btn-action" onclick="loadOfficerTable(${_officerPage+1})" ${_officerPage>=d.pages?'disabled':''}>▶</button>` : '');
        } catch(e) {
            tbl.innerHTML = '<tr><td colspan="6" style="color:#ff6b6b;text-align:center;padding:16px;">Failed to load officers.</td></tr>';
        }
    }

    function editOfficerRow(officerId, currentPhone) {
        const phoneCell  = document.getElementById(`phone-${officerId}`);
        const row        = document.getElementById(`orow-${officerId}`);
        const actionCell = row.querySelector('td:last-child');
        const jsId = JSON.stringify(officerId);
        phoneCell.innerHTML = `<input type="text" id="ep-${officerId}" value="${currentPhone.replace(/"/g,'&quot;')}"
            placeholder="+2348012345678"
            style="background:rgba(255,255,255,0.08);border:1px solid rgba(0,135,81,0.5);border-radius:6px;color:#fff;padding:4px 8px;width:155px;font-size:0.78rem;"
            onkeydown="if(event.key==='Enter')saveOfficerEdit(${jsId});if(event.key==='Escape')loadOfficerTable();">`;
        actionCell.innerHTML = `
            <button class="btn-action btn-save"   onclick='saveOfficerEdit(${jsId})'>💾 Save</button>
            <button class="btn-action btn-cancel" onclick="loadOfficerTable()">✕</button>`;
        document.getElementById(`ep-${officerId}`).focus();
    }

    async function saveOfficerEdit(officerId) {
        const phone = document.getElementById(`ep-${officerId}`).value.trim();
        try {
            const res = await fetch('/api/officers/update', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + _adminKey, 'Content-Type': 'application/json' },
                body: JSON.stringify({ officer_id: officerId, phone })
            });
            const out = await res.json();
            if (!res.ok) { alert(out.detail || 'Update failed'); return; }
            loadOfficerTable();
            loadStats();
        } catch(e) { alert('Server error'); }
    }

    async function deleteOfficer(officerId) {
        if (!confirm(`Remove officer ${officerId}? They will not be able to log in via OTP.`)) return;
        try {
            const res = await fetch('/api/officers/delete', {
                method: 'POST',
                headers: { 'Authorization': 'Bearer ' + _adminKey, 'Content-Type': 'application/json' },
                body: JSON.stringify({ officer_id: officerId })
            });
            const out = await res.json();
            if (!res.ok) { alert(out.detail || 'Delete failed'); return; }
            loadOfficerTable();
            loadStats();
        } catch(e) { alert('Server error'); }
    }

    // loadStats also refreshes the officer table after login
    async function loadStats() {
        await _loadStatsCore();
        if (_adminKey) loadOfficerTable(1, 'all', '');
    }
</script>
</body>
</html>
"""

# ── Admin portal ─────────────────────────────────────────────────────────────

# ── Admin: list registered officers (with phone) ──────────────────────────────
@app.get("/api/admin/list-officers")
async def list_officers(request: Request, page: int = 1, q: str = "", filter: str = "all"):
    """
    filter: all | registered | unregistered
    """
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    PAGE_SIZE = 50
    offset = (page - 1) * PAGE_SIZE
    conditions = ["LOWER(state)='osun'"]
    params_q = []
    if filter == "registered":
        conditions.append("officer_phone IS NOT NULL AND officer_phone != ''")
    elif filter == "unregistered":
        conditions.append("(officer_phone IS NULL OR officer_phone = '')")
    if q:
        like = f"%{q}%"
        conditions.append("(ward_code LIKE ? OR pu_code LIKE ? OR LOWER(lg) LIKE ? OR LOWER(ward) LIKE ? OR officer_phone LIKE ?)")
        params_q = [like, like, like.lower(), like.lower(), like]
    where = "WHERE " + " AND ".join(conditions)
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) as c FROM polling_units {where}", params_q)
            total = cur.fetchone()["c"]
            cur.execute(f"""SELECT ward_code, pu_code, lg, ward, location, officer_phone
                           FROM polling_units {where}
                           ORDER BY lg, ward_code, pu_code LIMIT ? OFFSET ?""",
                        params_q + [PAGE_SIZE, offset])
            rows = cur.fetchall()
    return {
        "officers": [dict(r) for r in rows],
        "total": total,
        "page": page,
        "pages": max(1, -(-total // PAGE_SIZE))
    }


# ── Admin: update a single officer phone ─────────────────────────────────────
@app.put("/api/admin/update-officer")
async def update_officer(request: Request):
    """Body: { "ward_code": "10", "pu_code": "001", "phone": "+2348012345678" }"""
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    ward_code = str(body.get("ward_code", "")).strip()
    pu_code   = str(body.get("pu_code", "")).strip()
    phone     = _clean_phone(str(body.get("phone", "")).strip())
    if not ward_code or not pu_code:
        raise HTTPException(status_code=400, detail="ward_code and pu_code required")
    if phone and len(phone) < 10:
        raise HTTPException(status_code=400, detail="Invalid phone number")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""UPDATE polling_units SET officer_phone = ?
                           WHERE ward_code = ? AND pu_code = ? AND LOWER(state)='osun'""",
                        (phone or None, ward_code, pu_code))
            changed = conn._conn.total_changes
        conn.commit()
    if not changed:
        raise HTTPException(status_code=404, detail="Officer not found")
    return {"status": "ok", "phone": phone}


# ── Admin: delete officer phone (clear it) ───────────────────────────────────
@app.delete("/api/admin/delete-officer")
async def delete_officer(request: Request):
    """Body: { "ward_code": "10", "pu_code": "001" }"""
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")
    ward_code = str(body.get("ward_code", "")).strip()
    pu_code   = str(body.get("pu_code", "")).strip()
    if not ward_code or not pu_code:
        raise HTTPException(status_code=400, detail="ward_code and pu_code required")
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""UPDATE polling_units SET officer_phone = NULL
                           WHERE ward_code = ? AND pu_code = ? AND LOWER(state)='osun'""",
                        (ward_code, pu_code))
            changed = conn._conn.total_changes
        conn.commit()
    if not changed:
        raise HTTPException(status_code=404, detail="Officer not found")
    return {"status": "ok"}


# ── Admin: list all submitted results (paginated) ────────────────────────────
@app.get("/api/admin/results")
async def admin_list_results(request: Request, page: int = 1, lga: str = "", status: str = ""):
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    PAGE_SIZE = 50
    offset = (page - 1) * PAGE_SIZE
    filters, params = [], []
    if lga:
        filters.append("LOWER(lg) = LOWER(%s)")
        params.append(lga)
    if status == "reviewed":
        filters.append("reviewed = 1")
    elif status == "pending":
        filters.append("(reviewed IS NULL OR reviewed = 0)")
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT COUNT(*) as c FROM field_submissions {where}", params)
            total = cur.fetchone()["c"]
            cur.execute(f"""SELECT id, officer_id, state, lg, ward, ward_code, pu_code, location,
                                   reg_voters, total_accredited, valid_votes, rejected_votes, total_cast,
                                   votes_json, edited_votes_json, ec8e_image, timestamp,
                                   reviewed, reviewed_by, reviewed_at, edit_note
                            FROM field_submissions {where}
                            ORDER BY timestamp DESC LIMIT %s OFFSET %s""",
                        params + [PAGE_SIZE, offset])
            rows = cur.fetchall()
    results = []
    for r in rows:
        votes = json.loads(r.get("edited_votes_json") or r.get("votes_json") or "{}")
        results.append({
            "id": r["id"],
            "officer_id": r["officer_id"],
            "lg": r["lg"],
            "ward": r["ward"],
            "location": r["location"],
            "pu_code": r["pu_code"],
            "reg_voters": r["reg_voters"] or 0,
            "total_accredited": r["total_accredited"] or 0,
            "valid_votes": r["valid_votes"] or 0,
            "rejected_votes": r["rejected_votes"] or 0,
            "total_cast": r["total_cast"] or 0,
            "votes": votes,
            "ec8e_image": r["ec8e_image"],
            "timestamp": r["timestamp"],
            "reviewed": bool(r["reviewed"]),
            "reviewed_by": r["reviewed_by"],
            "reviewed_at": r["reviewed_at"],
            "edit_note": r["edit_note"],
        })
    return {"results": results, "total": total, "page": page, "pages": max(1, -(-total // PAGE_SIZE))}


# ── Admin: edit a submitted result ───────────────────────────────────────────
@app.put("/api/admin/results/{submission_id}")
async def admin_edit_result(submission_id: int, request: Request):
    """
    Editable fields: votes (dict), total_accredited, total_cast,
                     valid_votes, rejected_votes, reg_voters, edit_note
    """
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    now = datetime.now().isoformat()
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM field_submissions WHERE id=%s", (submission_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Submission not found")

            updates, vals, audit_entries = [], [], []
            old_votes = json.loads(row["edited_votes_json"] or row["votes_json"] or "{}")

            if "votes" in body:
                new_votes = {k: int(v) for k, v in body["votes"].items()}
                updates.append("edited_votes_json = %s")
                vals.append(json.dumps(new_votes))
                audit_entries.append(("edit", "votes", json.dumps(old_votes), json.dumps(new_votes)))

            for field in ["total_accredited", "total_cast", "valid_votes", "rejected_votes", "reg_voters"]:
                if field in body:
                    old_val = row[field]
                    new_val = int(body[field])
                    updates.append(f"{field} = %s")
                    vals.append(new_val)
                    audit_entries.append(("edit", field, str(old_val), str(new_val)))

            if "edit_note" in body:
                updates.append("edit_note = %s")
                vals.append(str(body["edit_note"])[:500])

            if not updates:
                raise HTTPException(status_code=400, detail="No editable fields provided")

            vals.append(submission_id)
            cur.execute(f"UPDATE field_submissions SET {', '.join(updates)} WHERE id=%s", vals)

            for action, field, old_v, new_v in audit_entries:
                cur.execute("""INSERT INTO result_audit_log
                               (submission_id, action, field, old_value, new_value, changed_by, changed_at)
                               VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                            (submission_id, action, field, old_v, new_v, "admin", now))
        conn.commit()
    return {"status": "ok", "id": submission_id}


# ── Admin: approve / unapprove a result ──────────────────────────────────────
@app.post("/api/admin/results/{submission_id}/approve")
async def admin_approve_result(submission_id: int, request: Request):
    auth = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )) and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    try:
        body = await request.json()
    except Exception:
        body = {}
    approve = bool(body.get("approve", True))
    now = datetime.now().isoformat()
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT reviewed FROM field_submissions WHERE id=%s", (submission_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Submission not found")
            cur.execute("""UPDATE field_submissions
                           SET reviewed=%s, reviewed_by='admin', reviewed_at=%s
                           WHERE id=%s""",
                        (1 if approve else 0, now, submission_id))
            cur.execute("""INSERT INTO result_audit_log
                           (submission_id, action, field, old_value, new_value, changed_by, changed_at)
                           VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                        (submission_id,
                         "approve" if approve else "unapprove",
                         "reviewed",
                         str(bool(row["reviewed"])),
                         str(approve),
                         "admin", now))
        conn.commit()
    return {"status": "ok", "reviewed": approve}


# ── Admin: pending review count (for dashboard badge) ────────────────────────
@app.get("/api/admin/pending-review-count")
async def pending_review_count(request: Request):
    auth  = request.headers.get("Authorization", "")
    token = request.cookies.get("ds_session")
    bearer_ok = auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(), _DASHBOARD_KEY_HASH
    )
    if not bearer_ok and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as c FROM field_submissions WHERE reviewed=0 OR reviewed IS NULL")
            pending = cur.fetchone()["c"]
            cur.execute("SELECT COUNT(*) as c FROM field_submissions")
            total = cur.fetchone()["c"]
    return {"pending": pending, "total": total}


# ── Admin: results portal page ───────────────────────────────────────────────
@app.get("/admin/results", response_class=HTMLResponse)
async def admin_results_page(request: Request):
    return HTMLResponse(content=RESULTS_PORTAL_HTML)


@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    """Protected admin portal — requires DASHBOARD_KEY as Bearer token via
    a login form (same key as dashboard). Served as a standalone HTML page."""
    return HTMLResponse(content=ADMIN_HTML)


@app.get("/admin/template.csv")
async def download_template(request: Request):
    """Download the officer phone CSV template."""
    auth = request.headers.get("Authorization", "")
    # Also allow cookie-authenticated dashboard users
    token = request.cookies.get("ds_session")
    header_ok = auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(),
        _DASHBOARD_KEY_HASH
    )
    if not header_ok and not _is_valid_token(token):
        raise HTTPException(status_code=403, detail="Not authorised")
    csv_content = (
        "officer_id,phone\n"
        "WARD001-PU001,+2348012345678\n"
        "WARD001-PU002,+2348023456789\n"
        "WARD002-PU001,+2348034567890\n"
    )
    from fastapi.responses import Response as _R
    return _R(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=officer_phones_template.csv"}
    )
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def homepage():
    return HTMLResponse(content=HOMEPAGE_HTML)

HOMEPAGE_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <meta name="theme-color" content="#008751">
    <title>TKC — Osun 2026 Election Observation System</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700;900&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

        :root {
            --green: #008751;
            --green-light: #00b368;
            --gold: #ffc107;
            --dark: #020c06;
            --panel: rgba(255,255,255,0.04);
            --border: rgba(255,255,255,0.08);
        }

        body {
            font-family: 'Inter', sans-serif;
            background: var(--dark);
            color: #fff;
            min-height: 100vh;
            overflow-x: hidden;
        }

        /* ── Multi-layer background ── */
        body::before {
            content: '';
            position: fixed;
            inset: 0;
            z-index: 0;
            background:
                radial-gradient(ellipse 120% 80% at 50% -10%, rgba(0,135,81,0.18) 0%, transparent 60%),
                radial-gradient(ellipse 80% 60% at 80% 100%, rgba(0,80,40,0.15) 0%, transparent 55%),
                radial-gradient(ellipse 60% 50% at 10% 80%, rgba(255,193,7,0.04) 0%, transparent 50%),
                linear-gradient(160deg, #020c06 0%, #041508 40%, #020c06 100%);
            pointer-events: none;
        }

        /* ── Animated canvas background ── */
        #bg-canvas {
            position: fixed;
            inset: 0;
            z-index: 1;
            opacity: 0.65;
        }

        /* ── Precision grid overlay ── */
        .grid-overlay {
            position: fixed;
            inset: 0;
            z-index: 2;
            background-image:
                linear-gradient(rgba(0,135,81,0.06) 1px, transparent 1px),
                linear-gradient(90deg, rgba(0,135,81,0.06) 1px, transparent 1px),
                linear-gradient(rgba(0,135,81,0.02) 1px, transparent 1px),
                linear-gradient(90deg, rgba(0,135,81,0.02) 1px, transparent 1px);
            background-size: 80px 80px, 80px 80px, 20px 20px, 20px 20px;
            pointer-events: none;
        }

        /* ── Corner accent lines ── */
        .grid-overlay::before {
            content: '';
            position: absolute;
            inset: 0;
            background:
                linear-gradient(135deg, rgba(0,135,81,0.12) 0%, transparent 30%),
                linear-gradient(315deg, rgba(255,193,7,0.05) 0%, transparent 30%);
            pointer-events: none;
        }

        /* ── Radial glow ── */
        .glow-center {
            position: fixed;
            top: 50%; left: 50%;
            transform: translate(-50%, -50%);
            width: 1100px; height: 1100px;
            background: radial-gradient(circle, rgba(0,135,81,0.14) 0%, rgba(0,60,30,0.06) 40%, transparent 70%);
            z-index: 2;
            pointer-events: none;
        }

        /* ── Top scan line ── */
        .scan-line {
            position: fixed;
            top: 0; left: 0; right: 0;
            height: 2px;
            background: linear-gradient(90deg, transparent, rgba(0,135,81,0.6), rgba(255,193,7,0.4), rgba(0,135,81,0.6), transparent);
            z-index: 3;
            animation: scan 4s ease-in-out infinite;
        }

        @keyframes scan {
            0%, 100% { opacity: 0.4; transform: scaleX(0.8); }
            50% { opacity: 1; transform: scaleX(1); }
        }

        /* ── Main layout ── */
        .wrapper {
            position: relative;
            z-index: 2;
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            padding: 40px 20px;
        }

        /* ── Header ── */
        .header {
            text-align: center;
            margin-bottom: 56px;
            animation: fadeUp 0.8s ease both;
        }

        .logo-ring {
            width: 110px; height: 110px;
            border-radius: 50%;
            border: 2px solid rgba(0,135,81,0.5);
            display: flex; align-items: center; justify-content: center;
            margin: 0 auto 24px;
            background: rgba(0,135,81,0.08);
            box-shadow: 0 0 40px rgba(0,135,81,0.25), inset 0 0 20px rgba(0,135,81,0.05);
            position: relative;
            animation: pulse-ring 3s ease-in-out infinite;
        }

        .logo-ring::before {
            content: '';
            position: absolute;
            inset: -8px;
            border-radius: 50%;
            border: 1px solid rgba(0,135,81,0.2);
            animation: spin-slow 12s linear infinite;
        }

        .logo-ring::after {
            content: '';
            position: absolute;
            inset: -16px;
            border-radius: 50%;
            border: 1px dashed rgba(255,193,7,0.15);
            animation: spin-slow 20s linear infinite reverse;
        }

        .logo-ring img {
            width: 72px; height: 72px;
            object-fit: contain;
            filter: drop-shadow(0 0 12px rgba(0,135,81,0.6));
        }

        .system-tag {
            display: inline-block;
            font-size: 0.65rem;
            font-weight: 700;
            letter-spacing: 0.2em;
            color: var(--gold);
            background: rgba(255,193,7,0.1);
            border: 1px solid rgba(255,193,7,0.25);
            border-radius: 20px;
            padding: 4px 14px;
            margin-bottom: 16px;
            text-transform: uppercase;
        }

        h1 {
            font-size: clamp(1.8rem, 5vw, 3rem);
            font-weight: 900;
            line-height: 1.1;
            letter-spacing: -0.02em;
            background: linear-gradient(135deg, #fff 0%, rgba(255,255,255,0.7) 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 12px;
        }

        h1 span {
            background: linear-gradient(135deg, var(--green-light), var(--gold));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .subtitle {
            font-size: 0.95rem;
            color: rgba(255,255,255,0.45);
            font-weight: 400;
            max-width: 480px;
            margin: 0 auto;
            line-height: 1.6;
        }

        /* ── Stats bar ── */
        .stats-bar {
            display: flex;
            gap: 32px;
            justify-content: center;
            flex-wrap: wrap;
            margin-bottom: 56px;
            animation: fadeUp 0.8s 0.15s ease both;
        }

        .stat {
            text-align: center;
        }

        .stat-val {
            font-size: 1.5rem;
            font-weight: 900;
            color: var(--green-light);
            display: block;
            line-height: 1;
        }

        .stat-label {
            font-size: 0.65rem;
            color: rgba(255,255,255,0.3);
            text-transform: uppercase;
            letter-spacing: 0.1em;
            margin-top: 4px;
        }

        .stat-divider {
            width: 1px;
            background: var(--border);
            align-self: stretch;
        }

        /* ── Cards grid ── */
        .cards-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 16px;
            width: 100%;
            max-width: 780px;
            animation: fadeUp 0.8s 0.25s ease both;
        }

        .nav-card {
            position: relative;
            background: var(--panel);
            border: 1px solid var(--border);
            border-radius: 20px;
            padding: 28px 24px;
            cursor: pointer;
            text-decoration: none;
            color: #fff;
            display: flex;
            flex-direction: column;
            gap: 12px;
            overflow: hidden;
            transition: transform 0.25s ease, border-color 0.25s ease, box-shadow 0.25s ease;
            backdrop-filter: blur(12px);
        }

        .nav-card::before {
            content: '';
            position: absolute;
            inset: 0;
            border-radius: 20px;
            opacity: 0;
            transition: opacity 0.25s ease;
        }

        .nav-card:hover {
            transform: translateY(-4px);
            box-shadow: 0 20px 60px rgba(0,0,0,0.4);
        }

        .nav-card:hover::before { opacity: 1; }

        /* Card colour themes */
        .card-vote {
            border-color: rgba(0,135,81,0.3);
        }
        .card-vote:hover {
            border-color: rgba(0,135,81,0.7);
            box-shadow: 0 20px 60px rgba(0,135,81,0.15);
        }
        .card-vote::before {
            background: linear-gradient(135deg, rgba(0,135,81,0.08), transparent);
        }

        .card-report {
            border-color: rgba(220,53,69,0.3);
        }
        .card-report:hover {
            border-color: rgba(220,53,69,0.7);
            box-shadow: 0 20px 60px rgba(220,53,69,0.15);
        }
        .card-report::before {
            background: linear-gradient(135deg, rgba(220,53,69,0.08), transparent);
        }

        .card-vdash {
            border-color: rgba(255,193,7,0.3);
        }
        .card-vdash:hover {
            border-color: rgba(255,193,7,0.7);
            box-shadow: 0 20px 60px rgba(255,193,7,0.12);
        }
        .card-vdash::before {
            background: linear-gradient(135deg, rgba(255,193,7,0.06), transparent);
        }

        .card-idash {
            border-color: rgba(13,202,240,0.3);
        }
        .card-idash:hover {
            border-color: rgba(13,202,240,0.7);
            box-shadow: 0 20px 60px rgba(13,202,240,0.12);
        }
        .card-idash::before {
            background: linear-gradient(135deg, rgba(13,202,240,0.06), transparent);
        }

        .card-icon {
            width: 48px; height: 48px;
            border-radius: 14px;
            display: flex; align-items: center; justify-content: center;
            font-size: 1.4rem;
            flex-shrink: 0;
        }

        .icon-vote   { background: rgba(0,135,81,0.15);  }
        .icon-report { background: rgba(220,53,69,0.15); }
        .icon-vdash  { background: rgba(255,193,7,0.12); }
        .icon-idash  { background: rgba(13,202,240,0.12);}

        .card-body { flex: 1; }

        .card-title {
            font-size: 1rem;
            font-weight: 700;
            margin-bottom: 6px;
            display: flex; align-items: center; gap: 8px;
        }

        .card-badge {
            font-size: 0.55rem;
            font-weight: 700;
            letter-spacing: 0.08em;
            padding: 2px 8px;
            border-radius: 20px;
            text-transform: uppercase;
        }

        .badge-field    { background: rgba(0,135,81,0.2);  color: #00cc66; }
        .badge-security { background: rgba(220,53,69,0.2); color: #ff6b6b; }
        .badge-live     { background: rgba(255,193,7,0.2); color: #ffc107; }
        .badge-command  { background: rgba(13,202,240,0.2);color: #0dcaf0; }

        .card-desc {
            font-size: 0.78rem;
            color: rgba(255,255,255,0.4);
            line-height: 1.5;
        }

        .card-arrow {
            position: absolute;
            bottom: 24px; right: 24px;
            font-size: 1.1rem;
            opacity: 0.3;
            transition: opacity 0.2s, transform 0.2s;
        }

        .nav-card:hover .card-arrow {
            opacity: 0.8;
            transform: translate(3px, -3px);
        }

        /* ── Footer ── */
        .footer {
            margin-top: 48px;
            text-align: center;
            font-size: 0.7rem;
            color: rgba(255,255,255,0.2);
            animation: fadeUp 0.8s 0.4s ease both;
        }

        .footer strong { color: rgba(255,255,255,0.4); }

        /* ── Animations ── */
        @keyframes fadeUp {
            from { opacity: 0; transform: translateY(20px); }
            to   { opacity: 1; transform: translateY(0); }
        }

        @keyframes pulse-ring {
            0%, 100% { box-shadow: 0 0 40px rgba(0,135,81,0.25), inset 0 0 20px rgba(0,135,81,0.05); }
            50%       { box-shadow: 0 0 60px rgba(0,135,81,0.4),  inset 0 0 30px rgba(0,135,81,0.1); }
        }

        @keyframes spin-slow {
            from { transform: rotate(0deg); }
            to   { transform: rotate(360deg); }
        }

        /* ── Mobile ── */
        @media (max-width: 600px) {
            .cards-grid { grid-template-columns: 1fr; max-width: 400px; }
            .stats-bar { gap: 20px; }
            .stat-divider { display: none; }
            h1 { font-size: 1.7rem; }
        }
    </style>
</head>
<body>

<!-- ══════════════════════════════════════════════════════
     TKC PREFACE SPLASH — shown once on first visit
     Press "Proceed" to dismiss and see the homepage
     ══════════════════════════════════════════════════════ -->
<div id="tkcSplash" style="
    position:fixed;inset:0;z-index:99999;
    background:linear-gradient(160deg,#0a1f0e 0%,#0d2b10 50%,#0a1f0e 100%);
    display:flex;flex-direction:column;align-items:center;justify-content:flex-start;
    overflow-y:auto;padding:40px 20px 60px;
">
  <!-- Gold top border -->
  <div style="position:fixed;top:0;left:0;right:0;height:5px;background:linear-gradient(90deg,#ffc107,#008751,#ffc107);z-index:100000;"></div>

  <!-- Logo -->
  <img src="/static/logos/tkc-logo.png"
       alt="The Kukah Centre (TKC) Logo"
       style="width:200px;height:200px;object-fit:contain;border-radius:16px;
              box-shadow:0 0 60px rgba(255,193,7,0.35);margin-bottom:28px;margin-top:16px;
              background:#fff;padding:12px;">

  <!-- Title block -->
  <div style="text-align:center;margin-bottom:28px;">
    <div style="font-size:0.6rem;font-weight:800;letter-spacing:0.25em;color:#ffc107;
                text-transform:uppercase;background:rgba(255,193,7,0.1);
                border:1px solid rgba(255,193,7,0.25);border-radius:20px;
                padding:4px 16px;display:inline-block;margin-bottom:14px;">
      The Kukah Centre (TKC) — Independent Election Observation
    </div>
    <h1 style="font-family:'Inter',sans-serif;font-size:clamp(1.3rem,4vw,2rem);
               font-weight:900;color:#fff;line-height:1.25;margin-bottom:6px;">
      A Message from <span style="color:#ffc107;">TKC</span>
    </h1>
    <p style="color:rgba(255,255,255,0.4);font-size:0.82rem;letter-spacing:0.04em;">
      Independent Field Observation System — 2026 Osun Governorship Election
    </p>
  </div>

  <!-- 4 Paragraphs -->
  <div style="max-width:680px;width:100%;">

    <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(0,135,81,0.25);
                border-left:4px solid #008751;border-radius:12px;padding:22px 24px;
                margin-bottom:16px;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0;">
        <strong style="color:#ffc107;">The Kukah Centre (TKC)</strong> is a non-governmental, non-partisan
        organisation committed to promoting faith-based leadership, sound public policy, and credible
        democratic processes across Nigeria. TKC is not affiliated with, and does not endorse, any
        political party or candidate. Our sole interest in the 2026 Osun State Governorship Election is
        an accurate, transparent, and peaceful process whose outcome reflects the genuine will of the
        people of Osun State.
      </p>
    </div>

    <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(0,135,81,0.25);
                border-left:4px solid #ffc107;border-radius:12px;padding:22px 24px;
                margin-bottom:16px;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0;">
        This platform exists to support that mission. It allows trained, accredited TKC observers stationed
        at polling units across all 30 Local Government Areas of Osun State to relay results as they are
        announced, so that the public collation process can be independently monitored and cross-checked
        in real time — strengthening confidence in the outcome, whichever way it falls.
      </p>
    </div>

    <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(0,135,81,0.25);
                border-left:4px solid #00cc66;border-radius:12px;padding:22px 24px;
                margin-bottom:16px;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0;">
        TKC observers are guided by strict codes of neutrality and professional conduct. We do not support
        any party, candidate, or political outcome. Our commitment is to the process — to free, fair, and
        credible elections, and to the peaceful expression of the will of the electorate.
      </p>
    </div>

    <div style="background:linear-gradient(135deg,rgba(0,135,81,0.15),rgba(255,193,7,0.08));
                border:1px solid rgba(255,193,7,0.3);border-radius:12px;padding:22px 24px;
                margin-bottom:32px;text-align:center;">
      <p style="font-family:'Inter',sans-serif;font-size:0.92rem;color:rgba(255,255,255,0.88);
                line-height:1.8;margin:0 0 12px 0;">
        TKC calls on every voter, every party agent, and every citizen of Osun State to embrace peace
        before, during, and after the election — and to trust a process that is watched, verified, and
        reported by independent eyes.
      </p>
      <div style="display:inline-block;background:rgba(255,193,7,0.15);border:1px solid rgba(255,193,7,0.4);
                  border-radius:20px;padding:6px 20px;font-size:0.72rem;font-weight:800;
                  letter-spacing:0.12em;color:#ffc107;text-transform:uppercase;">
        — The Kukah Centre (TKC), Independent Election Observation
      </div>
    </div>

    <!-- Proceed button -->
    <div style="text-align:center;">
      <button onclick="document.getElementById('tkcSplash').style.display='none'"
              style="background:linear-gradient(135deg,#008751,#00b368);
                     border:none;border-radius:14px;color:#fff;font-family:'Inter',sans-serif;
                     font-size:1rem;font-weight:800;padding:16px 52px;cursor:pointer;
                     box-shadow:0 0 40px rgba(0,135,81,0.4);letter-spacing:0.04em;
                     transition:opacity 0.2s;"
              onmouseover="this.style.opacity='0.88'"
              onmouseout="this.style.opacity='1'">
        ✅ &nbsp; Proceed to Observation System &nbsp; →
      </button>
      <p style="color:rgba(255,255,255,0.25);font-size:0.7rem;margin-top:12px;">
        Faith · Leadership · Public Policy — TKC 2026
      </p>
    </div>
  </div>

  <!-- Gold bottom border -->
  <div style="position:fixed;bottom:0;left:0;right:0;height:4px;background:linear-gradient(90deg,#ffc107,#008751,#ffc107);z-index:100000;"></div>
</div>

<script>
  // Splash now shows on every page load/refresh (no longer suppressed by sessionStorage)
</script>


<div class="scan-line"></div>
<canvas id="bg-canvas"></canvas>
<div class="grid-overlay"></div>
<div class="glow-center"></div>

<div class="wrapper">

    <!-- Header -->
    <div class="header">
        <div class="logo-ring">
            <img src="/static/logos/tkc-logo.png" alt="The Kukah Centre (TKC)" onerror="this.src='https://via.placeholder.com/72x72/008751/ffffff?text=TKC'">
        </div>
        <div class="system-tag">Osun 2026 Governorship Election · Independent Observation</div>
        <h1>TKC <span>Intelligence</span><br>Operations Centre</h1>
        <p class="subtitle">Real-time election collation, incident reporting and live analytics for the 2026 Osun Governorship Election.</p>
    </div>

    <!-- Stats bar -->
    <div class="stats-bar">
        <div class="stat">
            <span class="stat-val">3,763</span>
            <span class="stat-label">Polling Units</span>
        </div>
        <div class="stat-divider"></div>
        <div class="stat">
            <span class="stat-val">30</span>
            <span class="stat-label">LGAs</span>
        </div>
        <div class="stat-divider"></div>
        <div class="stat">
            <span class="stat-val">332</span>
            <span class="stat-label">Wards</span>
        </div>
        <div class="stat-divider"></div>
        <div class="stat">
            <span class="stat-val">14</span>
            <span class="stat-label">Parties</span>
        </div>
    </div>

    <!-- Navigation cards -->
    <div class="cards-grid">

        <a href="/vote" class="nav-card card-vote">
            <div class="card-icon icon-vote">🗳️</div>
            <div class="card-body">
                <div class="card-title">
                    Vote Submission
                    <span class="card-badge badge-field">Field</span>
                </div>
                <p class="card-desc">Field officers submit polling unit results in real-time. Secure officer ID validation with auto-filled PU data.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

        <a href="/report" class="nav-card card-report">
            <div class="card-icon icon-report">🚨</div>
            <div class="card-body">
                <div class="card-title">
                    Incident Report
                    <span class="card-badge badge-security">Security</span>
                </div>
                <p class="card-desc">Report election irregularities, security threats, or misconduct. Attach photo evidence and GPS location.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

        <a href="/dashboard" class="nav-card card-vdash">
            <div class="card-icon icon-vdash">📊</div>
            <div class="card-body">
                <div class="card-title">
                    Vote Dashboard
                    <span class="card-badge badge-live">Live</span>
                </div>
                <p class="card-desc">Live vote tallies, party breakdowns, LGA completion tracker, swing PU analysis and AI analytics log.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

        <a href="/incident-dashboard" class="nav-card card-idash">
            <div class="card-icon icon-idash">🛡️</div>
            <div class="card-body">
                <div class="card-title">
                    Incident Dashboard
                    <span class="card-badge badge-command">Command</span>
                </div>
                <p class="card-desc">Security command centre — live incident feed, severity heatmap, evidence viewer and real-time map.</p>
            </div>
            <span class="card-arrow">↗</span>
        </a>

    </div>

    <!-- Footer -->
    <div class="footer">
        <img src="/static/logos/popson-logo.png" style="height:18px;vertical-align:middle;margin-right:6px;opacity:0.4;" onerror="this.style.display='none'">
        Powered by <strong>Popson Geospatial Services</strong> &nbsp;·&nbsp; The Kukah Centre (TKC) — Independent Election Observation
    </div>

</div>

<script>
// ── Enhanced particle + ring canvas animation ────────────────────────────────
(function() {
    const canvas = document.getElementById('bg-canvas');
    const ctx = canvas.getContext('2d');
    let W, H, particles = [];
    let rings = [
        { r: 0, maxR: 0, alpha: 0, speed: 0.6 },
        { r: 0, maxR: 0, alpha: 0, speed: 0.4 },
        { r: 0, maxR: 0, alpha: 0, speed: 0.5 }
    ];
    let ringTimer = 0;
    const COUNT = 90;
    const GREEN = '0,135,81';
    const GOLD  = '255,193,7';
    const TEAL  = '0,179,104';

    function resize() {
        W = canvas.width  = window.innerWidth;
        H = canvas.height = window.innerHeight;
        rings.forEach((r, i) => { r.maxR = Math.max(W, H) * 0.7; });
    }

    function Particle() {
        this.reset = function() {
            this.x  = Math.random() * W;
            this.y  = Math.random() * H;
            this.vx = (Math.random() - 0.5) * 0.35;
            this.vy = (Math.random() - 0.5) * 0.35;
            this.r  = Math.random() * 2 + 0.3;
            const rnd = Math.random();
            this.color = rnd > 0.92 ? GOLD : rnd > 0.75 ? TEAL : GREEN;
            this.alpha = Math.random() * 0.55 + 0.15;
            this.pulse = Math.random() * Math.PI * 2;
        };
        this.reset();
    }

    function spawnRing() {
        const inactive = rings.find(r => r.alpha <= 0);
        if (!inactive) return;
        inactive.r = 0;
        inactive.alpha = 0.35;
        inactive.cx = W * 0.5 + (Math.random() - 0.5) * W * 0.3;
        inactive.cy = H * 0.5 + (Math.random() - 0.5) * H * 0.3;
    }

    function init() {
        particles = [];
        for (let i = 0; i < COUNT; i++) particles.push(new Particle());
    }

    function draw() {
        ctx.clearRect(0, 0, W, H);
        ringTimer++;
        if (ringTimer % 180 === 0) spawnRing();

        // Draw expanding rings
        rings.forEach(ring => {
            if (ring.alpha <= 0) return;
            ring.r += ring.speed;
            ring.alpha -= 0.001;
            if (ring.r > ring.maxR) { ring.alpha = 0; return; }
            ctx.beginPath();
            ctx.arc(ring.cx || W/2, ring.cy || H/2, ring.r, 0, Math.PI * 2);
            ctx.strokeStyle = `rgba(${GREEN},${ring.alpha * 0.5})`;
            ctx.lineWidth = 1;
            ctx.stroke();
        });

        // Connecting lines
        for (let i = 0; i < particles.length; i++) {
            for (let j = i + 1; j < particles.length; j++) {
                const dx = particles[i].x - particles[j].x;
                const dy = particles[i].y - particles[j].y;
                const dist = Math.sqrt(dx*dx + dy*dy);
                if (dist < 120) {
                    const alpha = (1 - dist / 120) * 0.15;
                    ctx.beginPath();
                    ctx.strokeStyle = `rgba(${GREEN},${alpha})`;
                    ctx.lineWidth = 0.5;
                    ctx.moveTo(particles[i].x, particles[i].y);
                    ctx.lineTo(particles[j].x, particles[j].y);
                    ctx.stroke();
                }
            }
        }

        // Particles with subtle pulse
        particles.forEach(p => {
            p.pulse += 0.02;
            const pr = p.r + Math.sin(p.pulse) * 0.4;
            ctx.beginPath();
            ctx.arc(p.x, p.y, pr, 0, Math.PI * 2);
            ctx.fillStyle = `rgba(${p.color},${p.alpha})`;
            ctx.fill();
            p.x += p.vx; p.y += p.vy;
            if (p.x < -10) p.x = W + 10;
            if (p.x > W + 10) p.x = -10;
            if (p.y < -10) p.y = H + 10;
            if (p.y > H + 10) p.y = -10;
        });

        requestAnimationFrame(draw);
    }

    window.addEventListener('resize', () => { resize(); init(); });
    resize();
    init();
    spawnRing();
    draw();
})();
</script>

</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/vote", response_class=HTMLResponse)
async def vote_form():
    parties = ["ACCORD", "AA", "AAC", "ADC", "ADP", "APGA", "APC", "APM", "APP", "BP", "NNPP", "PRP", "YPP", "ZLP"]
    party_cards = "".join([f'''
        <div class="col-4 col-md-2 mb-2">
            <div class="p-2 border rounded text-center bg-white shadow-sm">
                <img src="/logos/{p}.png" onerror="this.src='https://via.placeholder.com/30?text={p}'" style="height:30px">
                <small class="d-block fw-bold">{p}</small>
                <input type="number" class="form-control form-control-sm party-v text-center" data-p="{p}" value="0" inputmode="numeric" oninput="calculateTotals()">
            </div>
        </div>''' for p in parties])

    return f"""
<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <meta name="mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="theme-color" content="#008751">
    <title>THE KUKAH CENTRE (TKC)</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body {{ background: linear-gradient(160deg,#0a1f0e 0%,#0d2b10 50%,#0a1f0e 100%); background-attachment: fixed; min-height: 100vh; margin: 0; }}
        .navbar {{ background: rgba(0, 135, 81, 0.9) !important; backdrop-filter: blur(10px); color: white; border-bottom: 4px solid #ffc107; }}
        .card {{ background: rgba(255, 255, 255, 0.95) !important; border-radius: 12px; border: none; box-shadow: 0 10px 30px rgba(0,0,0,0.3) !important; margin-bottom: 20px; color: #222; }}
        .section-label {{ font-size: 0.75rem; font-weight: bold; color: #008751; text-transform: uppercase; border-left: 3px solid #ffc107; padding-left: 10px; margin-bottom: 15px; display: block; }}
        input[readonly] {{ background-color: #e9ecef !important; font-weight: bold; }}
        #loginArea {{ margin-top: 100px; }}

        /* ── MOBILE OPTIMISATIONS ── */
        @media (max-width: 768px) {{
            body {{ background-attachment: scroll; }}
            .navbar h5 {{ font-size: 0.75rem; padding: 0 8px; }}
            .container {{ padding-left: 12px !important; padding-right: 12px !important; }}
            .card {{ border-radius: 14px; margin-bottom: 14px; }}
            .card.p-4 {{ padding: 14px !important; }}
            .card.p-5 {{ padding: 22px 16px !important; }}

            /* Stack the 3-col selects vertically */
            .row .col-4 {{ width: 100% !important; flex: 0 0 100%; max-width: 100%; }}
            /* But keep party vote grid at 3-col */
            .row .col-4.col-md-2 {{ width: 33.33% !important; flex: 0 0 33.33%; max-width: 33.33%; }}

            /* Big touch targets */
            .form-control, .form-select {{
                min-height: 48px !important;
                font-size: 1rem !important;
                border-radius: 10px !important;
            }}
            /* Number pad for vote inputs */
            .party-v {{ inputmode: numeric; min-height: 44px !important; font-size: 1rem !important; font-weight: 700; }}

            /* Big submit button */
            .btn.btn-success.btn-lg {{
                min-height: 56px !important;
                font-size: 1.05rem !important;
                border-radius: 14px !important;
            }}
            .btn.btn-outline-dark {{
                min-height: 50px !important;
                font-size: 0.95rem !important;
                border-radius: 12px !important;
            }}
            /* Modal full-width on mobile */
            .modal-dialog {{ margin: 8px; }}
            .modal-dialog-centered {{ align-items: flex-end; }}
            .modal-content {{ border-radius: 16px 16px 0 0 !important; }}
            /* Login card */
            #loginArea {{ margin-top: 50px; }}
            /* Section labels */
            .section-label {{ font-size: 0.72rem; }}
            /* Party card images */
            .col-4.col-md-2 img {{ height: 26px !important; }}
        }}
    </style>
</head>
<body>
    <nav class="navbar py-2 mb-4 text-center"><h5>THE KUKAH CENTRE (TKC) OFFICIAL FIELD COLLATION</h5></nav>
    <div class="container pb-5" style="max-width: 850px;">
        <!-- ── Step 1: LGA + Officer ID ── -->
        <div id="step1Area" class="card p-5 text-center mx-auto" style="max-width: 420px;">
            <div style="width:56px;height:56px;border-radius:50%;background:rgba(0,135,81,0.12);border:1px solid rgba(0,135,81,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">🗳️</div>
            <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ffc107;background:rgba(255,193,7,0.1);border:1px solid rgba(255,193,7,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 1 of 2 — Officer Verification</div>
            <h6 class="fw-bold mb-1">Select Your LGA &amp; Enter Officer ID</h6>
            <p class="small text-muted mb-3">Both fields must match your official assignment</p>
            <select id="lgaSelect" class="form-select mb-2 text-center fw-bold" style="background:#1a1a1a;color:#fff;border:1px solid rgba(0,135,81,0.4);border-radius:8px;">
                <option value="">— Select Your LGA —</option>
                <option value="Atakumosa East">Atakumosa East</option>
                <option value="Atakumosa West">Atakumosa West</option>
                <option value="Ayedaade">Ayedaade</option>
                <option value="Ayedire">Ayedire</option>
                <option value="Boluwaduro">Boluwaduro</option>
                <option value="Boripe">Boripe</option>
                <option value="Ede North">Ede North</option>
                <option value="Ede South">Ede South</option>
                <option value="Egbedore">Egbedore</option>
                <option value="Ejigbo">Ejigbo</option>
                <option value="Ife Central">Ife Central</option>
                <option value="Ife East">Ife East</option>
                <option value="Ife North">Ife North</option>
                <option value="Ife South">Ife South</option>
                <option value="Ifedayo">Ifedayo</option>
                <option value="Ifelodun">Ifelodun</option>
                <option value="Ila">Ila</option>
                <option value="Ilesa East">Ilesa East</option>
                <option value="Ilesa West">Ilesa West</option>
                <option value="Irepodun">Irepodun</option>
                <option value="Irewole">Irewole</option>
                <option value="Isokan">Isokan</option>
                <option value="Iwo">Iwo</option>
                <option value="Obokun">Obokun</option>
                <option value="Odo-Otin">Odo-Otin</option>
                <option value="Ola-Oluwa">Ola-Oluwa</option>
                <option value="Olorunda">Olorunda</option>
                <option value="Oriade">Oriade</option>
                <option value="Orolu">Orolu</option>
                <option value="Osogbo">Osogbo</option>
            </select>
            <input type="text" id="oid" class="form-control mb-2 text-center fw-bold" placeholder="e.g. 10-001" autocomplete="off" style="letter-spacing:0.08em;" onkeydown="if(event.key==='Enter')requestOTP()">
            <div id="step1Error" class="alert alert-danger d-none small py-2 mb-2"></div>
            <button class="btn btn-success w-100 fw-bold" id="step1Btn" onclick="requestOTP()">Send OTP to My WhatsApp →</button>
        </div>

        <!-- ── Step 2: OTP verification ── -->
        <div id="step2Area" class="card p-5 text-center mx-auto d-none" style="max-width: 420px;">
            <div style="width:56px;height:56px;border-radius:50%;background:rgba(0,135,81,0.12);border:1px solid rgba(0,135,81,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">📱</div>
            <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ffc107;background:rgba(255,193,7,0.1);border:1px solid rgba(255,193,7,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 2 of 2 — OTP Verification</div>
            <h6 class="fw-bold mb-1">Check Your WhatsApp</h6>
            <p class="small text-muted mb-1">A 6-digit code was sent to</p>
            <p class="fw-bold mb-3" id="phoneHintDisplay" style="color:#008751;font-size:1rem;letter-spacing:0.1em;">+234***0000</p>
            <input type="text" id="otpInput" class="form-control mb-2 text-center fw-bold" placeholder="000000" maxlength="6" inputmode="numeric" autocomplete="one-time-code" style="font-size:1.4rem;letter-spacing:0.3em;" onkeydown="if(event.key==='Enter')verifyOTP()">
            <div id="step2Error" class="alert alert-danger d-none small py-2 mb-2"></div>
            <button class="btn btn-success w-100 fw-bold mb-2" id="step2Btn" onclick="verifyOTP()">Verify OTP &amp; Unlock Form →</button>
            <div class="d-flex justify-content-between align-items-center">
                <button class="btn btn-link btn-sm text-muted p-0" onclick="backToStep1()">← Change ID</button>
                <button class="btn btn-link btn-sm p-0" id="resendBtn" onclick="resendOTP()" style="color:#008751;">Resend OTP</button>
            </div>
            <div id="resendCountdown" class="small text-muted mt-1 d-none"></div>
        </div>

        <div id="formArea" class="d-none">
            <div class="card p-4">
                <span class="section-label">1. Polling Unit — Auto-filled from Officer ID</span>
                <div class="row g-2">
                    <div class="col-6"><small class="text-muted">State</small><input type="text" id="s" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">LGA</small><input type="text" id="l" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">Ward</small><input type="text" id="w" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">Ward Code</small><input type="text" id="wc" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">PU Code</small><input type="text" id="pc" class="form-control" readonly></div>
                    <div class="col-6"><small class="text-muted">Polling Unit</small><input type="text" id="loc" class="form-control" readonly></div>
                </div>
            </div>

            <div class="card p-4">
                <span class="section-label">2. Official 14-Party Scorecard — 2026 Osun Governorship</span>
                <div class="row g-2">{party_cards}</div>
            </div>

            <div class="card p-4">
                <span class="section-label">3. Accreditation & Audit</span>
                <div class="row g-3">
                    <div class="col-4"><label class="small">Registered</label><input type="number" id="rv" class="form-control" value="0"></div>
                    <div class="col-4"><label class="small">Accredited</label><input type="number" id="ta" class="form-control" oninput="calculateTotals()"></div>
                    <div class="col-4"><label class="small">Rejected</label><input type="number" id="rj" class="form-control" value="0" oninput="calculateTotals()"></div>
                    <div class="col-6"><label class="small text-success">Valid</label><input type="number" id="vv" class="form-control" readonly></div>
                    <div class="col-6"><label class="small text-primary">Total Cast</label><input type="number" id="tc" class="form-control" readonly></div>
                </div>
                <div id="auditStatus" class="mt-3 p-2 rounded text-center d-none small fw-bold"></div>
            </div>

            <!-- BUG FIX #5: GPS no longer hard-blocks — shows warning but allows submission -->
            <button class="btn btn-outline-dark w-100 mb-2" onclick="getGPS()">
                <span id="gpsLabel">📍 Fix GPS Location (Recommended)</span>
            </button>
            <div id="gpsWarning" class="alert alert-warning d-none small py-2 mb-2">
                ⚠️ GPS not captured. You can still submit, but location will not be recorded.
            </div>

            <div class="card mb-3">
                <div class="card-body p-3">
                    <span class="section-label">4. EC 8E FORM IMAGE UPLOAD</span>
                    <div class="mt-2">
                        <label class="form-label small text-muted mb-1">Attach a clear photo of the signed EC 8E form (optional)</label>
                        <input type="file" id="ec8eFile" accept="image/*" capture="environment" class="form-control form-control-sm bg-dark text-white border-secondary">
                        <div id="ec8ePreview" class="mt-2 text-center d-none">
                            <img id="ec8eImg" src="#" alt="EC8E Preview" style="max-width:100%;max-height:200px;border-radius:8px;border:1px solid #ffc107;">
                        </div>
                    </div>
                </div>
            </div>
            <div id="submitError" class="alert alert-danger d-none small py-2 mb-2 fw-bold" style="border-radius:8px;"></div>
            <button class="btn btn-success btn-lg w-100 py-3 fw-bold" onclick="reviewSubmission()">UPLOAD PU RESULT</button>
        </div>
    </div>

        </div><!-- /formArea -->
    </div>
    <!-- Confirmation Modal -->
    <div class="modal fade" id="confirmModal" tabindex="-1">
      <div class="modal-dialog modal-dialog-centered">
        <div class="modal-content" style="background:#1a1a1a;color:#fff;border:1px solid #ffc107;">
          <div class="modal-header" style="border-bottom:1px solid #333;">
            <h6 class="modal-title text-warning fw-bold">⚠️ CONFIRM RESULT SUBMISSION</h6>
          </div>
          <div class="modal-body">
            <div id="confirmPUInfo" class="mb-3 p-2 rounded" style="background:#111;font-size:0.8rem;"></div>
            <table class="table table-sm table-dark table-bordered mb-2" style="font-size:0.8rem;">
              <thead><tr><th>Party</th><th class="text-end">Votes</th></tr></thead>
              <tbody id="confirmPartyRows"></tbody>
            </table>
            <div id="confirmAuditRows" class="p-2 rounded" style="background:#111;font-size:0.8rem;"></div>
          </div>
          <div class="modal-footer" style="border-top:1px solid #333;">
            <button type="button" class="btn btn-outline-secondary btn-sm" data-bs-dismiss="modal">← EDIT</button>
            <button type="button" class="btn btn-success btn-sm fw-bold" onclick="confirmAndSubmit()">✅ CONFIRM & SUBMIT</button>
          </div>
        </div>
      </div>
    </div>
    <div class="container pb-2" style="max-width:850px;">
        <div style="text-align:center;padding:18px 0 10px;border-top:1px solid rgba(255,255,255,0.12);margin-top:10px;">
            <img src="/static/logos/popson-logo.png" style="height:28px;object-fit:contain;vertical-align:middle;margin-right:8px;opacity:0.85;">
            <span style="color:rgba(255,255,255,0.55);font-size:0.75rem;vertical-align:middle;">Powered by <strong style="color:#fff;">Popson Geospatial Services</strong></span>
        </div>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>

    <script>
        let lat = null, lon = null, officerId, officerLga = '', submitToken = null, pendingPayload = null;
        let _resendTimer = null;

        // ── Step 1: request OTP ──────────────────────────────────────────────
        async function requestOTP() {{
            const rawId = document.getElementById('oid').value.trim().toUpperCase();
            const lga   = document.getElementById('lgaSelect').value;
            if (!lga) {{
                const errEl = document.getElementById('step1Error');
                errEl.innerText = 'Please select your LGA first.';
                errEl.classList.remove('d-none');
                return;
            }}
            if (!rawId) return;
            const btn  = document.getElementById('step1Btn');
            const errEl = document.getElementById('step1Error');
            btn.disabled = true; btn.innerText = 'Sending OTP...';
            errEl.classList.add('d-none');
            try {{
                const res = await fetch('/api/request-otp', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{ officer_id: rawId, lg: lga }})
                }});
                const out = await res.json();
                if (!res.ok) {{
                    errEl.innerText = out.detail || 'Error. Try again.';
                    errEl.classList.remove('d-none');
                    btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                    return;
                }}
                if (out.status === 'no_phone') {{
                    errEl.innerText = out.message;
                    errEl.classList.remove('d-none');
                    btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                    return;
                }}
                // OTP sent — move to step 2
                officerId = rawId;
                officerLga = lga;
                document.getElementById('phoneHintDisplay').innerText = out.phone_hint;
                document.getElementById('step1Area').classList.add('d-none');
                document.getElementById('step2Area').classList.remove('d-none');
                document.getElementById('otpInput').focus();
                startResendCountdown(60);
            }} catch(e) {{
                errEl.innerText = 'Server error. Try again.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
            }}
        }}

        // ── Step 2: verify OTP ───────────────────────────────────────────────
        async function verifyOTP() {{
            const otp   = document.getElementById('otpInput').value.trim();
            if (!otp || otp.length !== 6) return;
            const btn   = document.getElementById('step2Btn');
            const errEl = document.getElementById('step2Error');
            btn.disabled = true; btn.innerText = 'Verifying...';
            errEl.classList.add('d-none');
            try {{
                const res = await fetch('/api/verify-otp', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{ officer_id: officerId, lg: officerLga, otp }})
                }});
                const out = await res.json();
                if (!res.ok) {{
                    errEl.innerText = out.detail || 'Incorrect OTP.';
                    errEl.classList.remove('d-none');
                    btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
                    return;
                }}
                // ✅ Auth complete — store token, fill PU fields, show form
                submitToken = out.token;
                document.getElementById('s').value   = (out.state    || 'osun').toUpperCase();
                document.getElementById('l').value   = (out.lg       || '').toUpperCase();
                document.getElementById('w').value   = (out.ward     || '').toUpperCase();
                document.getElementById('wc').value  = out.ward_code || '';
                document.getElementById('pc').value  = out.pu_code   || '';
                document.getElementById('loc').value = (out.location || '').toUpperCase();
                document.getElementById('step2Area').classList.add('d-none');
                document.getElementById('formArea').classList.remove('d-none');
                if (_resendTimer) clearInterval(_resendTimer);
            }} catch(e) {{
                errEl.innerText = 'Server error. Try again.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
            }}
        }}

        // ── Resend OTP ───────────────────────────────────────────────────────
        async function resendOTP() {{
            document.getElementById('resendBtn').disabled = true;
            document.getElementById('otpInput').value = '';
            document.getElementById('step2Error').classList.add('d-none');
            // Re-use requestOTP logic — put ID back and call it
            document.getElementById('step2Area').classList.add('d-none');
            document.getElementById('step1Area').classList.remove('d-none');
            document.getElementById('step1Btn').disabled = false;
            document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
            await requestOTP();
        }}

        function backToStep1() {{
            if (_resendTimer) clearInterval(_resendTimer);
            document.getElementById('step2Area').classList.add('d-none');
            document.getElementById('step1Area').classList.remove('d-none');
            document.getElementById('step1Btn').disabled = false;
            document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
            document.getElementById('step1Error').classList.add('d-none');
            document.getElementById('otpInput').value = '';
            officerLga = '';
        }}

        function startResendCountdown(seconds) {{
            const btn = document.getElementById('resendBtn');
            const cd  = document.getElementById('resendCountdown');
            btn.disabled = true;
            cd.classList.remove('d-none');
            let remaining = seconds;
            cd.innerText = `Resend available in ${{remaining}}s`;
            _resendTimer = setInterval(() => {{
                remaining--;
                if (remaining <= 0) {{
                    clearInterval(_resendTimer);
                    btn.disabled = false;
                    cd.classList.add('d-none');
                }} else {{
                    cd.innerText = `Resend available in ${{remaining}}s`;
                }}
            }}, 1000);
        }}

        // PU fields are now auto-filled from verify-otp — no manual dropdowns needed
        function calculateTotals() {{
            let valid = 0;
            document.querySelectorAll('.party-v').forEach(i => valid += parseInt(i.value || 0));
            const rej = parseInt(document.getElementById('rj').value || 0);
            const acc = parseInt(document.getElementById('ta').value || 0);
            const cast = valid + rej;
            document.getElementById('vv').value = valid;
            document.getElementById('tc').value = cast;
            const msg = document.getElementById('auditStatus');
            if (acc > 0 && cast > acc) {{
                msg.innerHTML = "⚠️ ERROR: Over-voting!";
                msg.className = "mt-3 p-2 bg-danger text-white rounded text-center small fw-bold d-block";
            }} else if (cast > 0 && cast === acc) {{
                msg.innerHTML = "✅ AUDIT BALANCED";
                msg.className = "mt-3 p-2 bg-success text-white rounded text-center small fw-bold d-block";
            }} else {{ msg.className = "d-none"; }}
        }}

        // BUG FIX #5: GPS is now optional — shows label + warning, does not block submission
        function getGPS() {{
            navigator.geolocation.getCurrentPosition(
                pos => {{
                    lat = pos.coords.latitude;
                    lon = pos.coords.longitude;
                    document.getElementById('gpsLabel').innerText = `✅ GPS Fixed (${{lat.toFixed(4)}}, ${{lon.toFixed(4)}})`;
                    document.getElementById('gpsWarning').classList.add('d-none');
                }},
                err => {{
                    document.getElementById('gpsWarning').classList.remove('d-none');
                }}
            );
        }}

        // BUG FIX #2: ec8eFile listener wrapped in DOMContentLoaded
        document.addEventListener('DOMContentLoaded', function() {{
            const ec8eInput = document.getElementById('ec8eFile');
            if (ec8eInput) {{
                ec8eInput.addEventListener('change', function() {{
                    const file = this.files[0];
                    if (file) {{
                        const reader = new FileReader();
                        reader.onload = e => {{
                            document.getElementById('ec8eImg').src = e.target.result;
                            document.getElementById('ec8ePreview').classList.remove('d-none');
                        }};
                        reader.readAsDataURL(file);
                    }}
                }});
            }}
        }});

        async function reviewSubmission() {{
            // BUG FIX #5: GPS warning shown but not blocking
            if(!lat) {{
                document.getElementById('gpsWarning').classList.remove('d-none');
            }}
            const v = {{}};
            document.querySelectorAll('.party-v').forEach(i => v[i.dataset.p] = parseInt(i.value || 0));
            pendingPayload = {{
                officer_id: officerId, submit_token: submitToken,
                state: document.getElementById('s').value, lg: document.getElementById('l').value,
                ward: document.getElementById('w').value, ward_code: document.getElementById('wc').value,
                pu_code: document.getElementById('pc').value, location: document.getElementById('loc').value,
                reg_voters: parseInt(document.getElementById('rv').value || 0), total_accredited: parseInt(document.getElementById('ta').value || 0),
                valid_votes: parseInt(document.getElementById('vv').value || 0), rejected_votes: parseInt(document.getElementById('rj').value || 0),
                total_cast: parseInt(document.getElementById('tc').value || 0), lat, lon, votes: v
            }};
            document.getElementById('confirmPUInfo').innerHTML =
                `<b>📍 PU:</b> ${{pendingPayload.location}}<br>` +
                `<b>🗳 Ward:</b> ${{pendingPayload.ward}} &nbsp;|&nbsp; <b>LGA:</b> ${{pendingPayload.lg}}<br>` +
                `<b>🔑 PU Code:</b> ${{pendingPayload.pu_code}} &nbsp;|&nbsp; <b>Officer:</b> ${{pendingPayload.officer_id}}` +
                (lat ? `<br><b>📡 GPS:</b> ${{lat.toFixed(4)}}, ${{lon.toFixed(4)}}` : `<br><span class="text-warning">⚠️ No GPS captured</span>`);
            const tbody = document.getElementById('confirmPartyRows');
            tbody.innerHTML = '';
            Object.entries(v).forEach(([p, score]) => {{
                const tr = document.createElement('tr');
                tr.innerHTML = `<td>${{p}}</td><td class="text-end fw-bold ${{score > 0 ? 'text-warning' : 'text-secondary'}}">${{score}}</td>`;
                tbody.appendChild(tr);
            }});
            document.getElementById('confirmAuditRows').innerHTML =
                `<b>Registered:</b> ${{pendingPayload.reg_voters}} &nbsp;|&nbsp; ` +
                `<b>Accredited:</b> ${{pendingPayload.total_accredited}} &nbsp;|&nbsp; ` +
                `<b>Valid:</b> ${{pendingPayload.valid_votes}} &nbsp;|&nbsp; ` +
                `<b>Rejected:</b> ${{pendingPayload.rejected_votes}} &nbsp;|&nbsp; ` +
                `<b>Total Cast:</b> ${{pendingPayload.total_cast}}`;
            new bootstrap.Modal(document.getElementById('confirmModal')).show();
        }}

        async function confirmAndSubmit() {{
            const btn = document.querySelector('#confirmModal .btn-success');
            btn.disabled = true; btn.innerText = 'Submitting...';
            const fd = new FormData();
            fd.append('data', JSON.stringify(pendingPayload));
            const ec8eInput = document.getElementById('ec8eFile');
            if (ec8eInput && ec8eInput.files[0]) {{ fd.append('ec8e_image', ec8eInput.files[0]); }}
            const res = await fetch('/submit', {{ method: 'POST', body: fd }});
            const out = await res.json();
            bootstrap.Modal.getInstance(document.getElementById('confirmModal')).hide();
            alert(out.message);
            if(out.status === 'success') location.reload();
            else {{ btn.disabled = false; btn.innerText = '✅ CONFIRM & SUBMIT'; }}
        }}
    </script>
</body>
</html>
"""



# ── Incident Report OTP (no LGA required — officer ID lookup only) ─────────────
@app.post("/api/request-incident-otp")
async def request_incident_otp(request: Request):
    """
    Same as /api/request-otp but without requiring LGA.
    Looks up officer in Supabase first (by officer_id only), then SQLite (without LGA filter).
    Used by the Incident Report form so officers don't need to select their LGA.
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    import re as _re_inc
    officer_id = _re_inc.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)

    parts = officer_id.split("-", 1)
    if len(parts) != 2:
        raise HTTPException(status_code=400, detail="Invalid officer ID format. Expected: WARDCODE-PUCODE")
    ward_code, pu_code = parts[0].strip(), parts[1].strip()

    # Lockout check (no LGA in key for incidents)
    otp_key = f"incident|{officer_id}"
    entry = _OTP_STORE.get(otp_key, {})
    locked_until = entry.get("locked_until", 0)
    if time.time() < locked_until:
        remaining = int(locked_until - time.time())
        raise HTTPException(
            status_code=429,
            detail=f"Too many failed attempts. Try again in {remaining // 60}m {remaining % 60}s."
        )

    phone = None
    pu_data = {}

    # ── 1. Try Supabase (no LGA filter) ──────────────────────────────────────
    sb_officer = _get_supabase_officer(officer_id, "")  # empty lg = no filter
    if sb_officer:
        phone = _clean_phone(sb_officer.get("phone", "") or "")
        lg = sb_officer.get("lga", "")
        try:
            with get_db() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT ward, lg, location, pu_code, ward_code, state
                           FROM polling_units
                           WHERE ward_code = ? AND pu_code = ?
                           AND LOWER(state) = 'osun'
                           LIMIT 1""",
                        (ward_code, pu_code)
                    )
                    row = cur.fetchone()
                    if row:
                        pu_data = {
                            "state":     row["state"] or "osun",
                            "ward":      row["ward"],
                            "lg":        row["lg"],
                            "location":  row["location"],
                            "pu_code":   row["pu_code"],
                            "ward_code": row["ward_code"],
                        }
        except Exception:
            pass
        if not pu_data:
            pu_data = {
                "state":     "osun",
                "ward":      sb_officer.get("ward", ""),
                "lg":        lg,
                "location":  sb_officer.get("polling_unit", ""),
                "pu_code":   pu_code,
                "ward_code": ward_code,
            }
    else:
        # ── 2. Fallback: SQLite polling_units (no LGA filter) ─────────────────
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT ward, lg, location, pu_code, ward_code, state, officer_phone
                       FROM polling_units
                       WHERE ward_code = ? AND pu_code = ?
                       AND LOWER(state) = 'osun'
                       LIMIT 1""",
                    (ward_code, pu_code)
                )
                row = cur.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Officer ID not found. Access Denied.")

        phone = _clean_phone(row["officer_phone"] or "") if row["officer_phone"] else None
        pu_data = {
            "state":     row["state"] or "osun",
            "ward":      row["ward"],
            "lg":        row["lg"],
            "location":  row["location"],
            "pu_code":   row["pu_code"],
            "ward_code": row["ward_code"],
        }

    if not phone or len(phone) < 10:
        return {
            "status":  "no_phone",
            "message": "No phone number registered for this officer ID. Contact your supervisor."
        }

    # Generate & store OTP
    otp = _generate_otp()
    _OTP_STORE[otp_key] = {
        "otp":          otp,
        "expiry":       time.time() + _OTP_TTL,
        "phone_hint":   _mask_phone(phone),
        "phone":        phone,
        "used":         False,
        "attempts":     0,
        "locked_until": 0,
        "pu_data":      pu_data,
    }

    # Send via Twilio WhatsApp
    account_sid = os.environ.get("TWILIO_ACCOUNT_SID", "")
    auth_token  = os.environ.get("TWILIO_AUTH_TOKEN", "")
    from_number = os.environ.get("TWILIO_WHATSAPP_FROM", "+14155238886")

    if not account_sid or not auth_token:
        dev_mode = os.environ.get("OTP_DEV_MODE", "").lower() in ("1", "true", "yes")
        if dev_mode:
            logger.warning("Twilio not configured — OTP not sent (dev mode)")
            logger.info(f"[DEV OTP] {officer_id}: {otp}")
        else:
            raise HTTPException(status_code=503, detail="OTP service is not configured. Contact your administrator.")
    else:
        try:
            from twilio.rest import Client as _TC2
            client = _TC2(account_sid, auth_token)
            msg = (
                f"🚨 *TKC INCIDENT REPORTING*\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"Your One-Time Password:\n\n"
                f"*{otp}*\n\n"
                f"Valid for *5 minutes*. Do NOT share this code.\n"
                f"━━━━━━━━━━━━━━━━━━\n"
                f"_Powered by Popson Geospatial Services_"
            )
            client.messages.create(from_=f"whatsapp:{from_number}", to=f"whatsapp:{phone}", body=msg)
            logger.info(f"✅ Incident OTP sent to {_mask_phone(phone)} for officer {officer_id}")
        except Exception as e:
            logger.error(f"Twilio OTP send failed for {officer_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to send OTP: {str(e)}")

    return {"status": "sent", "phone_hint": _mask_phone(phone)}


@app.post("/api/verify-incident-otp")
async def verify_incident_otp(request: Request):
    """
    Verify OTP for incident reporting. Returns officer/PU data on success.
    No submit token required — incident form uses its own payload submission.
    """
    _check_rate_limit(request.client.host)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    officer_id = str(body.get("officer_id", "")).strip()[:60]
    import re as _re_inc2
    officer_id = _re_inc2.sub(r"[^A-Za-z0-9\-/_ ]", "", officer_id)
    given_otp  = str(body.get("otp", "")).strip()[:6]

    otp_key = f"incident|{officer_id}"
    entry   = _OTP_STORE.get(otp_key)

    if not entry:
        raise HTTPException(status_code=400, detail="No OTP requested for this officer. Start again.")

    if time.time() < entry.get("locked_until", 0):
        remaining = int(entry["locked_until"] - time.time())
        raise HTTPException(status_code=429, detail=f"Account locked. Try again in {remaining // 60}m {remaining % 60}s.")

    if entry.get("used"):
        raise HTTPException(status_code=400, detail="OTP already used. Request a new one.")

    if time.time() > entry["expiry"]:
        _OTP_STORE.pop(otp_key, None)
        raise HTTPException(status_code=400, detail="OTP expired. Request a new one.")

    if not secrets.compare_digest(given_otp, entry["otp"]):
        entry["attempts"] = entry.get("attempts", 0) + 1
        remaining_tries = _OTP_MAX_TRIES - entry["attempts"]
        if entry["attempts"] >= _OTP_MAX_TRIES:
            entry["locked_until"] = time.time() + _OTP_LOCKOUT
            entry["used"] = True
            raise HTTPException(status_code=429, detail=f"Too many wrong attempts. Account locked for {_OTP_LOCKOUT // 60} minutes.")
        raise HTTPException(status_code=401, detail=f"Incorrect OTP. {remaining_tries} attempt{'s' if remaining_tries != 1 else ''} remaining.")

    # ✅ Correct OTP
    entry["used"] = True
    pu_data = entry["pu_data"]
    return {
        "status":    "ok",
        "officer_id": officer_id,
        "state":     pu_data["state"],
        "ward":      pu_data["ward"],
        "lg":        pu_data["lg"],
        "location":  pu_data["location"],
        "pu_code":   pu_data["pu_code"],
        "ward_code": pu_data["ward_code"],
    }

# ── INCIDENT REPORT FORM ──────────────────────────────────────────────────────
@app.get("/report", response_class=HTMLResponse)
async def report_page():
    return HTMLResponse(content=REPORT_HTML)

REPORT_HTML = """
<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
    <meta name="mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="theme-color" content="#cc0000">
    <title>TKC — INCIDENT REPORT</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background: linear-gradient(160deg,#0a1f0e 0%,#0d2b10 50%,#0a1f0e 100%); background-attachment: fixed; min-height: 100vh; margin: 0; }
        .navbar { background: rgba(180,0,0,0.92) !important; backdrop-filter: blur(10px); color: white; border-bottom: 4px solid #ff6600; }
        .card { background: rgba(255,255,255,0.96) !important; border-radius: 12px; border: none; box-shadow: 0 10px 30px rgba(0,0,0,0.4) !important; margin-bottom: 20px; color: #222; }
        .section-label { font-size: 0.75rem; font-weight: bold; color: #cc0000; text-transform: uppercase; border-left: 3px solid #ff6600; padding-left: 10px; margin-bottom: 15px; display: block; }
        .severity-btn { border: 2px solid #ddd; border-radius: 8px; padding: 10px; cursor: pointer; text-align: center; transition: all 0.2s; background: #fff; }
        .severity-btn.active-low { border-color: #28a745; background: #d4edda; }
        .severity-btn.active-medium { border-color: #ffc107; background: #fff3cd; }
        .severity-btn.active-critical { border-color: #dc3545; background: #f8d7da; }
        #step1Area, #step2Area { margin-top: 100px; }
        @media (max-width: 768px) {
            body { background-attachment: scroll; }
            .navbar h5 { font-size: 0.75rem; }
            #step1Area, #step2Area { margin-top: 50px; }
            .form-control, .form-select { min-height: 48px !important; font-size: 1rem !important; border-radius: 10px !important; }
            .btn-lg { min-height: 56px !important; font-size: 1.05rem !important; border-radius: 14px !important; }
        }
    </style>
</head>
<body>
<nav class="navbar py-2 mb-4 text-center"><h5>⚠️ TKC INCIDENT REPORTING SYSTEM</h5></nav>
<div class="container pb-5" style="max-width: 750px;">

    <!-- Step 1: Officer ID only (no LGA needed) -->
    <div id="step1Area" class="card p-5 text-center mx-auto" style="max-width: 420px;">
        <div style="width:56px;height:56px;border-radius:50%;background:rgba(180,0,0,0.1);border:1px solid rgba(180,0,0,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">🚨</div>
        <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ff6600;background:rgba(255,102,0,0.1);border:1px solid rgba(255,102,0,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 1 of 2 — Officer Verification</div>
        <h6 class="fw-bold mb-1 text-danger">🔐 Officer Verification</h6>
        <p class="small text-muted mb-3">Enter your Officer ID to receive an OTP on WhatsApp</p>
        <input type="text" id="oid" class="form-control mb-3 text-center fw-bold" placeholder="e.g. 10-001" autocomplete="off" style="letter-spacing:0.08em;" onkeydown="if(event.key==='Enter')requestOTP()">
        <div id="step1Error" class="alert alert-danger d-none small py-2 mb-2"></div>
        <button class="btn btn-danger w-100 fw-bold" id="step1Btn" onclick="requestOTP()">Send OTP to My WhatsApp →</button>
        <div class="mt-3">
            <a href="/vote" class="small text-muted">← Back to Vote Submission</a>
        </div>
    </div>

    <!-- Step 2: OTP entry -->
    <div id="step2Area" class="card p-5 text-center mx-auto d-none" style="max-width: 420px;">
        <div style="width:56px;height:56px;border-radius:50%;background:rgba(180,0,0,0.1);border:1px solid rgba(180,0,0,0.3);display:flex;align-items:center;justify-content:center;font-size:1.4rem;margin:0 auto 16px;">📱</div>
        <div style="font-size:0.6rem;font-weight:700;letter-spacing:0.15em;color:#ff6600;background:rgba(255,102,0,0.1);border:1px solid rgba(255,102,0,0.2);border-radius:20px;padding:3px 12px;display:inline-block;margin-bottom:14px;text-transform:uppercase;">Step 2 of 2 — OTP Verification</div>
        <h6 class="fw-bold mb-1">Check Your WhatsApp</h6>
        <p class="small text-muted mb-1">A 6-digit code was sent to</p>
        <p class="fw-bold mb-3" id="phoneHintDisplay" style="color:#cc0000;font-size:1rem;letter-spacing:0.1em;">+234***0000</p>
        <input type="text" id="otpInput" class="form-control mb-2 text-center fw-bold" placeholder="000000" maxlength="6" inputmode="numeric" autocomplete="one-time-code" style="font-size:1.4rem;letter-spacing:0.3em;" onkeydown="if(event.key==='Enter')verifyOTP()">
        <div id="step2Error" class="alert alert-danger d-none small py-2 mb-2"></div>
        <button class="btn btn-danger w-100 fw-bold mb-2" id="step2Btn" onclick="verifyOTP()">Verify OTP &amp; Unlock Form →</button>
        <div class="d-flex justify-content-between align-items-center">
            <button class="btn btn-link btn-sm text-muted p-0" onclick="backToStep1()">← Change ID</button>
            <button class="btn btn-link btn-sm p-0" id="resendBtn" onclick="resendOTP()" style="color:#cc0000;">Resend OTP</button>
        </div>
        <div id="resendCountdown" class="small text-muted mt-1 d-none"></div>
    </div>

    <!-- Incident Form (shown after OTP verified) -->
    <div id="formArea" class="d-none">

        <div class="card p-4">
            <span class="section-label">1. Officer &amp; Location</span>
            <div class="row g-2">
                <div class="col-6"><small class="text-muted">Officer ID</small><input type="text" id="disp_officer" class="form-control" readonly></div>
                <div class="col-6"><small class="text-muted">PU Code</small><input type="text" id="disp_pu" class="form-control" readonly></div>
                <div class="col-12"><small class="text-muted">Polling Unit</small><input type="text" id="disp_loc" class="form-control" readonly></div>
                <div class="col-6"><small class="text-muted">Ward</small><input type="text" id="disp_ward" class="form-control" readonly></div>
                <div class="col-6"><small class="text-muted">LGA</small><input type="text" id="disp_lg" class="form-control" readonly></div>
            </div>
        </div>

        <div class="card p-4">
            <span class="section-label">2. Incident Type</span>
            <select id="incident_type" class="form-select">
                <option value="">-- Select Incident Type --</option>
                <option value="Violence / security threat">⚔️ Violence / Security Threat</option>
                <option value="Ballot box snatching">🗳️ Ballot Box Snatching</option>
                <option value="INEC official misconduct">🏛️ INEC Official Misconduct</option>
                <option value="Voter intimidation">😰 Voter Intimidation</option>
                <option value="Equipment failure (BVAS etc)">🖥️ Equipment Failure (BVAS etc)</option>
                <option value="Other irregularities">⚠️ Other Irregularities</option>
            </select>
        </div>

        <div class="card p-4">
            <span class="section-label">3. Severity Level</span>
            <div class="row g-2">
                <div class="col-4">
                    <div class="severity-btn" id="btn-low" onclick="setSeverity('Low')">
                        <div style="font-size:1.5rem;">🟢</div>
                        <div class="fw-bold small">LOW</div>
                        <div style="font-size:0.65rem; color:#666;">Minor issue</div>
                    </div>
                </div>
                <div class="col-4">
                    <div class="severity-btn" id="btn-medium" onclick="setSeverity('Medium')">
                        <div style="font-size:1.5rem;">🟡</div>
                        <div class="fw-bold small">MEDIUM</div>
                        <div style="font-size:0.65rem; color:#666;">Needs attention</div>
                    </div>
                </div>
                <div class="col-4">
                    <div class="severity-btn" id="btn-critical" onclick="setSeverity('Critical')">
                        <div style="font-size:1.5rem;">🔴</div>
                        <div class="fw-bold small">CRITICAL</div>
                        <div style="font-size:0.65rem; color:#666;">Urgent response</div>
                    </div>
                </div>
            </div>
            <input type="hidden" id="severity_val" value="">
        </div>

        <div class="card p-4">
            <span class="section-label">4. Description</span>
            <textarea id="description" class="form-control" rows="4" placeholder="Describe what happened in detail — who, what, when, how many people involved..."></textarea>
        </div>

        <div class="card p-4">
            <span class="section-label">5. Evidence (Photo / Video)</span>
            <label class="form-label small text-muted">Attach photo or video evidence (optional but strongly recommended)</label>
            <input type="file" id="evidenceFile" accept="image/*,video/*" capture="environment" class="form-control bg-dark text-white border-secondary">
            <div id="evidencePreview" class="mt-2 text-center d-none">
                <img id="evidenceImg" src="#" alt="Preview" style="max-width:100%;max-height:200px;border-radius:8px;border:2px solid #ff6600;">
            </div>
        </div>

        <button class="btn btn-outline-dark w-100 mb-2" onclick="getGPS()">
            <span id="gpsLabel">📍 Fix GPS Location (Recommended)</span>
        </button>
        <div id="gpsWarning" class="alert alert-warning d-none small py-2 mb-2">
            ⚠️ GPS not captured. You can still submit.
        </div>

        <div id="submitError" class="alert alert-danger d-none small py-2 mb-2 fw-bold"></div>
        <button class="btn btn-danger btn-lg w-100 py-3 fw-bold" onclick="submitIncident()">🚨 SUBMIT INCIDENT REPORT</button>
        <div class="mt-3 text-center">
            <a href="/vote" class="small text-white">← Back to Vote Submission</a>
        </div>
    </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script>
    let lat = null, lon = null, officerData = {}, _resendTimer = null;

    // ── Step 1: request OTP ──────────────────────────────────────────────────
    async function requestOTP() {
        const rawId = document.getElementById('oid').value.trim().toUpperCase();
        if (!rawId) return;
        const btn   = document.getElementById('step1Btn');
        const errEl = document.getElementById('step1Error');
        btn.disabled = true; btn.innerText = 'Sending OTP...';
        errEl.classList.add('d-none');
        try {
            const res = await fetch('/api/request-incident-otp', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ officer_id: rawId })
            });
            const out = await res.json();
            if (!res.ok) {
                errEl.innerText = out.detail || 'Error. Try again.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                return;
            }
            if (out.status === 'no_phone') {
                errEl.innerText = out.message;
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
                return;
            }
            // OTP sent — move to step 2
            officerData.officer_id = rawId;
            document.getElementById('phoneHintDisplay').innerText = out.phone_hint;
            document.getElementById('step1Area').classList.add('d-none');
            document.getElementById('step2Area').classList.remove('d-none');
            document.getElementById('otpInput').focus();
            startResendCountdown(60);
        } catch(e) {
            errEl.innerText = 'Server error. Try again.';
            errEl.classList.remove('d-none');
            btn.disabled = false; btn.innerText = 'Send OTP to My WhatsApp →';
        }
    }

    // ── Step 2: verify OTP ───────────────────────────────────────────────────
    async function verifyOTP() {
        const otp   = document.getElementById('otpInput').value.trim();
        if (!otp || otp.length !== 6) return;
        const btn   = document.getElementById('step2Btn');
        const errEl = document.getElementById('step2Error');
        btn.disabled = true; btn.innerText = 'Verifying...';
        errEl.classList.add('d-none');
        try {
            const res = await fetch('/api/verify-incident-otp', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({ officer_id: officerData.officer_id, otp })
            });
            const out = await res.json();
            if (!res.ok) {
                errEl.innerText = out.detail || 'Incorrect OTP.';
                errEl.classList.remove('d-none');
                btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
                return;
            }
            // ✅ Auth done — fill officer/PU data and show incident form
            officerData = {
                officer_id: out.officer_id,
                pu_code:    out.pu_code,
                ward:       out.ward,
                ward_code:  out.ward_code,
                lg:         out.lg,
                state:      out.state,
                location:   out.location
            };
            document.getElementById('disp_officer').value = out.officer_id;
            document.getElementById('disp_pu').value      = out.pu_code  || '';
            document.getElementById('disp_loc').value     = out.location || '';
            document.getElementById('disp_ward').value    = out.ward     || '';
            document.getElementById('disp_lg').value      = out.lg       || '';
            document.getElementById('step2Area').classList.add('d-none');
            document.getElementById('formArea').classList.remove('d-none');
            if (_resendTimer) clearInterval(_resendTimer);
        } catch(e) {
            errEl.innerText = 'Server error. Try again.';
            errEl.classList.remove('d-none');
            btn.disabled = false; btn.innerText = 'Verify OTP & Unlock Form →';
        }
    }

    // ── Resend OTP ───────────────────────────────────────────────────────────
    async function resendOTP() {
        document.getElementById('resendBtn').disabled = true;
        document.getElementById('otpInput').value = '';
        document.getElementById('step2Error').classList.add('d-none');
        document.getElementById('step2Area').classList.add('d-none');
        document.getElementById('step1Area').classList.remove('d-none');
        document.getElementById('step1Btn').disabled = false;
        document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
        await requestOTP();
    }

    function backToStep1() {
        if (_resendTimer) clearInterval(_resendTimer);
        document.getElementById('step2Area').classList.add('d-none');
        document.getElementById('step1Area').classList.remove('d-none');
        document.getElementById('step1Btn').disabled = false;
        document.getElementById('step1Btn').innerText = 'Send OTP to My WhatsApp →';
        document.getElementById('step1Error').classList.add('d-none');
        document.getElementById('otpInput').value = '';
    }

    function startResendCountdown(seconds) {
        const btn = document.getElementById('resendBtn');
        const cd  = document.getElementById('resendCountdown');
        btn.disabled = true;
        cd.classList.remove('d-none');
        let remaining = seconds;
        cd.innerText = `Resend available in ${remaining}s`;
        _resendTimer = setInterval(() => {
            remaining--;
            if (remaining <= 0) {
                clearInterval(_resendTimer);
                btn.disabled = false;
                cd.classList.add('d-none');
            } else {
                cd.innerText = `Resend available in ${remaining}s`;
            }
        }, 1000);
    }

    function setSeverity(level) {
        document.getElementById('severity_val').value = level;
        ['low','medium','critical'].forEach(l => {
            const btn = document.getElementById('btn-' + l);
            btn.className = 'severity-btn' + (l === level.toLowerCase() ? ' active-' + l : '');
        });
    }

    function getGPS() {
        navigator.geolocation.getCurrentPosition(
            pos => {
                lat = pos.coords.latitude;
                lon = pos.coords.longitude;
                document.getElementById('gpsLabel').innerText = `✅ GPS Fixed (${lat.toFixed(4)}, ${lon.toFixed(4)})`;
                document.getElementById('gpsWarning').classList.add('d-none');
            },
            () => { document.getElementById('gpsWarning').classList.remove('d-none'); }
        );
    }

    document.addEventListener('DOMContentLoaded', function() {
        const ef = document.getElementById('evidenceFile');
        if (ef) {
            ef.addEventListener('change', function() {
                const file = this.files[0];
                if (file && file.type.startsWith('image/')) {
                    const reader = new FileReader();
                    reader.onload = e => {
                        document.getElementById('evidenceImg').src = e.target.result;
                        document.getElementById('evidencePreview').classList.remove('d-none');
                    };
                    reader.readAsDataURL(file);
                }
            });
        }
    });

    async function submitIncident() {
        const errEl = document.getElementById('submitError');
        errEl.classList.add('d-none');

        const incident_type = document.getElementById('incident_type').value;
        const severity = document.getElementById('severity_val').value;
        const description = document.getElementById('description').value.trim();

        if (!incident_type) { errEl.innerText = 'Please select an incident type.'; errEl.classList.remove('d-none'); return; }
        if (!severity) { errEl.innerText = 'Please select a severity level.'; errEl.classList.remove('d-none'); return; }
        if (!description) { errEl.innerText = 'Please describe the incident.'; errEl.classList.remove('d-none'); return; }

        const btn = document.querySelector('button.btn-danger.btn-lg');
        btn.disabled = true; btn.innerText = 'Submitting...';

        const payload = {
            officer_id: officerData.officer_id,
            pu_code: officerData.pu_code,
            ward: officerData.ward,
            ward_code: officerData.ward_code || '',
            lg: officerData.lg,
            state: officerData.state || 'osun',
            location: officerData.location,
            incident_type, severity, description, lat, lon
        };

        const fd = new FormData();
        fd.append('data', JSON.stringify(payload));
        const ef = document.getElementById('evidenceFile');
        if (ef && ef.files[0]) fd.append('evidence', ef.files[0]);

        try {
            const res = await fetch('/submit-incident', { method: 'POST', body: fd });
            const out = await res.json();
            alert(out.message);
            if (out.status === 'success') {
                document.getElementById('incident_type').value = '';
                document.getElementById('severity_val').value = '';
                document.getElementById('description').value = '';
                document.getElementById('evidencePreview').classList.add('d-none');
                ['low','medium','critical'].forEach(l => { document.getElementById('btn-'+l).className = 'severity-btn'; });
                btn.disabled = false; btn.innerText = '🚨 SUBMIT INCIDENT REPORT';
            } else {
                btn.disabled = false; btn.innerText = '🚨 SUBMIT INCIDENT REPORT';
            }
        } catch (e) {
            errEl.innerText = 'Server error. Try again.';
            errEl.classList.remove('d-none');
            btn.disabled = false; btn.innerText = '🚨 SUBMIT INCIDENT REPORT';
        }
    }
</script>
</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────

# --- DASHBOARD PAGE ---
# ── INCIDENT DASHBOARD ────────────────────────────────────────────────────────
@app.get("/incident-dashboard", response_class=HTMLResponse)
async def incident_dashboard_page(request: Request):
    token = request.cookies.get("ds_session")
    if not _is_valid_token(token):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url="/dashboard", status_code=302)
    return HTMLResponse(content=INCIDENT_DASHBOARD_HTML)

INCIDENT_DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TKC — Incident Observation Dashboard</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css">
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <style>
        :root { --red: #cc0000; --orange: #ff6600; --gold: #ffc107; --dark: #0a0a0a; --panel: #141414; }
        body { background: var(--dark); color: #fff; font-family: 'Segoe UI', sans-serif; margin: 0; overflow-y: auto; }

        .navbar-custom { background: #1a0000; border-bottom: 3px solid var(--red); padding: 10px 20px; display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 10px; }
        .brand-title { color: var(--red); font-weight: 900; font-size: 1.1rem; letter-spacing: 1px; }

        .kpi-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; padding: 10px; }
        .kpi-card { background: var(--panel); border-radius: 10px; padding: 15px; text-align: center; border: 1px solid #222; }
        .kpi-val { font-size: 2rem; font-weight: 900; display: block; }
        .kpi-label { font-size: 0.65rem; color: #aaa; text-transform: uppercase; margin-top: 4px; }
        .kpi-critical { border-top: 3px solid #ff0000; }
        .kpi-medium   { border-top: 3px solid #ffc107; }
        .kpi-low      { border-top: 3px solid #00cc44; }
        .kpi-total    { border-top: 3px solid var(--orange); }

        .main-grid { display: grid; grid-template-columns: 340px 1fr; gap: 10px; padding: 0 10px 10px; min-height: 0; }
        .side-panel { background: var(--panel); border-radius: 12px; border: 1px solid #222; display: flex; flex-direction: column; overflow: hidden; max-height: calc(100vh - 200px); }
        .panel-header { background: #1c0000; padding: 10px 15px; font-size: 0.75rem; font-weight: bold; color: var(--red); border-bottom: 1px solid #330000; text-transform: uppercase; display: flex; align-items: center; justify-content: space-between; }

        .filter-row { padding: 8px; display: flex; gap: 6px; flex-wrap: wrap; background: #0f0f0f; border-bottom: 1px solid #1a1a1a; }
        .filter-row select { flex: 1; min-width: 90px; background: #1a1a1a; color: #fff; border: 1px solid #333; border-radius: 6px; font-size: 0.72rem; padding: 4px 6px; }

        .feed-container { flex: 1; overflow-y: auto; padding: 8px; }
        .incident-card { background: #1e1e1e; border-radius: 8px; padding: 10px 12px; margin-bottom: 8px; cursor: pointer; transition: background 0.15s; }
        .incident-card:hover { background: #2a2a2a; }
        .incident-card.critical { border-left: 4px solid #ff0000; }
        .incident-card.medium   { border-left: 4px solid #ffc107; }
        .incident-card.low      { border-left: 4px solid #00cc44; }

        .sev-badge { display: inline-block; font-size: 0.6rem; font-weight: bold; padding: 2px 7px; border-radius: 10px; margin-left: 6px; }
        .sev-critical { background: #ff0000; color: #fff; }
        .sev-medium   { background: #ffc107; color: #000; }
        .sev-low      { background: #00cc44; color: #000; }

        #map { height: 380px; border-radius: 12px; background: #111; margin-bottom: 10px; }

        .evidence-panel { background: var(--panel); border-radius: 12px; border: 1px solid #222; padding: 12px; margin-bottom: 10px; min-height: 120px; }
        .evidence-panel-title { font-size: 0.7rem; color: var(--orange); font-weight: bold; text-transform: uppercase; margin-bottom: 8px; border-bottom: 1px solid #2a2a2a; padding-bottom: 6px; }

        .detail-panel { background: var(--panel); border-radius: 12px; border: 1px solid #222; padding: 14px; }
        .detail-row { display: flex; justify-content: space-between; padding: 5px 0; border-bottom: 1px solid #1a1a1a; font-size: 0.75rem; }
        .detail-row:last-child { border-bottom: none; }
        .detail-label { color: #666; }
        .detail-val { color: #fff; font-weight: bold; text-align: right; max-width: 60%; }

        @media (max-width: 768px) {
            .kpi-row { grid-template-columns: repeat(2, 1fr); }
            .main-grid { grid-template-columns: 1fr; }
            .side-panel { max-height: 400px; }
        }
    </style>
</head>
<body>

<nav class="navbar-custom">
    <div>
        <div class="brand-title">🚨 TKC INCIDENT COMMAND — OSUN 2026</div>
        <div style="font-size:0.65rem;color:#666;margin-top:2px;">Security & Command Team Dashboard · Auto-refresh every 30s</div>
    </div>
    <div style="display:flex;gap:8px;align-items:center;">
        <span id="last-refresh" style="font-size:0.65rem;color:#555;"></span>
        <button class="btn btn-sm btn-outline-danger py-1 px-3" style="font-size:11px;" onclick="loadIncidents()">
            <i class="bi bi-arrow-clockwise"></i> REFRESH
        </button>
        <a href="/dashboard" class="btn btn-sm btn-outline-warning py-1 px-3" style="font-size:11px;">
            📊 Vote Dashboard
        </a>
    </div>
</nav>

<!-- KPI Row -->
<div class="kpi-row">
    <div class="kpi-card kpi-total">
        <span id="kpi-total" class="kpi-val" style="color:var(--orange);">0</span>
        <div class="kpi-label">Total Incidents</div>
    </div>
    <div class="kpi-card kpi-critical">
        <span id="kpi-critical" class="kpi-val" style="color:#ff4444;">0</span>
        <div class="kpi-label">🔴 Critical</div>
    </div>
    <div class="kpi-card kpi-medium">
        <span id="kpi-medium" class="kpi-val" style="color:#ffc107;">0</span>
        <div class="kpi-label">🟡 Medium</div>
    </div>
    <div class="kpi-card kpi-low">
        <span id="kpi-low" class="kpi-val" style="color:#00cc44;">0</span>
        <div class="kpi-label">🟢 Low</div>
    </div>
</div>

<div class="main-grid">

    <!-- Left: Incident Feed -->
    <div class="side-panel">
        <div class="panel-header">
            LIVE INCIDENT FEED
            <span id="feed-count" class="badge bg-danger ms-1">0</span>
        </div>
        <div class="filter-row">
            <select id="f-severity" onchange="applyFilters()">
                <option value="">All Severity</option>
                <option value="Critical">🔴 Critical</option>
                <option value="Medium">🟡 Medium</option>
                <option value="Low">🟢 Low</option>
            </select>
            <select id="f-type" onchange="applyFilters()">
                <option value="">All Types</option>
                <option value="Violence / security threat">Violence</option>
                <option value="Ballot box snatching">Ballot Snatching</option>
                <option value="INEC official misconduct">INEC Misconduct</option>
                <option value="Voter intimidation">Voter Intimidation</option>
                <option value="Equipment failure (BVAS etc)">Equipment Failure</option>
                <option value="Other irregularities">Other</option>
            </select>
            <select id="f-lga" onchange="applyFilters()">
                <option value="">All LGAs</option>
            </select>
        </div>
        <div class="feed-container" id="incidentFeed"></div>
    </div>

    <!-- Right: Map + Detail -->
    <div>
        <div id="map"></div>

        <div class="evidence-panel" id="evidencePanel">
            <div class="evidence-panel-title">📷 Evidence Viewer — Click an incident to view</div>
            <div id="evidenceContent" style="text-align:center;color:#444;font-size:0.75rem;padding:20px 0;">
                No incident selected
            </div>
        </div>

        <div class="detail-panel" id="detailPanel">
            <div style="font-size:0.7rem;color:var(--orange);font-weight:bold;text-transform:uppercase;margin-bottom:10px;border-bottom:1px solid #2a2a2a;padding-bottom:6px;">
                📋 Incident Detail
            </div>
            <div id="detailContent" style="color:#555;font-size:0.75rem;">Select an incident from the feed</div>
        </div>
    </div>
</div>

<script>
    let map, allIncidents = [], markers = [];

    const SEV_COLOR = { Critical: '#ff0000', Medium: '#ffc107', Low: '#00cc44' };
    const TYPE_ICON = {
        'Violence / security threat': '⚔️',
        'Ballot box snatching': '🗳️',
        'INEC official misconduct': '🏛️',
        'Voter intimidation': '😰',
        'Equipment failure (BVAS etc)': '🖥️',
        'Other irregularities': '⚠️'
    };

    function initMap() {
        map = L.map('map', { zoomControl: true }).setView([7.56, 4.52], 9);
        L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png').addTo(map);
    }

    async function loadIncidents() {
        try {
            const res = await fetch('/api/incidents', { credentials: 'include' });
            allIncidents = await res.json();
            document.getElementById('last-refresh').textContent = 'Last refresh: ' + new Date().toLocaleTimeString('en-NG');
            populateLGAFilter();
            applyFilters();
        } catch(e) { console.error('Load error', e); }
    }

    function populateLGAFilter() {
        const lgas = [...new Set(allIncidents.map(i => i.lg).filter(Boolean))].sort();
        const sel = document.getElementById('f-lga');
        const cur = sel.value;
        sel.innerHTML = '<option value="">All LGAs</option>';
        lgas.forEach(l => sel.add(new Option(l.toUpperCase(), l)));
        sel.value = cur;
    }

    function applyFilters() {
        const sev = document.getElementById('f-severity').value;
        const typ = document.getElementById('f-type').value;
        const lga = document.getElementById('f-lga').value;

        let filtered = allIncidents;
        if (sev) filtered = filtered.filter(i => i.severity === sev);
        if (typ) filtered = filtered.filter(i => i.incident_type === typ);
        if (lga) filtered = filtered.filter(i => i.lg === lga);

        updateKPIs(filtered);
        renderFeed(filtered);
        renderMap(filtered);
    }

    function updateKPIs(data) {
        document.getElementById('kpi-total').textContent = data.length;
        document.getElementById('kpi-critical').textContent = data.filter(i => i.severity === 'Critical').length;
        document.getElementById('kpi-medium').textContent = data.filter(i => i.severity === 'Medium').length;
        document.getElementById('kpi-low').textContent = data.filter(i => i.severity === 'Low').length;
        document.getElementById('feed-count').textContent = data.length;
    }

    function renderFeed(data) {
        const feed = document.getElementById('incidentFeed');
        if (!data.length) {
            feed.innerHTML = '<div style="color:#555;font-size:0.75rem;padding:20px;text-align:center;">No incidents found</div>';
            return;
        }
        feed.innerHTML = data.map((inc, idx) => {
            const sev = (inc.severity || 'low').toLowerCase();
            const icon = TYPE_ICON[inc.incident_type] || '⚠️';
            const ts = inc.timestamp ? new Date(inc.timestamp).toLocaleString('en-NG', {day:'2-digit',month:'short',hour:'2-digit',minute:'2-digit'}) : '--';
            return `<div class="incident-card ${sev}" onclick="selectIncident(${idx})">
                <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:4px;">
                    <span style="font-size:0.8rem;font-weight:bold;">${icon} ${inc.incident_type || 'Unknown'}</span>
                    <span class="sev-badge sev-${sev}">${inc.severity || ''}</span>
                </div>
                <div style="font-size:0.72rem;color:#aaa;">${(inc.location || '').toUpperCase()} · ${(inc.lg || '').toUpperCase()}</div>
                <div style="font-size:0.68rem;color:#666;margin-top:3px;">👤 ${inc.officer_id || ''} · ${ts}</div>
                ${inc.description ? `<div style="font-size:0.7rem;color:#ccc;margin-top:5px;border-top:1px solid #2a2a2a;padding-top:5px;">${inc.description.substring(0,80)}${inc.description.length > 80 ? '...' : ''}</div>` : ''}
            </div>`;
        }).join('');
    }

    function renderMap(data) {
        markers.forEach(m => map.removeLayer(m));
        markers = [];
        data.forEach((inc, idx) => {
            if (!inc.lat || !inc.lon) return;
            const color = SEV_COLOR[inc.severity] || '#aaa';
            const m = L.circleMarker([inc.lat, inc.lon], {
                radius: inc.severity === 'Critical' ? 10 : inc.severity === 'Medium' ? 7 : 5,
                color: color, fillColor: color, fillOpacity: 0.85, weight: 2
            }).addTo(map);
            m.bindPopup(`<b>${inc.incident_type}</b><br>${inc.location}<br><span style="color:${color};font-weight:bold;">${inc.severity}</span>`);
            m.on('click', () => selectIncident(idx));
            markers.push(m);
        });
    }

    function selectIncident(idx) {
        const filtered = getFiltered();
        const inc = filtered[idx];
        if (!inc) return;

        // Pan map
        if (inc.lat && inc.lon) map.setView([inc.lat, inc.lon], 14);

        // Evidence
        const evEl = document.getElementById('evidenceContent');
        if (inc.evidence_url) {
            evEl.innerHTML = `<img src="${inc.evidence_url}" style="max-width:100%;max-height:220px;border-radius:8px;border:2px solid var(--orange);cursor:zoom-in;" onclick="window.open('${inc.evidence_url}','_blank')"><div style="font-size:0.65rem;color:#555;margin-top:4px;">Click to open full size</div>`;
        } else {
            evEl.innerHTML = '<div style="color:#444;font-size:0.75rem;padding:20px 0;">No evidence uploaded for this incident</div>';
        }

        // Detail
        const sev = (inc.severity || '').toLowerCase();
        const sevColor = SEV_COLOR[inc.severity] || '#aaa';
        const ts = inc.timestamp ? new Date(inc.timestamp).toLocaleString('en-NG') : '--';
        document.getElementById('detailContent').innerHTML = `
            <div class="detail-row"><span class="detail-label">Type</span><span class="detail-val">${TYPE_ICON[inc.incident_type] || ''} ${inc.incident_type || '--'}</span></div>
            <div class="detail-row"><span class="detail-label">Severity</span><span class="detail-val"><span class="sev-badge sev-${sev}">${inc.severity || '--'}</span></span></div>
            <div class="detail-row"><span class="detail-label">Officer</span><span class="detail-val">${inc.officer_id || '--'}</span></div>
            <div class="detail-row"><span class="detail-label">Polling Unit</span><span class="detail-val">${(inc.location || '--').toUpperCase()}</span></div>
            <div class="detail-row"><span class="detail-label">PU Code</span><span class="detail-val">${inc.pu_code || '--'}</span></div>
            <div class="detail-row"><span class="detail-label">Ward</span><span class="detail-val">${(inc.ward || '--').toUpperCase()}</span></div>
            <div class="detail-row"><span class="detail-label">LGA</span><span class="detail-val">${(inc.lg || '--').toUpperCase()}</span></div>
            <div class="detail-row"><span class="detail-label">GPS</span><span class="detail-val">${inc.lat ? inc.lat.toFixed(4) + ', ' + inc.lon.toFixed(4) : 'Not captured'}</span></div>
            <div class="detail-row"><span class="detail-label">Reported</span><span class="detail-val">${ts}</span></div>
            ${inc.description ? `<div style="margin-top:10px;padding:8px;background:#1a1a1a;border-radius:6px;font-size:0.72rem;color:#ccc;border-left:3px solid ${sevColor};">${inc.description}</div>` : ''}
        `;
    }

    function getFiltered() {
        const sev = document.getElementById('f-severity').value;
        const typ = document.getElementById('f-type').value;
        const lga = document.getElementById('f-lga').value;
        let filtered = allIncidents;
        if (sev) filtered = filtered.filter(i => i.severity === sev);
        if (typ) filtered = filtered.filter(i => i.incident_type === typ);
        if (lga) filtered = filtered.filter(i => i.lg === lga);
        return filtered;
    }

    document.addEventListener('DOMContentLoaded', () => {
        initMap();
        loadIncidents();
        setInterval(loadIncidents, 30000);
    });
</script>
</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────



# ── INSIGHT API ENDPOINTS ─────────────────────────────────────────────────────

@app.get("/api/lga_completion")
async def lga_completion(request: Request):
    _require_dashboard(request)
    TOTAL_PUS_PER_LGA = {
        "osogbo": 218,"olorunda": 210,"egbedore": 72,"ede north": 88,"ede south": 50,
        "ejigbo": 120,"ife central": 148,"ife east": 118,"ife north": 74,"ife south": 84,
        "ifedayo": 50,"ifelodun": 136,"ila": 90,"ilesa east": 98,"ilesa west": 100,
        "irepodun": 130,"irewole": 100,"isokan": 72,"iwo": 200,"obokun": 88,
        "odo-otin": 110,"ola-oluwa": 64,"oriade": 100,"orolu": 96,
        "atakumosa east": 76,"atakumosa west": 68,"ayedaade": 112,"ayedire": 68,
        "boluwaduro": 56,"boripe": 90
    }
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT LOWER(lg) as lga, COUNT(*) as submitted FROM field_submissions GROUP BY LOWER(lg)")
        rows = cur.fetchall(); cur.close(); conn.close()
        result = []
        for r in rows:
            lga = r["lga"] or ""
            submitted = r["submitted"]
            total = TOTAL_PUS_PER_LGA.get(lga.lower(), 100)
            pct = round((submitted / total) * 100, 1)
            result.append({"lga": lga.upper(), "submitted": submitted, "total": total, "pct": pct})
        result.sort(key=lambda x: x["pct"], reverse=True)
        return result
    except Exception as e:
        return []

@app.get("/api/swing_pus")
async def swing_pus(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT * FROM field_submissions")
        rows = cur.fetchall(); cur.close(); conn.close()
        PARTIES_LIST = ["ACCORD","AA","AAC","ADC","ADP","APGA","APC","APM","APP","BP","NNPP","PRP","YPP","ZLP"]
        swing = []
        for r in rows:
            try:
                v = json.loads(r.get("edited_votes_json") or r["votes_json"] or "{}") if isinstance(r.get("edited_votes_json") or r.get("votes_json"), str) else (r.get("edited_votes_json") or r.get("votes_json") or {})
            except Exception:
                v = {}
            accord = v.get("ACCORD", 0)
            rivals = {p: v.get(p, 0) for p in PARTIES_LIST if p != "ACCORD"}
            if not rivals: continue
            top_rival = max(rivals, key=rivals.get)
            margin = accord - rivals[top_rival]
            if abs(margin) <= 15:
                swing.append({"pu_name": r["location"], "lga": r["lg"], "ward": r["ward"],
                               "accord": accord, "rival": top_rival,
                               "rival_votes": rivals[top_rival], "margin": margin})
        swing.sort(key=lambda x: abs(x["margin"]))
        return swing
    except Exception as e:
        return []

@app.get("/api/integrity_flags")
async def integrity_flags(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT * FROM field_submissions")
        rows = cur.fetchall(); cur.close(); conn.close()
        flags = []
        for r in rows:
            issues = []
            total_cast = r.get("total_cast") or 0
            accredited = r.get("total_accredited") or 0
            ec8e = r.get("ec8e_image")
            if accredited > 0 and total_cast > accredited:
                issues.append("Overvoting: votes exceed accredited")
            if not ec8e:
                issues.append("No EC8E image uploaded")
            if total_cast == 0:
                issues.append("Zero total votes recorded")
            if issues:
                flags.append({"pu_name": r["location"], "lga": r["lg"], "ward": r["ward"],
                               "issues": issues,
                               "severity": "high" if any("Overvoting" in i for i in issues) else "medium"})
        return flags
    except Exception as e:
        return []

@app.get("/api/collation_timeline")
async def collation_timeline(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("SELECT location, lg, timestamp FROM field_submissions ORDER BY timestamp ASC")
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"pu_name": r["location"], "lga": r["lg"], "timestamp": str(r["timestamp"])} for r in rows]
    except Exception as e:
        return []

@app.get("/api/agent_leaderboard")
async def agent_leaderboard(request: Request):
    _require_dashboard(request)
    try:
        conn = get_pg(); cur = conn.cursor()
        cur.execute("""SELECT officer_id, COUNT(*) as submissions, MAX(timestamp) as last_submission
                       FROM field_submissions GROUP BY officer_id ORDER BY submissions DESC LIMIT 20""")
        rows = cur.fetchall(); cur.close(); conn.close()
        return [dict(r) for r in rows]
    except Exception as e:
        return []

# ─────────────────────────────────────────────────────────────────────────────
# ── INCIDENT ENDPOINTS ────────────────────────────────────────────────────────

@app.get("/api/incidents")
async def get_incidents(request: Request):
    _require_dashboard(request)
    try:
        with get_pg() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM incidents ORDER BY timestamp DESC")
                rows = cur.fetchall()
                return [dict(r) for r in rows]
    except Exception as e:
        return []

@app.post("/submit-incident")
async def submit_incident(
    data: str = Form(...),
    evidence: UploadFile = File(None)
):
    try:
        payload = json.loads(data)
        evidence_url = None

        if evidence and evidence.filename:
            safe_pu = str(payload.get("pu_code", "unk")).replace("/", "_").replace(" ", "_")
            public_id = f"incidents/{safe_pu}_{uuid.uuid4().hex[:8]}"
            try:
                img_bytes = await _safe_read_image(evidence)
                upload_result = cloudinary.uploader.upload(
                    img_bytes,
                    public_id=public_id,
                    resource_type="auto",
                    overwrite=True
                )
                evidence_url = upload_result["secure_url"]
                logger.info(f"Evidence uploaded to Cloudinary: {evidence_url}")
            except Exception as cloud_err:
                logger.error(f"Cloudinary upload failed for incident: {cloud_err}")

        with get_pg() as conn:
            with conn.cursor() as cur:
                cur.execute("""INSERT INTO incidents (
                    officer_id, pu_code, ward, ward_code, lg, state, location,
                    incident_type, severity, description, evidence_url,
                    lat, lon, timestamp, status
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""", (
                    payload.get("officer_id"),
                    payload.get("pu_code"),
                    payload.get("ward"),
                    payload.get("ward_code"),
                    payload.get("lg"),
                    payload.get("state"),
                    payload.get("location"),
                    payload.get("incident_type"),
                    payload.get("severity"),
                    payload.get("description"),
                    evidence_url,
                    payload.get("lat"),
                    payload.get("lon"),
                    datetime.now().isoformat(),
                    "open"
                ))
            conn.commit()

        alert_payload = {**payload, "timestamp": datetime.now().strftime("%d %b %Y %H:%M")}
        threading.Thread(target=send_incident_alert, args=(alert_payload,)).start()

        return {"status": "success", "message": "Incident reported successfully"}
    except Exception as e:
        logger.error(f"Incident submission error: {e}", exc_info=True)
        return {"status": "error", "message": str(e)}

# ─────────────────────────────────────────────────────────────────────────────
# ── Dashboard login page ──────────────────────────────────────────────────────
DASHBOARD_LOGIN_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TKC — Dashboard Access</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;900&display=swap" rel="stylesheet">
    <style>
        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            font-family: 'Inter', sans-serif;
            background: #020c06;
            color: #fff;
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
            overflow: hidden;
        }
        body::before {
            content: '';
            position: fixed;
            inset: 0;
            background:
                radial-gradient(ellipse 120% 80% at 50% -10%, rgba(0,135,81,0.18) 0%, transparent 60%),
                linear-gradient(160deg, #020c06 0%, #041508 40%, #020c06 100%);
            z-index: 0;
        }
        .grid {
            position: fixed; inset: 0; z-index: 0;
            background-image:
                linear-gradient(rgba(0,135,81,0.05) 1px, transparent 1px),
                linear-gradient(90deg, rgba(0,135,81,0.05) 1px, transparent 1px);
            background-size: 60px 60px;
            pointer-events: none;
        }
        .card {
            position: relative; z-index: 1;
            background: rgba(255,255,255,0.04);
            border: 1px solid rgba(0,135,81,0.25);
            border-radius: 20px;
            padding: 48px 40px;
            width: 100%;
            max-width: 420px;
            text-align: center;
            box-shadow: 0 0 60px rgba(0,135,81,0.1);
        }
        .lock-icon {
            width: 64px; height: 64px;
            border-radius: 50%;
            background: rgba(0,135,81,0.12);
            border: 1px solid rgba(0,135,81,0.3);
            display: flex; align-items: center; justify-content: center;
            font-size: 1.6rem;
            margin: 0 auto 24px;
            box-shadow: 0 0 30px rgba(0,135,81,0.2);
        }
        .tag {
            display: inline-block;
            font-size: 0.6rem; font-weight: 700;
            letter-spacing: 0.18em; text-transform: uppercase;
            color: #ffc107;
            background: rgba(255,193,7,0.1);
            border: 1px solid rgba(255,193,7,0.2);
            border-radius: 20px;
            padding: 3px 12px;
            margin-bottom: 14px;
        }
        h1 { font-size: 1.4rem; font-weight: 900; margin-bottom: 8px; }
        p { font-size: 0.8rem; color: rgba(255,255,255,0.4); margin-bottom: 32px; line-height: 1.5; }
        .input-wrap { position: relative; margin-bottom: 12px; }
        input[type="password"] {
            width: 100%;
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(0,135,81,0.3);
            border-radius: 12px;
            color: #fff;
            font-size: 1rem;
            padding: 14px 48px 14px 18px;
            outline: none;
            transition: border-color 0.2s;
            letter-spacing: 0.12em;
        }
        input[type="password"]:focus { border-color: rgba(0,135,81,0.7); }
        .toggle-eye {
            position: absolute; right: 14px; top: 50%;
            transform: translateY(-50%);
            background: none; border: none; color: rgba(255,255,255,0.3);
            cursor: pointer; font-size: 1rem; padding: 0;
        }
        .toggle-eye:hover { color: rgba(255,255,255,0.7); }
        button.submit {
            width: 100%;
            background: linear-gradient(135deg, #008751, #00b368);
            border: none; border-radius: 12px;
            color: #fff; font-size: 0.95rem; font-weight: 700;
            padding: 14px;
            cursor: pointer;
            transition: opacity 0.2s, transform 0.1s;
            margin-bottom: 16px;
        }
        button.submit:hover { opacity: 0.9; }
        button.submit:active { transform: scale(0.98); }
        button.submit:disabled { opacity: 0.5; cursor: not-allowed; }
        .error {
            background: rgba(220,53,69,0.15);
            border: 1px solid rgba(220,53,69,0.3);
            border-radius: 8px;
            color: #ff6b6b;
            font-size: 0.78rem;
            padding: 10px 14px;
            margin-bottom: 12px;
            display: none;
        }
        .back { font-size: 0.72rem; color: rgba(255,255,255,0.25); text-decoration: none; }
        .back:hover { color: rgba(255,255,255,0.5); }
    </style>
</head>
<body>
<div class="grid"></div>
<div class="card">
    <div class="lock-icon">🔐</div>
    <div class="tag">Restricted Access</div>
    <h1>Dashboard Access</h1>
    <p>This dashboard is restricted to authorised command team members only. Enter your access key to continue.</p>
    <div class="error" id="err"></div>
    <div class="input-wrap">
        <input type="password" id="key" placeholder="Enter access key" autocomplete="off" onkeydown="if(event.key==='Enter')verify()">
        <button class="toggle-eye" onclick="toggleEye()" id="eyeBtn" title="Show/hide">👁</button>
    </div>
    <button class="submit" id="btn" onclick="verify()">Unlock Dashboard →</button>
    <a href="/" class="back">← Back to home</a>
</div>
<script>
    function toggleEye() {
        const inp = document.getElementById('key');
        inp.type = inp.type === 'password' ? 'text' : 'password';
    }
    async function verify() {
        const key = document.getElementById('key').value.trim();
        const btn = document.getElementById('btn');
        const err = document.getElementById('err');
        if (!key) return;
        btn.disabled = true; btn.textContent = 'Verifying...';
        err.style.display = 'none';
        try {
            const res = await fetch('/api/verify-dashboard', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({ key })
            });
            if (res.ok) {
                window.location.href = '/dashboard';
            } else {
                err.textContent = 'Invalid access key. Contact your system administrator.';
                err.style.display = 'block';
                btn.disabled = false; btn.textContent = 'Unlock Dashboard →';
            }
        } catch(e) {
            err.textContent = 'Connection error. Try again.';
            err.style.display = 'block';
            btn.disabled = false; btn.textContent = 'Unlock Dashboard →';
        }
    }
</script>
</body>
</html>
"""
# ─────────────────────────────────────────────────────────────────────────────



# ── Demo: clear all submissions ────────────────────────────────────────────────
@app.post("/api/admin/clear-submissions")
async def clear_submissions(request: Request):
    auth = request.headers.get("Authorization", "")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(),
        _DASHBOARD_KEY_HASH
    )):
        raise HTTPException(status_code=403, detail="Not authorised")
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as c FROM field_submissions")
            row   = cur.fetchone()
            count = row["c"] if row else 0
            cur.execute("TRUNCATE TABLE field_submissions RESTART IDENTITY")
    return {"status": "ok", "message": f"Cleared {count} submissions", "deleted": count}


# ── Demo: create submission directly (no OTP) ───────────────────────────────────
@app.post("/api/admin/create-submission")
async def create_submission(request: Request):
    auth = request.headers.get("Authorization", "")
    if not (auth.startswith("Bearer ") and secrets.compare_digest(
        hashlib.sha256(auth.split(" ", 1)[1].strip().encode()).hexdigest(),
        _DASHBOARD_KEY_HASH
    )):
        raise HTTPException(status_code=403, detail="Not authorised")
    body = await request.json()
    now  = datetime.now().isoformat()
    with get_pg() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO field_submissions
                (officer_id, state, lg, ward, ward_code, pu_code, location,
                 reg_voters, total_accredited, valid_votes, rejected_votes,
                 total_cast, lat, lon, timestamp, votes_json)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (pu_code) DO UPDATE SET
                    officer_id = EXCLUDED.officer_id,
                    state = EXCLUDED.state,
                    lg = EXCLUDED.lg,
                    ward = EXCLUDED.ward,
                    ward_code = EXCLUDED.ward_code,
                    location = EXCLUDED.location,
                    reg_voters = EXCLUDED.reg_voters,
                    total_accredited = EXCLUDED.total_accredited,
                    valid_votes = EXCLUDED.valid_votes,
                    rejected_votes = EXCLUDED.rejected_votes,
                    total_cast = EXCLUDED.total_cast,
                    lat = EXCLUDED.lat,
                    lon = EXCLUDED.lon,
                    timestamp = EXCLUDED.timestamp,
                    votes_json = EXCLUDED.votes_json
            """, (
                body.get("officer_id", "DEMO"),
                (body.get("state", "osun")).lower(),
                body.get("lg", ""),
                body.get("ward", ""),
                body.get("ward_code", ""),
                body.get("pu_code", ""),
                body.get("location", ""),
                body.get("reg_voters", 0),
                body.get("total_accredited", 0),
                body.get("valid_votes", 0),
                body.get("rejected_votes", 0),
                body.get("total_cast", 0),
                body.get("lat", 0.0),
                body.get("lon", 0.0),
                now,
                json.dumps(body.get("votes", {})),
            ))
    return {"status": "ok", "message": "Submission created"}


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard_page(request: Request):
    from fastapi.responses import HTMLResponse as _HR
    token = request.cookies.get("ds_session")
    if not _is_valid_token(token):
        return _HR(content=DASHBOARD_LOGIN_HTML)
    return _HR(content=DASHBOARD_HTML)

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>TKC — Election Observation Room · Osun 2026</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.1/font/bootstrap-icons.css" rel="stylesheet">
    <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;900&display=swap" rel="stylesheet">
    <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
    <script src="https://unpkg.com/leaflet.heat@0.2.0/dist/leaflet-heat.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.0.0"></script>
    <style>
    :root {
        --gold:   #F5A623;
        --apc:    #1A56DB;
        --adc:    #00875A;
        --pdp:    #E02424;
        --bg:     #0D0F14;
        --s1:     #161B26;
        --s2:     #1E2433;
        --s3:     #252D3D;
        --border: rgba(255,255,255,0.07);
        --text:   #F0F2F5;
        --muted:  #6B7280;
        --nav-h:  62px;
        --kpi-h:  82px;
    }
    *,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}
    html,body{height:100%;background:var(--bg);color:var(--text);
        font-family:'Inter','Segoe UI',sans-serif;overflow:hidden;}

    /* ── NAVBAR ── */
    .navbar-custom{
        height:var(--nav-h);
        background:linear-gradient(180deg,rgba(22,27,38,0.98) 0%,rgba(13,15,20,0.98) 100%);
        border-bottom:1px solid rgba(245,166,35,0.25);
        backdrop-filter:blur(16px);
        display:flex;align-items:center;justify-content:space-between;
        padding:0 20px;gap:12px;flex-shrink:0;position:relative;z-index:10;
    }
    .brand{display:flex;align-items:center;gap:10px;}
    .live-dot{width:8px;height:8px;background:#22c55e;border-radius:50%;flex-shrink:0;
        box-shadow:0 0 0 0 rgba(34,197,94,0.5);animation:pulse-dot 1.5s infinite;}
    @keyframes pulse-dot{0%{box-shadow:0 0 0 0 rgba(34,197,94,0.5);}
        70%{box-shadow:0 0 0 8px rgba(34,197,94,0);}
        100%{box-shadow:0 0 0 0 rgba(34,197,94,0);}}
    .nav-divider{width:1px;height:32px;background:var(--border);flex-shrink:0;}
    .brand-name{font-weight:900;font-size:1rem;color:var(--gold);letter-spacing:1.5px;}
    .brand-tag{font-size:0.55rem;color:var(--muted);letter-spacing:2px;text-transform:uppercase;margin-top:1px;}

    .party-pills{display:flex;gap:8px;}
    .pill{background:var(--s2);border:1px solid var(--border);border-radius:10px;
        padding:5px 12px 5px 8px;display:flex;align-items:center;gap:8px;
        min-width:115px;transition:border-color 0.2s;}
    .pill-accord{border-bottom:2px solid var(--gold);}
    .pill-apc   {border-bottom:2px solid var(--apc);}
    .pill-adc   {border-bottom:2px solid var(--adc);}
    .pill-logo{width:28px;height:28px;border-radius:6px;background:var(--s3);
        display:flex;align-items:center;justify-content:center;overflow:hidden;flex-shrink:0;}
    .pill-logo img{width:100%;height:100%;object-fit:contain;padding:2px;}
    .pill-info label{display:block;font-size:0.52rem;color:var(--muted);
        text-transform:uppercase;letter-spacing:0.5px;}
    .pill-info span{font-size:1rem;font-weight:800;line-height:1.1;}
    .pill-accord .pill-info span{color:var(--gold);}
    .pill-apc    .pill-info span{color:#60a5fa;}
    .pill-adc    .pill-info span{color:#34d399;}

    .nav-right{display:flex;align-items:center;gap:12px;flex-shrink:0;}
    .pu-badge{background:var(--s2);border:1px solid var(--border);border-radius:8px;
        padding:5px 12px;font-size:0.7rem;color:var(--muted);
        display:flex;align-items:center;gap:6px;}
    .pu-badge b{color:var(--text);font-weight:700;}
    .pu-badge i{color:#a855f7;font-size:0.75rem;}
    .live-clock{font-size:0.88rem;font-weight:700;color:var(--gold);
        font-variant-numeric:tabular-nums;letter-spacing:1px;}
    .logout-btn{background:none;border:1px solid var(--border);color:var(--muted);
        border-radius:8px;padding:5px 10px;font-size:0.7rem;cursor:pointer;
        transition:all 0.2s;font-family:inherit;}
    .logout-btn:hover{border-color:var(--pdp);color:var(--pdp);}

    /* ── KPI STRIP ── */
    .kpi-strip{display:grid;grid-template-columns:repeat(6,1fr);
        gap:8px;padding:8px 16px;flex-shrink:0;}
    .kpi{background:var(--s1);border:1px solid var(--border);border-radius:10px;
        padding:12px 14px;position:relative;overflow:hidden;cursor:default;}
    .kpi::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;border-radius:2px 2px 0 0;}
    .kpi-accord::before{background:linear-gradient(90deg,var(--gold),#ffd666);}
    .kpi-margin::before{background:linear-gradient(90deg,var(--apc),#60a5fa);}
    .kpi-pus::before   {background:linear-gradient(90deg,#7c3aed,#a855f7);}
    .kpi-accred::before{background:linear-gradient(90deg,#0891b2,#22d3ee);}
    .kpi-inc::before   {background:linear-gradient(90deg,var(--pdp),#f87171);}
    .kpi-turn::before  {background:linear-gradient(90deg,var(--adc),#34d399);}
    .kpi-icon{font-size:1rem;margin-bottom:4px;display:block;opacity:0.9;}
    .kpi-accord .kpi-icon{color:var(--gold);}
    .kpi-margin .kpi-icon{color:#60a5fa;}
    .kpi-pus    .kpi-icon{color:#c084fc;}
    .kpi-accred .kpi-icon{color:#22d3ee;}
    .kpi-inc    .kpi-icon{color:#f87171;}
    .kpi-turn   .kpi-icon{color:#34d399;}
    .kpi-val{font-size:1.25rem;font-weight:900;display:block;line-height:1.1;
        transition:all 0.4s;letter-spacing:-0.5px;}
    .kpi-label{font-size:0.58rem;color:#9CA3AF;text-transform:uppercase;
        letter-spacing:0.5px;margin-top:4px;display:block;font-weight:500;}

    /* ── MAIN GRID ── */
    .main-content{
        display:grid;grid-template-columns:290px 1fr 280px;
        gap:8px;padding:0 8px 8px;
        height:calc(100vh - var(--nav-h) - var(--kpi-h));overflow:hidden;
    }
    .side-panel{background:var(--s1);border:1px solid var(--border);border-radius:12px;
        display:flex;flex-direction:column;overflow:hidden;min-height:0;}
    .panel-header{padding:9px 14px;font-size:0.63rem;font-weight:700;
        color:#9CA3AF;text-transform:uppercase;letter-spacing:1px;
        border-bottom:1px solid var(--border);flex-shrink:0;
        display:flex;align-items:center;justify-content:space-between;}
    .ph-badge{background:var(--s2);border:1px solid var(--border);
        border-radius:20px;padding:1px 8px;font-size:0.58rem;color:var(--gold);}

    /* Filter bar */
    .filter-row{padding:7px;display:flex;gap:5px;border-bottom:1px solid var(--border);flex-shrink:0;}
    .filter-row input,.filter-row select{flex:1;background:var(--s2);color:var(--text);
        border:1px solid var(--border);border-radius:8px;font-size:0.68rem;
        padding:5px 8px;font-family:inherit;outline:none;}
    .filter-row input:focus,.filter-row select:focus{border-color:rgba(245,166,35,0.4);}
    .filter-row input::placeholder{color:var(--muted);}

    /* PU Feed */
    .feed-container{flex:1;overflow-y:auto;padding:8px;min-height:0;}
    .feed-container::-webkit-scrollbar{width:3px;}
    .feed-container::-webkit-scrollbar-thumb{background:var(--s3);border-radius:3px;}

    .pu-card{background:var(--s2);border:1px solid var(--border);border-radius:10px;
        padding:9px 11px;margin-bottom:6px;cursor:pointer;
        transition:background 0.15s,border-color 0.15s,transform 0.1s;}
    .pu-card:hover{background:var(--s3);border-color:rgba(245,166,35,0.2);transform:translateX(2px);}
    .pu-card-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:3px;}
    .pu-card-name{font-size:0.72rem;font-weight:600;flex:1;min-width:0;
        overflow:hidden;text-overflow:ellipsis;white-space:nowrap;margin-right:6px;}
    .winner-badge{font-size:0.52rem;font-weight:700;padding:2px 7px;border-radius:20px;flex-shrink:0;}
    .wb-ACCORD{background:rgba(245,166,35,0.12);color:var(--gold);border:1px solid rgba(245,166,35,0.2);}
    .wb-APC   {background:rgba(26,86,219,0.12);color:#60a5fa;border:1px solid rgba(26,86,219,0.2);}
    .wb-ADC   {background:rgba(0,135,90,0.12);color:#34d399;border:1px solid rgba(0,135,90,0.2);}
    .wb-PDP   {background:rgba(224,36,36,0.12);color:#f87171;border:1px solid rgba(224,36,36,0.2);}
    .wb-other {background:rgba(107,114,128,0.12);color:var(--muted);border:1px solid var(--border);}
    .pu-card-meta{font-size:0.6rem;color:var(--muted);margin-bottom:5px;}
    .score-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:3px;}
    .score-item{background:var(--s1);border-radius:6px;padding:4px 5px;text-align:center;border:1px solid transparent;}
    .si-ACCORD{border-color:rgba(245,166,35,0.15);}
    .si-APC   {border-color:rgba(26,86,219,0.15);}
    .si-ADC   {border-color:rgba(0,135,90,0.15);}
    .score-label{font-size:0.5rem;color:var(--muted);display:block;margin-bottom:1px;}
    .score-val{font-size:0.72rem;font-weight:700;}
    .sv-ACCORD{color:var(--gold);}
    .sv-APC   {color:#60a5fa;}
    .sv-ADC   {color:#34d399;}
    .sv-PDP   {color:#f87171;}
    .sv-other {color:var(--muted);}
    .ec8e-badge{font-size:0.55rem;font-weight:700;background:rgba(245,166,35,0.15);
        color:var(--gold);padding:1px 5px;border-radius:4px;margin-left:4px;}

    /* Centre column */
    .centre-col{display:flex;flex-direction:column;gap:8px;min-height:0;overflow:hidden;}
    #map{height:45%;border-radius:12px;background:#111;
        border:1px solid rgba(245,166,35,0.15);flex-shrink:0;z-index:1;}
    .leaflet-heatmap-layer{opacity:0.55!important;}
    .leaflet-tile-pane{z-index:2;opacity:1!important;}
    .leaflet-overlay-pane{z-index:3;opacity:0.65!important;}
    .leaflet-marker-pane{z-index:4;}
    .leaflet-tooltip-pane{z-index:5;}
    .leaflet-popup-pane{z-index:6;}
    .chart-row{display:grid;grid-template-columns:1fr 1fr;gap:8px;flex:1;min-height:120px;}
    .chart-box{background:var(--s1);border:1px solid var(--border);border-radius:12px;
        padding:12px;display:flex;flex-direction:column;min-height:0;overflow:hidden;}
    .chart-title{font-size:0.62rem;font-weight:700;color:#9CA3AF;
        text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;flex-shrink:0;}
    .chart-box canvas{display:block;width:100%!important;flex:1;min-height:0;}

    /* Right panel */
    .margin-card{margin:10px;background:var(--s2);border:1px solid rgba(245,166,35,0.2);
        border-radius:10px;padding:12px;text-align:center;flex-shrink:0;}
    .margin-val{font-size:1.8rem;font-weight:900;color:var(--gold);display:block;line-height:1;}
    .margin-label{font-size:0.58rem;color:var(--muted);margin-top:3px;}
    .margin-rival{font-size:0.63rem;color:#34d399;margin-top:4px;
        display:flex;align-items:center;justify-content:center;gap:3px;}

    .proj-card{margin:0 10px 8px;background:var(--s2);
        border:1px solid rgba(0,135,90,0.2);border-radius:10px;
        padding:9px 12px;flex-shrink:0;display:flex;align-items:center;justify-content:space-between;}
    .proj-val{font-size:1.25rem;font-weight:900;color:#34d399;}
    .proj-label{font-size:0.57rem;color:var(--muted);}

    .lga-list{flex:1;overflow-y:auto;padding:8px;min-height:0;}
    .lga-list::-webkit-scrollbar{width:3px;}
    .lga-list::-webkit-scrollbar-thumb{background:var(--s3);}
    .lga-row{display:flex;align-items:center;gap:7px;margin-bottom:6px;}
    .lga-name{font-size:0.64rem;width:80px;flex-shrink:0;color:var(--text);
        white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
    .lga-bar-bg{flex:1;height:5px;background:var(--s3);border-radius:3px;overflow:hidden;}
    .lga-bar-fill{height:100%;border-radius:3px;
        background:linear-gradient(90deg,var(--adc),var(--gold));
        transition:width 0.6s ease;}
    .lga-pct{font-size:0.58rem;color:var(--muted);width:26px;text-align:right;flex-shrink:0;}

    .ai-box{margin:0 10px 10px;background:#000;border:1px solid rgba(34,197,94,0.15);
        border-radius:10px;padding:10px;font-family:'Courier New',monospace;
        font-size:0.62rem;color:#22c55e;flex:1;overflow-y:auto;min-height:0;
        line-height:1.7;white-space:pre-wrap;}

    /* EC8E viewer */
    #ec8eViewerPanel{background:var(--s2);border-top:1px solid var(--border);
        padding:8px;flex-shrink:0;display:none;}
    #ec8eViewerPanel.active{display:block;}

    /* Overlays */
    .ov-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.85);
        z-index:1000;align-items:center;justify-content:center;}
    .ov-overlay.active{display:flex;}
    .ov-inner{background:var(--s1);border:1px solid var(--border);border-radius:16px;
        padding:24px;max-width:600px;width:90%;max-height:80vh;overflow-y:auto;position:relative;}
    .ov-close{position:absolute;top:12px;right:14px;background:none;border:none;
        color:var(--muted);font-size:1.2rem;cursor:pointer;}
    .ov-close:hover{color:var(--text);}
    </style>
</head>
<body>

<!-- NAVBAR -->
<nav class="navbar-custom">
    <div class="brand">
        <div class="live-dot"></div>
        <div class="nav-divider"></div>
        <div>
            <div class="brand-name">TKC · OSUN 2026</div>
            <div class="brand-tag">Situation Room — Live Command</div>
        </div>
    </div>

    <div class="party-pills">
        <div class="pill pill-accord">
            <div class="pill-logo">
                <img src="/static/logos/ACCORD.png" alt="ACCORD"
                     onerror="this.style.display='none'">
            </div>
            <div class="pill-info">
                <label>ACCORD</label>
                <span id="nav-ACCORD">0</span>
            </div>
        </div>
        <div class="pill pill-apc">
            <div class="pill-logo">
                <img src="/static/logos/APC.png" alt="APC"
                     onerror="this.style.display='none'">
            </div>
            <div class="pill-info">
                <label>APC</label>
                <span id="nav-APC">0</span>
            </div>
        </div>
        <div class="pill pill-adc">
            <div class="pill-logo">
                <img src="/static/logos/ADC.png" alt="ADC"
                     onerror="this.style.display='none'">
            </div>
            <div class="pill-info">
                <label>ADC</label>
                <span id="nav-ADC">0</span>
            </div>
        </div>
        <!-- Hidden spans for other parties (needed by existing JS) -->
        <span id="nav-PDP" style="display:none">0</span>
        <span id="nav-LP"  style="display:none">0</span>
        <span id="nav-NNPP" style="display:none">0</span>
        <span id="nav-APGA" style="display:none">0</span>
        <span id="nav-PRP"  style="display:none">0</span>
        <span id="nav-SDP"  style="display:none">0</span>
        <span id="nav-AA"   style="display:none">0</span>
        <span id="nav-AAC"  style="display:none">0</span>
        <span id="nav-ADP"  style="display:none">0</span>
        <span id="nav-APM"  style="display:none">0</span>
        <span id="nav-APP"  style="display:none">0</span>
        <span id="nav-BP"   style="display:none">0</span>
        <span id="nav-YPP"  style="display:none">0</span>
        <span id="nav-ZLP"  style="display:none">0</span>
    </div>

    <div class="nav-right">
        <div class="pu-badge">
            <i class="bi bi-geo-alt-fill"></i>
            <b id="pu-count">0</b>&nbsp;PUs Reported
        </div>
        <div class="live-clock" id="liveClock">--:--:--</div>
        <button class="logout-btn" onclick="logoutDash()">
            <i class="bi bi-box-arrow-right"></i> Logout
        </button>
    </div>
</nav>

<!-- KPI STRIP -->
<div class="kpi-strip">
    <div class="kpi kpi-accord">
        <i class="bi bi-check-circle-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-accord-val">0</span>
        <span class="kpi-label">ACCORD Total</span>
    </div>
    <div class="kpi kpi-margin">
        <i class="bi bi-graph-up-arrow kpi-icon"></i>
        <span class="kpi-val" id="kpi-margin-val">0</span>
        <span class="kpi-label">Lead Margin</span>
    </div>
    <div class="kpi kpi-pus">
        <i class="bi bi-geo-alt-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-pus-val">0</span>
        <span class="kpi-label">PUs Reported</span>
    </div>
    <div class="kpi kpi-accred">
        <i class="bi bi-people-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-accred-val">0</span>
        <span class="kpi-label">Accredited</span>
    </div>
    <div class="kpi kpi-inc">
        <i class="bi bi-exclamation-triangle-fill kpi-icon"></i>
        <span class="kpi-val" id="kpi-inc-val">0</span>
        <span class="kpi-label">Incidents</span>
    </div>
    <div class="kpi kpi-turn">
        <i class="bi bi-percent kpi-icon"></i>
        <span class="kpi-val" id="kpi-turn-val">0%</span>
        <span class="kpi-label">Avg Turnout</span>
    </div>
</div>

<!-- MAIN GRID -->
<div class="main-content">

    <!-- LEFT: PU FEED -->
    <div class="side-panel">
        <div class="panel-header">
            PU Results Feed
            <span class="ph-badge" id="feed-count">0</span>
        </div>
        <div class="filter-row">
            <input type="text" id="puSearch" placeholder="🔍  Search PU..." oninput="applyFilters()">
        </div>
        <div class="filter-row" style="border-top:none;padding-top:0;">
            <select id="fState" onchange="loadFilters()">
                <option value="">All States</option>
                <option value="osun" selected>Osun</option>
            </select>
            <select id="fLGA" onchange="updateWards()">
                <option value="">All LGAs</option>
            </select>
            <select id="fWard" onchange="applyFilters()">
                <option value="">All Wards</option>
            </select>
        </div>
        <div class="feed-container" id="feedList"></div>
        <div id="ec8eViewerPanel">
            <div id="ec8eContent" style="font-size:0.7rem;color:var(--muted);">
                Click a PU card to view EC8E form image
            </div>
        </div>
    </div>

    <!-- CENTRE: MAP + CHARTS -->
    <div class="centre-col">
        <div id="map"></div>
        <div class="chart-row">
            <div class="chart-box">
                <div class="chart-title">🗳 Vote Share %</div>
                <canvas id="pieChart"></canvas>
            </div>
            <div class="chart-box">
                <div class="chart-title">📊 Party Vote Count</div>
                <canvas id="barChart"></canvas>
            </div>
        </div>
    </div>

    <!-- RIGHT: ANALYTICS -->
    <div class="side-panel">
        <div class="panel-header">Analytics & Projection</div>

        <div class="margin-card">
            <span class="margin-val" id="marginVal">0</span>
            <div class="margin-label" id="marginLead">ACCORD LEAD</div>
            <div class="margin-rival">
                <i class="bi bi-arrow-up-short"></i>
                <span id="projectionNote">Awaiting data...</span>
            </div>
        </div>

        <div class="proj-card">
            <div>
                <div class="proj-val" id="projectionVal">—</div>
                <div class="proj-label">Projected ACCORD Total</div>
            </div>
            <i class="bi bi-lightning-fill" style="color:var(--gold);font-size:1.1rem;"></i>
        </div>

        <div class="panel-header" style="margin-top:2px;">
            LGA Completion
            <span class="ph-badge" id="lga-count">0 LGAs</span>
        </div>
        <div class="lga-list" id="lgaCompletionList"></div>

        <div class="panel-header">AI Insight</div>
        <div class="ai-box" id="ai_box">Awaiting data...</div>
    </div>
</div>

<!-- OVERLAYS (kept from original) -->
<div id="ov-lga" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-lga')">✕</button>
        <h5 style="color:var(--gold);margin-bottom:12px;">LGA COMPLETION</h5>
        <div id="ov-lga-inner" style="max-height:60vh;overflow-y:auto;"></div>
    </div>
</div>
<div id="ov-swing" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-swing')">✕</button>
        <h5 style="color:#ff4444;margin-bottom:12px;">SWING POLLING UNITS</h5>
        <div id="ov-swing-inner" style="max-height:60vh;overflow-y:auto;"></div>
    </div>
</div>
<div id="ov-flags" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-flags')">✕</button>
        <h5 style="color:#ff6600;margin-bottom:12px;">RESULT INTEGRITY FLAGS</h5>
        <div id="ov-flags-inner" style="max-height:60vh;overflow-y:auto;"></div>
    </div>
</div>
<div id="ov-proj" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-proj')">✕</button>
        <h5 style="color:#00ff88;margin-bottom:12px;">PROJECTED TALLY + AGENTS</h5>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;">
            <div>
                <div style="font-size:0.75rem;color:#aaa;margin-bottom:8px;">PROJECTED ACCORD TOTAL</div>
                <div id="ov-projVal" style="font-size:3rem;font-weight:900;color:#00ff88;">--</div>
                <div id="ov-projNote" style="font-size:0.75rem;color:#555;margin-top:4px;"></div>
            </div>
            <div>
                <div style="font-size:0.75rem;color:#ffc107;font-weight:bold;margin-bottom:8px;">AGENT LEADERBOARD</div>
                <div id="ov-agentList"></div>
            </div>
        </div>
    </div>
</div>
<div id="ov-timeline" class="ov-overlay">
    <div class="ov-inner">
        <button class="ov-close" onclick="closeOverlay('ov-timeline')">✕</button>
        <h5 style="color:var(--gold);margin-bottom:12px;">COLLATION TIMELINE</h5>
        <div style="position:relative;height:320px;"><canvas id="ov-timelineChart"></canvas></div>
        <div id="ov-timeline-list" style="max-height:200px;overflow-y:auto;margin-top:12px;"></div>
    </div>
</div>

<script>
    const PARTIES = ['ACCORD','AA','AAC','ADC','ADP','APGA','APC','APM','APP','BP','NNPP','PRP','YPP','ZLP'];
    const BIG3    = ['ACCORD','APC','ADC'];
    const COLORS  = { ACCORD:'#F5A623', APC:'#1A56DB', ADC:'#00875A', PDP:'#E02424' };

    let map, markers = [], heatLayer = null, allIncidents = [];
    let globalData = [], globalTotals = {};
    let pie = null, bar = null;

    // ── Live clock ──
    function tickClock() {
        const n = new Date();
        const el = document.getElementById('liveClock');
        if (el) el.textContent = n.toLocaleTimeString('en-NG',
            {hour:'2-digit',minute:'2-digit',second:'2-digit'});
    }
    setInterval(tickClock, 1000); tickClock();

    // ── Map init ──
    function init() {
        // Locked to Osun State
                map = L.map('map', {zoomControl:false}).setView([7.56, 4.52], 9);
        L.tileLayer('https://{s}.basemaps.cartocdn.com/dark_all/{z}/{x}/{y}{r}.png')
         .addTo(map);

        refreshData();
        loadInsights();
    }

    // ── Data refresh ──
    async function refreshData() {
        try {
            const res = await fetch(window.location.origin + '/submissions', {credentials:'include'});
            globalData = await res.json();
            globalData = globalData.map(x => ({
                ...x,
                state: (x.state||'').toLowerCase(),
                lga:   (x.lga||'').toLowerCase(),
                ward:  (x.ward||'').toLowerCase()
            }));
            applyFilters();
        } catch(e) { console.error('Data refresh error', e); }
    }

    // ── Filters ──
    async function loadFilters() {
        try {
            const s = document.getElementById('fState').value;
            if (!s) return;
            const res = await fetch('/api/lgas/' + s, {credentials:'include'});
            const lgas = await res.json();
            const sel = document.getElementById('fLGA');
            sel.innerHTML = '<option value="">All LGAs</option>';
            lgas.forEach(l => {
                const o = document.createElement('option');
                o.value = l; o.textContent = l.charAt(0).toUpperCase() + l.slice(1);
                sel.appendChild(o);
            });
            applyFilters();
        } catch(e) {}
    }

    async function updateWards() {
        const s = document.getElementById('fState').value;
        const l = document.getElementById('fLGA').value;
        const sel = document.getElementById('fWard');
        sel.innerHTML = '<option value="">All Wards</option>';
        if (s && l) {
            try {
                const res = await fetch(`/api/wards/${s}/${l}`, {credentials:'include'});
                const wards = await res.json();
                wards.forEach(w => {
                    const o = document.createElement('option');
                    o.value = w.name || w; o.textContent = w.name || w;
                    sel.appendChild(o);
                });
            } catch(e) {}
        }
        applyFilters();
    }

    function applyFilters() {
        const s = document.getElementById('fState').value.toLowerCase();
        const l = document.getElementById('fLGA').value.toLowerCase();
        const w = document.getElementById('fWard').value.toLowerCase();
        let filtered = globalData;
        if(s) filtered = filtered.filter(x => x.state === s);
        if(l) filtered = filtered.filter(x => x.lga   === l);
        if(w) filtered = filtered.filter(x => x.ward  === w);
        updateUI(filtered);
    }

    // ── Main UI update — wires ALL metrics ──
    function updateUI(data) {
        const t = {};
        PARTIES.forEach(p => t[p] = 0);
        let totalAccred = 0, totalReg = 0;

        const list = document.getElementById('feedList');
        list.innerHTML = '';
        markers.forEach(m => map.removeLayer(m));
        markers = [];
        const heatPoints = [];

        const searchTerm = (document.getElementById('puSearch').value || '').toLowerCase();

        data.forEach(d => {
            PARTIES.forEach(p => { t[p] += (d['votes_party_'+p] || 0); });
            totalAccred += (d.total_accredited || 0);
            totalReg    += (d.reg_voters       || 0);

            if (searchTerm && !(d.pu_name||'').toLowerCase().includes(searchTerm)) return;

            // Determine winner
            const partyVotes = BIG3.map(p => ({p, v: d['votes_party_'+p]||0}));
            partyVotes.sort((a,b) => b.v - a.v);
            const winner = partyVotes[0].p;
            const wbClass = 'wb-' + (COLORS[winner] ? winner : 'other');

            const ec8eBadge = d.ec8e_image
                ? `<span class="ec8e-badge">📷 EC8E</span>` : '';

            const card = document.createElement('div');
            card.className = 'pu-card';
            card.innerHTML = `
                <div class="pu-card-head">
                    <span class="pu-card-name">${d.pu_name||'—'}${ec8eBadge}</span>
                    <span class="winner-badge ${wbClass}">${winner} ✓</span>
                </div>
                <div class="pu-card-meta">${(d.lga||'').toUpperCase()} · ${d.ward||''}</div>
                <div class="score-grid">
                    <div class="score-item si-ACCORD">
                        <span class="score-label">ACCORD</span>
                        <span class="score-val sv-ACCORD">${(d.votes_party_ACCORD||0).toLocaleString()}</span>
                    </div>
                    <div class="score-item si-APC">
                        <span class="score-label">APC</span>
                        <span class="score-val sv-APC">${(d.votes_party_APC||0).toLocaleString()}</span>
                    </div>
                    <div class="score-item si-ADC">
                        <span class="score-label">ADC</span>
                        <span class="score-val sv-ADC">${(d.votes_party_ADC||0).toLocaleString()}</span>
                    </div>
                    <div class="score-item">
                        <span class="score-label">PDP</span>
                        <span class="score-val sv-PDP">${(d.votes_party_PDP||0).toLocaleString()}</span>
                    </div>
                </div>`;
            card.onclick = () => {
                if(d.latitude) map.setView([d.latitude, d.longitude], 14);
                showEc8e(d.ec8e_image, d.pu_name);
            };
            list.appendChild(card);

            // Map marker — coloured by winner
            if (d.latitude) {
                const col = COLORS[winner] || '#ffc107';
                const m = L.circleMarker([d.latitude, d.longitude], {
                    radius:7, color:col, fillColor:col, fillOpacity:0.85, weight:2
                }).addTo(map);
                m.bindPopup(`<b>${d.pu_name}</b><br>
                    ACCORD: <b style="color:#F5A623">${d.votes_party_ACCORD||0}</b>
                    &nbsp; APC: ${d.votes_party_APC||0}
                    &nbsp; ADC: <span style="color:#34d399">${d.votes_party_ADC||0}</span>`);
                markers.push(m);
                heatPoints.push([d.latitude, d.longitude,
                    Math.min(1, (d.votes_party_ACCORD||0) / 400)]);
            }
        });

        // Heatmap — subtle overlay, tiles stay visible
        if (heatLayer) { map.removeLayer(heatLayer); heatLayer = null; }
        if (typeof L.heatLayer !== 'undefined' && heatPoints.length > 0) {
            heatLayer = L.heatLayer(heatPoints, {
                radius:35, blur:25, maxZoom:14, minOpacity:0.4,
                gradient:{
                    0.0:'#00000000',
                    0.3:'#006400',
                    0.5:'#ffa500',
                    0.7:'#ff4500',
                    1.0:'#ffffff'
                }
            }).addTo(map);
            // Bring markers on top of heatmap
            markers.forEach(m => { try{ m.bringToFront(); }catch(e){} });
        }

        // ── Update ALL nav party spans ──
        PARTIES.forEach(p => {
            const el = document.getElementById('nav-'+p);
            if(el) el.innerText = (t[p]||0).toLocaleString();
        });

        // ── Update KPI cards ──
        const rivals = {};
        PARTIES.filter(p => p !== 'ACCORD').forEach(p => rivals[p] = t[p]);
        const topRival = Object.keys(rivals).reduce((a,b) => rivals[a]>rivals[b]?a:b, 'APC');
        const margin   = t.ACCORD - rivals[topRival];
        const turnout  = totalReg > 0 ? ((totalAccred/totalReg)*100).toFixed(1) : '0.0';

        const set = (id, val) => { const el=document.getElementById(id); if(el) el.innerText=val; };
        set('kpi-accord-val', (t.ACCORD||0).toLocaleString());
        set('kpi-margin-val', (margin>=0?'+':'')+margin.toLocaleString());
        set('kpi-pus-val',    data.length.toLocaleString());
        set('kpi-accred-val', totalAccred.toLocaleString());
        set('kpi-turn-val',   turnout+'%');
        set('pu-count',       data.length.toLocaleString());
        set('feed-count',     data.length.toLocaleString());

        // Margin card
        set('marginVal',  Math.abs(margin).toLocaleString());
        set('marginLead', margin>=0 ? `ACCORD LEAD OVER ${topRival}` : `TRAILING ${topRival}`);
        const mEl = document.getElementById('marginVal');
        if(mEl) mEl.style.color = margin>=0 ? '#F5A623' : '#f87171';

        globalTotals = {...t};
        updateProjection(globalTotals, data.length);
        updateCharts(t);
        runAI(t);
        loadInsights();
    }

    // ── Charts ──
    function updateCharts(t) {
        const labels = BIG3;
        const vals   = labels.map(p => t[p]||0);
        const colors = ['#F5A623','#1A56DB','#00875A'];
        const total  = vals.reduce((a,b)=>a+b,0);

        if(pie) pie.destroy();
        pie = new Chart(document.getElementById('pieChart'), {
            type:'doughnut',
            data:{labels, datasets:[{data:vals, backgroundColor:colors,
                borderWidth:2, borderColor:'#0D0F14', hoverOffset:6}]},
            options:{
                responsive:true, maintainAspectRatio:false,
                plugins:{
                    legend:{
                        position:'bottom',
                        labels:{color:'#9CA3AF',font:{size:9},
                            boxWidth:10,padding:10,usePointStyle:true}
                    },
                    datalabels:{
                        color:'#fff',
                        font:{weight:'900',size:11},
                        textShadowBlur:4,
                        textShadowColor:'rgba(0,0,0,0.8)',
                        formatter:(val,ctx)=>{
                            if(total===0||val===0) return '';
                            const pct=((val/total)*100).toFixed(1);
                            return pct>5 ? pct+'%' : '';
                        }
                    },
                    tooltip:{callbacks:{
                        label:ctx=>`${ctx.label}: ${ctx.parsed.toLocaleString()} votes (${total>0?((ctx.parsed/total)*100).toFixed(1):0}%)`
                    }}
                },
                cutout:'60%'
            }
        });

        if(bar) bar.destroy();
        bar = new Chart(document.getElementById('barChart'), {
            type:'bar',
            data:{labels, datasets:[{data:vals, backgroundColor:colors,
                borderRadius:4, borderSkipped:false}]},
            options:{
                responsive:true, maintainAspectRatio:false,
                indexAxis:'y',
                plugins:{
                    legend:{display:false},
                    datalabels:{color:'#fff',anchor:'end',align:'right',
                        font:{weight:'bold',size:9},
                        formatter:(val)=>val>0?val.toLocaleString():''}
                },
                scales:{
                    x:{beginAtZero:true, grid:{color:'rgba(255,255,255,0.04)'},
                       ticks:{color:'#6B7280',font:{size:9}}, border:{display:false}},
                    y:{grid:{display:false}, ticks:{color:'#F0F2F5',font:{size:10,weight:'600'}},
                       border:{display:false}}
                }
            }
        });
    }

    // ── Projection ──
    function updateProjection(totals, puCount) {
        const TOTAL_PUS = 500;
        if (puCount < 1) return;
        const proj = Math.round((totals.ACCORD / puCount) * TOTAL_PUS);
        const set = (id,v) => { const el=document.getElementById(id); if(el) el.innerText=v; };
        set('projectionVal',  '~' + proj.toLocaleString());
        set('ov-projVal',     '~' + proj.toLocaleString());
        set('projectionNote', `Based on ${puCount} of ~${TOTAL_PUS} PUs`);
        set('ov-projNote',    `Based on ${puCount} of ~${TOTAL_PUS} PUs`);
    }

    // ── AI Insight ──
    async function runAI(totals) {
        try {
            const s = document.getElementById('fState').value.toLowerCase();
            const l = document.getElementById('fLGA').value.toLowerCase();
            const payload = Object.assign({}, totals, {lg:l||'ALL', state:s||'Osun'});
            const res = await fetch(window.location.origin + '/api/ai_interpret', {
                method:'POST', headers:{'Content-Type':'application/json'},
                credentials:'include', body:JSON.stringify(payload)
            });
            const out = await res.json();
            const el = document.getElementById('ai_box');
            if(el) el.innerText = out.analysis || 'No insight available.';
        } catch(e) {}
    }

    // ── Insights (LGA completion etc) ──
    async function loadInsights() {
        loadLGACompletion();
    }

    async function loadLGACompletion() {
        try {
            const res = await fetch('/api/lga_completion', {credentials:'include'});
            const data = await res.json();
            const el = document.getElementById('lgaCompletionList');
            if(!el) return;
            el.innerHTML = '';
            const lgas = data.slice(0,15);
            lgas.forEach(row => {
                const pct = row.pct !== undefined ? row.pct
                    : Math.round((row.submitted||row.reported||0) / Math.max(row.total,1) * 100);
                const div = document.createElement('div');
                div.className = 'lga-row';
                div.innerHTML = `
                    <span class="lga-name">${row.lga||row.name||''}</span>
                    <div class="lga-bar-bg">
                        <div class="lga-bar-fill" style="width:${Math.min(pct,100)}%"></div>
                    </div>
                    <span class="lga-pct">${pct}%</span>`;
                el.appendChild(div);
            });
            const countEl = document.getElementById('lga-count');
            if(countEl) countEl.textContent =
                `${lgas.filter(r=>(r.submitted||r.reported||0)>0).length} / 30`;
        } catch(e) {}
    }

    // ── Incidents KPI ──
    async function loadIncidents() {
        try {
            const res = await fetch('/api/incidents', {credentials:'include'});
            const data = await res.json();
            const el = document.getElementById('kpi-inc-val');
            if(el) el.innerText = data.length;
        } catch(e) {}
    }

    // ── EC8E viewer ──
    function showEc8e(url, name) {
        const panel = document.getElementById('ec8eViewerPanel');
        const content = document.getElementById('ec8eContent');
        if(!panel || !content) return;
        if(url) {
            panel.classList.add('active');
            content.innerHTML = `<div style="font-size:0.65rem;color:var(--muted);margin-bottom:6px;">${name}</div>
                <img src="${url}" style="max-width:100%;max-height:180px;border-radius:8px;
                    border:1px solid var(--border);cursor:zoom-in;"
                    onclick="window.open('${url}','_blank')">`;
        } else {
            panel.classList.remove('active');
            content.innerHTML = 'No EC8E image for this PU.';
        }
    }

    // ── Overlays ──
    function openOverlay(id) {
        const el = document.getElementById(id);
        if(el) el.classList.add('active');
    }
    function closeOverlay(id) {
        const el = document.getElementById(id);
        if(el) el.classList.remove('active');
    }

    // ── Logout ──
    async function logoutDash() {
        await fetch('/api/logout-dashboard', {method:'POST'});
        window.location.href = '/dashboard';
    }

    // ── Boot ──
    document.addEventListener('DOMContentLoaded', init);
    setInterval(refreshData,   2000);
    setInterval(loadIncidents, 2000);
    setInterval(loadInsights,  2000);
</script>
</body>
</html>
"""
